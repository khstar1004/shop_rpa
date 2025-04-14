# Import refactored module structure while maintaining backward compatibility
import logging
import os
import re
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import queue
import threading

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from utils.caching import FileCache, cache_result
from utils.reporting import generate_primary_report, generate_secondary_report

from .data_models import MatchResult, ProcessingResult, Product
from .matching.image_matcher import ImageMatcher
from .matching.multimodal_matcher import MultiModalMatcher
from .matching.text_matcher import TextMatcher

# Import the refactored Processor class
from .processing.main_processor import ProductProcessor as Processor
from .scraping.koryo_scraper import KoryoScraper
from .scraping.naver_crawler import NaverShoppingCrawler

class ParallelProcessor:
    """병렬 처리를 위한 프로세서"""
    
    def __init__(
        self,
        max_workers: int = 5,
        queue_size: int = 100,
        batch_size: int = 10
    ):
        self.max_workers = max_workers
        self.queue_size = queue_size
        self.batch_size = batch_size
        self.task_queue = queue.Queue(maxsize=queue_size)
        self.result_queue = queue.Queue()
        self.error_queue = queue.Queue()
        self.stop_event = threading.Event()
        
    def process_items(
        self,
        items: List[Any],
        process_func: callable,
        **kwargs
    ) -> Tuple[List[Any], List[Exception]]:
        """
        아이템들을 병렬로 처리합니다.
        
        Args:
            items: 처리할 아이템 리스트
            process_func: 각 아이템을 처리할 함수
            **kwargs: process_func에 전달할 추가 인자
            
        Returns:
            처리된 결과와 에러 목록
        """
        results = []
        errors = []
        
        # 작업자 스레드 시작
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 작업 제출
            futures = []
            for i in range(0, len(items), self.batch_size):
                batch = items[i:i + self.batch_size]
                future = executor.submit(self._process_batch, batch, process_func, **kwargs)
                futures.append(future)
            
            # 결과 수집
            for future in as_completed(futures):
                try:
                    batch_results, batch_errors = future.result()
                    results.extend(batch_results)
                    errors.extend(batch_errors)
                except Exception as e:
                    errors.append(e)
                    
        return results, errors
        
    def _process_batch(
        self,
        batch: List[Any],
        process_func: callable,
        **kwargs
    ) -> Tuple[List[Any], List[Exception]]:
        """배치 단위로 아이템을 처리합니다."""
        results = []
        errors = []
        
        for item in batch:
            try:
                result = process_func(item, **kwargs)
                results.append(result)
            except Exception as e:
                errors.append(e)
                
        return results, errors

class OptimizedProcessor(Processor):
    """최적화된 제품 처리기"""
    
    def __init__(self, *args, max_workers: int = 5, **kwargs):
        super().__init__(*args, **kwargs)
        self.parallel_processor = ParallelProcessor(max_workers=max_workers)
        
    def process_products(self, products: List[Product]) -> List[ProcessingResult]:
        """제품 목록을 병렬로 처리합니다."""
        results, errors = self.parallel_processor.process_items(
            items=products,
            process_func=self._process_single_product
        )
        
        if errors:
            self.logger.warning(f"Processing completed with {len(errors)} errors")
            
        return results
        
    def _process_single_product(self, product: Product) -> ProcessingResult:
        """단일 제품을 처리합니다."""
        try:
            # 캐시 확인
            cache_key = f"process_{product.source}_{product.id}"
            cached_result = self.get_cached_result(cache_key)
            if cached_result:
                return cached_result
                
            # 실제 처리 로직
            result = super()._process_single_product(product)
            
            # 결과 캐싱
            self.cache_result(cache_key, result)
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error processing product {product.id}: {e}")
            return ProcessingResult(
                source_product=product,
                error=str(e)
            )
            
    def get_cached_result(self, key: str) -> Optional[ProcessingResult]:
        """캐시된 처리 결과를 가져옵니다."""
        if not hasattr(self, 'cache'):
            return None
            
        return self.cache.get(key)
        
    def cache_result(self, key: str, result: ProcessingResult) -> None:
        """처리 결과를 캐시에 저장합니다."""
        if not hasattr(self, 'cache'):
            return
            
        self.cache.set(key, result)

# For backward compatibility, re-export these classes
__all__ = ["Processor", "OptimizedProcessor", "Product", "MatchResult", "ProcessingResult"]

# Show deprecation warning
warnings.warn(
    "Direct import from 'processing' is deprecated. Use the new module structure instead: "
    "from core.processing import Processor",
    DeprecationWarning,
    stacklevel=2,
)

# Setup module logger
logger = logging.getLogger(__name__)
logger.info("Using optimized processing module structure")
