import logging
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Union, Any


class ExcelWriter:
    def __init__(self, config: dict, logger: logging.Logger = None):
        self.config = config
        self.logger = logger or logging.getLogger(__name__)
        
        # config 객체의 타입 확인
        if hasattr(config, 'sections') and callable(getattr(config, 'get', None)):
            # ConfigParser 객체인 경우 설정값 가져오기
            self._get_config = lambda section, key, default=None: config.get(section, key, fallback=default)
            try:
                self.excel_settings = dict(config["EXCEL"]) if "EXCEL" in config.sections() else {}
            except Exception as e:
                self.logger.warning(f"Excel 설정을 불러오는 중 오류 발생: {str(e)}")
                self.excel_settings = {}
        else:
            # dict 객체인 경우 설정값 가져오기
            self._get_config = lambda section, key, default=None: config.get(section, {}).get(key, default)
            self.excel_settings = config.get("EXCEL", {})
            
        # Style definitions
        self.header_font = Font(bold=True, size=11)
        self.default_font = Font(size=10)
        self.url_font = Font(color="0000FF", underline="single")
        self.error_font = Font(color="FF0000", bold=True)
        
        self.default_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        self.header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        self.error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def generate_enhanced_output(self, results: list, input_file: str, output_dir: str = None) -> str:
        """처리 결과를 엑셀로 저장하고 포맷팅을 적용합니다."""
        try:
            # 결과 데이터 준비
            report_data = []
            processed_count = 0

            for result in results:
                if hasattr(result, "source_product") and result.source_product:
                    processed_count += 1
                    row = {}

                    # 원본 제품 데이터 추출
                    source_data = getattr(result.source_product, "original_input_data", {})
                    if isinstance(source_data, dict):
                        for key, value in source_data.items():
                            row[key] = value

                    # 필수 필드 확인
                    row.setdefault("상품명", getattr(result.source_product, "name", "상품명 없음"))
                    row.setdefault("판매단가(V포함)", getattr(result.source_product, "price", 0))
                    row.setdefault("상품Code", getattr(result.source_product, "id", "DEFAULT-CODE"))
                    row.setdefault("본사상품링크", getattr(result.source_product, "url", ""))
                    
                    # 이미지 URL 처리
                    if "본사 이미지" not in row and hasattr(result.source_product, "image_url"):
                        row["본사 이미지"] = result.source_product.image_url
                    
                    # 고려기프트 매칭 데이터 추가
                    if hasattr(result, "best_koryo_match") and result.best_koryo_match:
                        self._add_match_data(row, result.best_koryo_match, "2", "고려기프트")
                    
                    # 네이버 매칭 데이터 추가
                    if hasattr(result, "best_naver_match") and result.best_naver_match:
                        self._add_match_data(row, result.best_naver_match, "3", "네이버")
                    elif hasattr(result, "naver_matches") and result.naver_matches:
                        # 최적 매칭이 없지만 다른 후보가 있는 경우 첫 번째 후보 사용
                        self._add_match_data(row, result.naver_matches[0], "3", "네이버")
                    
                    report_data.append(row)

            # 로깅 - 처리된 결과 수
            self.logger.info(f"총 {processed_count}개 제품 처리됨, 엑셀 파일에 {len(report_data)}행 작성")

            # 결과 데이터가 비어있는지 확인
            if not report_data:
                self.logger.warning("엑셀 보고서에 작성할 데이터 없음! 기본 데이터 생성")
                # 최소한의 데이터 생성
                empty_row = {
                    "상품명": "데이터 처리 중 오류 발생",
                    "판매단가(V포함)": 0,
                    "상품Code": "ERROR",
                    "구분": "ERROR",
                }
                report_data.append(empty_row)

            # DataFrame 생성
            result_df = pd.DataFrame(report_data)

            # 가격 차이 컬럼 계산
            self._calculate_price_differences(result_df)

            # 파일명 생성
            base_name = os.path.splitext(input_file)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
                output_file = os.path.join(output_dir, f'{os.path.basename(base_name)}-result_result_{timestamp}_upload_{timestamp}.xlsx')
            else:
                os.makedirs("output", exist_ok=True)
                output_file = os.path.join("output", f'{os.path.basename(base_name)}-result_result_{timestamp}_upload_{timestamp}.xlsx')

            # 엑셀로 저장
            result_df.to_excel(output_file, index=False)

            self.logger.info(f"결과 파일 생성 완료: {output_file} (총 {len(report_data)}행)")

            return output_file

        except Exception as e:
            self.logger.error(f"결과 파일 생성 중 오류 발생: {str(e)}", exc_info=True)

            # 오류가 발생해도 파일은 생성 시도
            try:
                # 기본 데이터로 빈 파일 생성
                error_data = [
                    {
                        "오류": str(e),
                        "상품명": "데이터 처리 중 오류 발생",
                        "시간": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    }
                ]

                error_df = pd.DataFrame(error_data)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                if output_dir:
                    os.makedirs(output_dir, exist_ok=True)
                    output_file = os.path.join(output_dir, f'{os.path.basename(os.path.splitext(input_file)[0])}-error-result_{timestamp}.xlsx')
                else:
                    os.makedirs("output", exist_ok=True)
                    output_file = os.path.join("output", f'{os.path.basename(os.path.splitext(input_file)[0])}-error-result_{timestamp}.xlsx')
                
                error_df.to_excel(output_file, index=False)

                self.logger.warning(f"오류 정보가 포함된 파일을 생성했습니다: {output_file}")
                return output_file

            except Exception as inner_e:
                self.logger.critical(f"오류 파일 생성 중 추가 오류 발생: {str(inner_e)}")
                return ""

    def _add_match_data(self, row: Dict, match_obj: Any, suffix: str, source: str) -> None:
        """매칭 데이터를 행에 추가"""
        try:
            if hasattr(match_obj, "matched_product"):
                product = match_obj.matched_product
                
                # 가격 데이터
                if hasattr(product, "price"):
                    row[f"판매단가(V포함)({suffix})"] = product.price
                
                # 기본 수량
                if hasattr(product, "min_order_quantity"):
                    row[f"기본수량({suffix})"] = product.min_order_quantity
                
                # 이미지 URL
                if hasattr(product, "image_url") and product.image_url:
                    row[f"{source} 이미지"] = product.image_url
                
                # 상품 링크
                if hasattr(product, "url") and product.url:
                    row[f"{source} 상품링크"] = product.url
                
                # 텍스트 유사도
                if hasattr(match_obj, "text_similarity"):
                    row[f"텍스트유사도({suffix})"] = round(match_obj.text_similarity, 2)
                
                # 상태 메시지
                text_similarity = getattr(match_obj, "text_similarity", 0)
                if text_similarity < 0.85:
                    row[f"매칭_상황({suffix})"] = f"텍스트 유사도 낮음 ({text_similarity:.2f})"
            else:
                row[f"매칭_상황({suffix})"] = f"{source}에서 상품을 찾지 못했습니다"
                row[f"판매단가(V포함)({suffix})"] = 0
                row[f"{source} 이미지"] = ""
                row[f"{source} 상품링크"] = ""
                
        except Exception as e:
            self.logger.error(f"{source} 매칭 데이터 추가 중 오류 발생: {str(e)}")
            row[f"매칭_상황({suffix})"] = f"오류 발생: {str(e)}"

    def _calculate_price_differences(self, df: pd.DataFrame) -> None:
        """가격 차이 및 백분율 계산"""
        try:
            # 1. 고려기프트 가격 차이 계산
            if "판매단가(V포함)" in df.columns and "판매단가(V포함)(2)" in df.columns:
                # 기존 열이 있으면 삭제
                if "가격차이(2)" in df.columns:
                    del df["가격차이(2)"]
                if "가격차이(2)(%)" in df.columns:
                    del df["가격차이(2)(%)"]

                # 가격 차이 계산
                df["가격차이(2)"] = df.apply(
                    lambda row: (
                        row["판매단가(V포함)(2)"] - row["판매단가(V포함)"]
                        if pd.notna(row.get("판매단가(V포함)(2)"))
                        and pd.notna(row.get("판매단가(V포함)"))
                        and row["판매단가(V포함)"] > 0
                        and row["판매단가(V포함)(2)"] > 0
                        else None
                    ),
                    axis=1,
                )

                # 가격 차이 백분율 계산
                df["가격차이(2)(%)"] = df.apply(
                    lambda row: (
                        (row["판매단가(V포함)(2)"] - row["판매단가(V포함)"]) / row["판매단가(V포함)"]
                        if pd.notna(row.get("판매단가(V포함)(2)"))
                        and pd.notna(row.get("판매단가(V포함)"))
                        and row["판매단가(V포함)"] > 0
                        and row["판매단가(V포함)(2)"] > 0
                        else None
                    ),
                    axis=1,
                )

            # 2. 네이버 가격 차이 계산
            if "판매단가(V포함)" in df.columns and "판매단가(V포함)(3)" in df.columns:
                # 기존 열이 있으면 삭제
                if "가격차이(3)" in df.columns:
                    del df["가격차이(3)"]
                if "가격차이(3)(%)" in df.columns:
                    del df["가격차이(3)(%)"]

                # 가격 차이 계산
                df["가격차이(3)"] = df.apply(
                    lambda row: (
                        row["판매단가(V포함)(3)"] - row["판매단가(V포함)"]
                        if pd.notna(row.get("판매단가(V포함)(3)"))
                        and pd.notna(row.get("판매단가(V포함)"))
                        and row["판매단가(V포함)"] > 0
                        and row["판매단가(V포함)(3)"] > 0
                        else None
                    ),
                    axis=1,
                )

                # 가격 차이 백분율 계산
                df["가격차이(3)(%)"] = df.apply(
                    lambda row: (
                        (row["판매단가(V포함)(3)"] - row["판매단가(V포함)"]) / row["판매단가(V포함)"]
                        if pd.notna(row.get("판매단가(V포함)(3)"))
                        and pd.notna(row.get("판매단가(V포함)"))
                        and row["판매단가(V포함)"] > 0
                        and row["판매단가(V포함)(3)"] > 0
                        else None
                    ),
                    axis=1,
                )
        except Exception as e:
            self.logger.error(f"가격 차이 계산 중 오류 발생: {str(e)}")

    def create_worksheet(self, workbook: Workbook, sheet_name: str):
        """워크시트 생성 및 헤더 설정"""
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # 헤더 정의
        headers = [
            "제품명", 
            "상품코드", 
            "가격", 
            "메인 이미지", 
            "추가 이미지", 
            "상품 URL", 
            "소스", 
            "브랜드", 
            "카테고리", 
            "스크래핑 시간", 
            "상세 설명",
            "원본 데이터"
        ]
        
        # 헤더 쓰기
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = self.header_fill
            cell.border = self.default_border
            
        # 열 너비 설정
        column_widths = {1:40, 2:15, 3:15, 4:25, 5:25, 6:30, 7:15, 8:20, 9:20, 10:20, 11:40, 12:50}
        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width
            
        # 상단 행 고정
        worksheet.freeze_panes = "A2"
        
        return worksheet

    def save_products(self, products: list, output_path: str, sheet_name: str = None, naver_results: list = None) -> str:
        """제품 목록을 Excel 파일로 저장"""
        try:
            # 출력 디렉토리 생성
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"Created output directory: {output_dir}")
                
            # 워크북 및 워크시트 생성
            workbook = Workbook()
            
            if not sheet_name:
                sheet_name = datetime.now().strftime("%Y%m%d_%H%M%S")
                
            # 기본 시트 제거 및 새 시트 생성
            workbook.remove(workbook.active)
            worksheet = self.create_worksheet(workbook, sheet_name)
            
            # 제품 데이터 작성
            current_row = 2  # 헤더 다음 행부터 시작
            total_products = 0
            
            if products:
                self.logger.info(f"Writing {len(products)} products to Excel")
                for product in products:
                    current_row = self._write_product_data(worksheet, current_row, product)
                    total_products += 1
                    
            if naver_results:
                self.logger.info(f"Writing {len(naver_results)} Naver results to Excel")
                for product in naver_results:
                    current_row = self._write_product_data(worksheet, current_row, product)
                    total_products += 1
                    
            # 결과 요약 추가
            summary_row = current_row + 1
            worksheet.cell(row=summary_row, column=1, value=f"총 제품 수: {total_products}")
            worksheet.cell(row=summary_row, column=2, value=f"처리 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # 파일 저장
            workbook.save(output_path)
            self.logger.info(f"Excel file saved to: {output_path}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error saving products to Excel: {str(e)}", exc_info=True)
            return output_path

    def _write_product_data(self, worksheet, row_idx: int, product) -> int:
        """제품 데이터를 워크시트에 작성"""
        try:
            # 기본 데이터 쓰기
            worksheet.cell(row=row_idx, column=1, value=getattr(product, 'name', '제품명 없음'))
            worksheet.cell(row=row_idx, column=2, value=getattr(product, 'product_code', getattr(product, 'id', '코드 없음')))
            worksheet.cell(row=row_idx, column=3, value=getattr(product, 'price', 0))
            
            # 이미지 URL 처리
            image_url = getattr(product, 'image_url', '')
            if image_url and isinstance(image_url, str) and image_url.startswith("http"):
                formula = f'=IMAGE("{image_url}", 2)'
                worksheet.cell(row=row_idx, column=4, value=formula)
            else:
                worksheet.cell(row=row_idx, column=4, value="이미지 없음")
                
            # 추가 이미지 갤러리
            gallery = getattr(product, 'image_gallery', [])
            if gallery and isinstance(gallery, (list, tuple)) and len(gallery) > 0:
                gallery_str = " | ".join(str(img) for img in gallery[:5])  # 첫 5개만 표시
                worksheet.cell(row=row_idx, column=5, value=gallery_str)
            else:
                worksheet.cell(row=row_idx, column=5, value="")
                
            # 상품 URL
            product_url = getattr(product, 'url', '')
            worksheet.cell(row=row_idx, column=6, value=product_url)
            if product_url and isinstance(product_url, str) and product_url.startswith("http"):
                cell = worksheet.cell(row=row_idx, column=6)
                cell.font = self.url_font
                cell.hyperlink = product_url
                
            # 기타 속성
            worksheet.cell(row=row_idx, column=7, value=getattr(product, 'source', '알 수 없음'))
            worksheet.cell(row=row_idx, column=8, value=getattr(product, 'brand', '알 수 없음'))
            worksheet.cell(row=row_idx, column=9, value=getattr(product, 'category', '알 수 없음'))
            worksheet.cell(row=row_idx, column=10, value=getattr(product, 'fetched_at', ''))
            worksheet.cell(row=row_idx, column=11, value=getattr(product, 'description', ''))
            
            # 원본 데이터 (긴 텍스트는 자르기)
            orig_data = str(getattr(product, 'original_input_data', ''))[:32000]
            worksheet.cell(row=row_idx, column=12, value=orig_data)
            
            # 모든 셀에 테두리 적용
            for col in range(1, 13):
                cell = worksheet.cell(row=row_idx, column=col)
                cell.border = self.default_border
                
                # 너무 긴 텍스트는 자동 줄바꿈
                if col in [1, 11, 12]:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
            return row_idx + 1
            
        except Exception as e:
            self.logger.error(f"Error writing product data: {str(e)}")
            # 오류 행 생성
            worksheet.cell(row=row_idx, column=1, value="데이터 처리 오류")
            worksheet.cell(row=row_idx, column=2, value=str(e)[:100])
            
            for col in range(1, 13):
                cell = worksheet.cell(row=row_idx, column=col)
                cell.border = self.default_border
                cell.fill = self.error_fill
                
            return row_idx + 1 