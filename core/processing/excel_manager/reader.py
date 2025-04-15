import logging
import os
import re
from pathlib import Path
from datetime import datetime
from configparser import ConfigParser
from typing import Optional, Dict, Union, Any, List, Tuple

import pandas as pd
import numpy as np
from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, config: Union[ConfigParser, Dict[str, Any]], logger: Optional[logging.Logger] = None):
        """
        Excel 리더 초기화
        
        Args:
            config: 설정 객체 (ConfigParser 또는 Dict)
            logger: 로거 객체
        """
        self.logger = logger or logging.getLogger(__name__)
        self.config = config
        
        # config가 ConfigParser인지 dict인지 확인하고 적절한 방식으로 값을 얻음
        self._get_config_value = self._make_config_getter(config)
        
        # Excel 설정
        self.excel_settings = {
            "sheet_name": self._get_config_value("EXCEL", "default_sheet_name", "Sheet1"),
            "alternative_sheet_names": self._get_config_value("EXCEL", "alternative_sheet_names", "").split(","),
            "header_row": int(self._get_config_value("EXCEL", "header_row", 1)),
            "start_row": int(self._get_config_value("EXCEL", "start_row", 2)),
            "alternative_start_rows": [int(x) for x in self._get_config_value("EXCEL", "alternative_start_rows", "0,1,3").split(",") if x],
            "max_rows": int(self._get_config_value("EXCEL", "max_rows", 1000)),
            "max_file_size_mb": int(self._get_config_value("EXCEL", "max_file_size_mb", 200)),
            "validation_rules": self._get_bool_config_value("EXCEL", "validation_rules", True),
            "enable_data_quality_metrics": self._get_bool_config_value("EXCEL", "enable_data_quality_metrics", True),
            "enable_duplicate_detection": self._get_bool_config_value("EXCEL", "enable_duplicate_detection", True),
            "enable_auto_correction": self._get_bool_config_value("EXCEL", "enable_auto_correction", True),
            "attempt_all_sheets": self._get_bool_config_value("EXCEL", "attempt_all_sheets", True),
            "flexible_column_mapping": self._get_bool_config_value("EXCEL", "flexible_column_mapping", True),
            "create_missing_columns": self._get_bool_config_value("EXCEL", "create_missing_columns", True),
            "enable_formatting": self._get_bool_config_value("EXCEL", "enable_formatting", True),
            "date_format": self._get_config_value("EXCEL", "date_format", "YYYY-MM-DD"),
            "number_format": self._get_config_value("EXCEL", "number_format", "#,##0")
        }
        
        # Get required columns from config
        required_cols_str = self._get_config_value("EXCEL", "required_columns", "상품명,판매단가(V포함),상품Code,본사 이미지,본사상품링크")
        self.excel_settings["required_columns"] = [col.strip() for col in required_cols_str.split(",") if col.strip()]
        
        # Column alternatives
        self.column_alternatives = {
            "상품명": self._get_config_value("EXCEL", "column_alternatives_상품명", "품명,제품명,상품,product,name,item,품목,상품이름").split(","),
            "판매단가(V포함)": self._get_config_value("EXCEL", "column_alternatives_판매단가", "단가,판매가,가격,price,가격(v포함),단가(vat),판매단가").split(","),
            "상품Code": self._get_config_value("EXCEL", "column_alternatives_상품code", "코드,code,item code,product code,품목코드,제품코드,상품코드").split(","),
            "본사 이미지": self._get_config_value("EXCEL", "column_alternatives_이미지", "이미지,image,상품이미지,제품이미지,이미지주소,image url").split(","),
            "본사상품링크": self._get_config_value("EXCEL", "column_alternatives_링크", "링크,link,url,상품링크,제품링크,상품url,제품url,홈페이지").split(",")
        }
        
        # 가격 유효성 검사 설정
        self.price_validation = {
            "min": float(self._get_config_value("EXCEL", "price_min", 0)),
            "max": float(self._get_config_value("EXCEL", "price_max", 10000000000))
        }
        
        self.logger.info(f"Excel 설정 초기화 완료: {len(self.excel_settings)} 항목")
        self.logger.debug(f"컬럼 대체어: {len(self.column_alternatives)} 항목")

    def _make_config_getter(self, config):
        """config 객체에 따라 적절한 getter 함수 반환"""
        if hasattr(config, 'get') and hasattr(config, 'sections'):
            # ConfigParser 객체인 경우
            return lambda section, option, fallback=None: config.get(section, option, fallback=fallback)
        else:
            # dict 객체인 경우
            return lambda section, option, fallback=None: config.get(section, {}).get(option, fallback)
    
    def _get_bool_config_value(self, section, option, fallback=False):
        """ConfigParser 또는 dict에서 불리언 값 얻기"""
        if hasattr(self.config, 'getboolean'):
            return self.config.getboolean(section, option, fallback=fallback)
        else:
            value = self._get_config_value(section, option, fallback)
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                return value.lower() in ('true', 'yes', '1', 'on')
            return bool(value)

    def _ensure_default_excel_settings(self) -> None:
        """설정에 필요한 기본값을 설정합니다."""
        defaults = {
            "sheet_name": "Sheet1",
            "start_row": 2,
            "required_columns": [
                "상품명",
                "판매단가(V포함)",
                "상품Code",
                "본사 이미지",
                "본사상품링크"
            ],
            "optional_columns": ["본사 단가", "가격", "상품코드"],
            "max_rows": 10000,
            "enable_formatting": True,
            "date_format": "YYYY-MM-DD",
            "number_format": "#,##0.00",
            "max_file_size_mb": 100,
            "enable_data_quality": True,
            "enable_duplicate_detection": True,
            "enable_auto_correction": True,
            "auto_correction_rules": ["price", "url", "product_code"],
            "report_formatting": True,
            "report_styles": True,
            "report_filters": True,
            "report_sorting": True,
            "report_freeze_panes": True,
            "report_auto_fit": True,
            "validation_rules": {
                "price": {"min": 0, "max": 1000000000},
                "product_code": {"pattern": r"^[A-Za-z0-9-]+$"},
                "url": {"pattern": r"^https?://.*$"}
            }
        }
        for key, value in defaults.items():
            if key not in self.excel_settings:
                self.excel_settings[key] = value
            elif key == "validation_rules" and isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    if sub_key not in self.excel_settings[key]:
                        self.excel_settings[key][sub_key] = sub_value

    def _clean_url(self, url: str) -> str:
        """URL 정제 및 이미지 URL 유효성 검사"""
        if not url or not isinstance(url, str):
            return ""
        cleaned_url = url.strip().replace('@', '')
        if cleaned_url.startswith("http:"):
            cleaned_url = "https:" + cleaned_url[5:]
        elif cleaned_url.startswith("//"):
            cleaned_url = "https:" + cleaned_url
        elif not cleaned_url.startswith("https:"):
            if '.' in cleaned_url:
                cleaned_url = "https://" + cleaned_url
            else:
                return cleaned_url
        cleaned_url = cleaned_url.replace('\\', '/').replace('"', '%22').replace(' ', '%20')
        return cleaned_url

    def _create_minimal_error_dataframe(self, error_message: str) -> pd.DataFrame:
        """오류 상황에서 기본 데이터프레임을 생성합니다."""
        required_columns = self.excel_settings["required_columns"]
        error_data = {col: [""] for col in required_columns}

        # 첫 번째 행에 오류 메시지 표시
        if "상품명" in required_columns:
            error_data["상품명"] = [f"Error: {error_message}"]

        if "판매단가(V포함)" in required_columns:
            error_data["판매단가(V포함)"] = [0]

        if "상품Code" in required_columns:
            error_data["상품Code"] = ["ERROR"]

        return pd.DataFrame(error_data)

    def _ensure_required_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """데이터프레임에 필수 컬럼이 있는지 확인하고 없으면 추가합니다."""
        df_copy = df.copy()
        required_columns = self.excel_settings["required_columns"]

        # 해오름기프트 엑셀 형식에 맞게 컬럼 형태 표준화
        # 샘플: 구분 담당자 업체명 업체코드 상품Code 중분류카테고리 상품명 기본수량(1) 판매단가(V포함) 본사상품링크
        haeoreum_columns = [
            "구분",
            "담당자",
            "업체명",
            "업체코드",
            "상품Code",
            "중분류카테고리",
            "상품명",
            "기본수량(1)",
            "판매단가(V포함)",
            "본사상품링크",
        ]

        # 컬럼명 공백 제거 및 표준화
        renamed_columns = {}
        for col in df_copy.columns:
            cleaned_col = str(col).strip()
            if cleaned_col != col:
                renamed_columns[col] = cleaned_col

        if renamed_columns:
            df_copy = df_copy.rename(columns=renamed_columns)
            self.logger.info(f"Cleaned column names: {renamed_columns}")

        # 컬럼명 매핑 (해오름 형식에 맞게)
        column_mapping = {
            # 기본 필드 매핑
            "Code": "상품Code",
            "상품코드": "상품Code",
            "상품 코드": "상품Code",
            "상품번호": "상품Code",
            "상품이름": "상품명",
            "상품 이름": "상품명",
            "제품명": "상품명",
            "판매가(V포함)": "판매단가(V포함)",
            "판매가(VAT포함)": "판매단가(V포함)",
            "판매가": "판매단가(V포함)",
            "가격": "판매단가(V포함)",
            "가격(V포함)": "판매단가(V포함)",
            "본사링크": "본사상품링크",
            "상품링크": "본사상품링크",
            "URL": "본사상품링크",
            "수량": "기본수량(1)",
            "기본수량": "기본수량(1)",
            "이미지": "본사 이미지",
            "이미지URL": "본사 이미지",
            "제품이미지": "본사 이미지",
            "상품이미지": "본사 이미지",
        }

        # 해오름 형식에 맞게 컬럼 리매핑
        for src_col, target_col in column_mapping.items():
            if src_col in df_copy.columns and target_col not in df_copy.columns:
                df_copy[target_col] = df_copy[src_col]
                self.logger.info(f"Mapped column '{src_col}' to '{target_col}'")

        # 필수 컬럼 확인 및 생성
        for col in required_columns:
            if col not in df_copy.columns:
                # 유사한 컬럼 찾기
                similar_col = self._find_similar_column(df_copy, col)

                if similar_col:
                    self.logger.info(f"Mapped '{similar_col}' to required column '{col}'")
                    df_copy[col] = df_copy[similar_col]
                else:
                    self.logger.warning(f"Creating default values for missing column '{col}'")
                    # 컬럼 유형에 따른 기본값 설정
                    if "단가" in col or "가격" in col:
                        df_copy[col] = 0
                    elif "Code" in col or "코드" in col:
                        df_copy[col] = df_copy.index.map(lambda i: f"GEN-{i}")
                    elif "이미지" in col:
                        df_copy[col] = ""
                    elif "링크" in col:
                        df_copy[col] = ""
                    else:
                        df_copy[col] = ""

        # 소스 컬럼 추가 (해오름기프트 소스 명시)
        if "source" not in df_copy.columns:
            df_copy["source"] = "haeoreum"
            self.logger.info("Added 'source' column with value 'haeoreum'")

        return df_copy

    def _find_similar_column(self, df: pd.DataFrame, target_column: str) -> Optional[str]:
        """데이터프레임에서 타겟 컬럼과 유사한 컬럼을 찾습니다."""
        # 대소문자 무시, 유사성 검사
        for col in df.columns:
            if col.lower() == target_column.lower():
                return col

        # 컬럼 대체어 검사
        if target_column in self.column_alternatives:
            for similar in self.column_alternatives[target_column]:
                similar = similar.strip()
                for col in df.columns:
                    if similar.lower() in col.lower():
                        return col

        return None

    def _try_read_with_various_settings(self, file_path: str) -> List[Tuple[pd.DataFrame, str, int]]:
        """다양한 설정으로 엑셀 파일 읽기 시도"""
        all_dataframes = []
        file_path_obj = Path(file_path)
        
        if not file_path_obj.exists():
            self.logger.error(f"Excel file not found: {file_path}")
            return all_dataframes

        # 시트 확인
        sheet_names = []
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()  # Close workbook to free resources
        except Exception as e:
            self.logger.warning(f"Could not inspect Excel structure: {str(e)}. Using default sheet names.")
            sheet_names = [self.excel_settings["sheet_name"]] + self.excel_settings["alternative_sheet_names"]

        # 시트별 읽기 시도
        for sheet_name in sheet_names:
            if not sheet_name:
                continue
                
            # 다양한 행 스킵 값으로 시도
            for skip_rows in [0] + self.excel_settings["alternative_start_rows"]:
                try:
                    df = pd.read_excel(
                        file_path, 
                        sheet_name=sheet_name,
                        skiprows=skip_rows,
                        engine="openpyxl",
                        na_values=['#N/A', '#NA', '#NULL', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?']
                    )
                    
                    if not df.empty:
                        # 열 이름 정리 (공백 제거)
                        df.columns = [str(col).strip() if isinstance(col, str) else col for col in df.columns]
                        self.logger.info(f"Found data in sheet '{sheet_name}' with skiprows={skip_rows}, {len(df)} rows")
                        all_dataframes.append((df, sheet_name, skip_rows))
                except Exception as e:
                    self.logger.debug(f"Failed with sheet='{sheet_name}', skiprows={skip_rows}: {str(e)}")

        # 대체 엔진 시도
        if not all_dataframes:
            try:
                # xlrd 엔진 시도 (오래된 Excel 형식)
                df = pd.read_excel(file_path, engine="xlrd")
                if not df.empty:
                    all_dataframes.append((df, "Unknown (xlrd)", 0))
            except Exception:
                pass

            try:
                # CSV 형식 시도
                df = pd.read_csv(file_path, encoding='utf-8')
                if not df.empty:
                    all_dataframes.append((df, "CSV", 0))
            except Exception:
                try:
                    # CP949 인코딩으로 시도
                    df = pd.read_csv(file_path, encoding='cp949')
                    if not df.empty:
                        all_dataframes.append((df, "CSV (CP949)", 0))
                except Exception:
                    pass

        return all_dataframes

    def read_excel_file(self, file_path: str) -> pd.DataFrame:
        """Excel 파일을 읽고 검증하여 DataFrame을 반환합니다."""
        try:
            # 파일 존재 확인
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                self.logger.error(f"Excel file not found: {file_path}")
                return self._create_minimal_error_dataframe("File not found")

            self.logger.info(f"Reading Excel file: {file_path}")

            # 다양한 방법으로 읽기 시도
            all_dataframes = self._try_read_with_various_settings(file_path)

            # 데이터프레임이 없으면 기본 형식 반환
            if not all_dataframes:
                self.logger.error(f"Could not read any valid data from {file_path}")
                return self._create_minimal_error_dataframe("No data found")

            # 가장 많은 행을 가진 데이터프레임 선택
            best_df, sheet, skiprows = max(all_dataframes, key=lambda x: len(x[0]))
            self.logger.info(f"Selected dataframe from sheet '{sheet}' with {len(best_df)} rows")

            # 필수 컬럼 확인 및 추가
            best_df = self._ensure_required_columns(best_df)

            # 가격 데이터 정제
            if "판매단가(V포함)" in best_df.columns:
                try:
                    best_df["판매단가(V포함)"] = pd.to_numeric(
                        best_df["판매단가(V포함)"].astype(str).str.replace(',', ''), 
                        errors='coerce'
                    )
                    best_df["판매단가(V포함)"].fillna(0, inplace=True)
                except Exception as e:
                    self.logger.warning(f"Error converting price data: {e}")

            # NaN 값을 적절한 값으로 대체
            best_df.fillna({
                "상품명": "제품명 없음",
                "상품Code": "코드 없음",
                "본사상품링크": "",
                "본사 이미지": ""
            }, inplace=True)

            # URL 정제
            if "본사상품링크" in best_df.columns:
                best_df["본사상품링크"] = best_df["본사상품링크"].apply(self._clean_url)
            
            if "본사 이미지" in best_df.columns:
                best_df["본사 이미지"] = best_df["본사 이미지"].apply(self._clean_url)

            self.logger.info(f"Excel file read successfully: {file_path}, {len(best_df)} rows")
            return best_df

        except Exception as e:
            self.logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
            return self._create_minimal_error_dataframe(f"Error: {str(e)}") 