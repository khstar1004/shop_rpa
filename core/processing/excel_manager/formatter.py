import logging
import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from typing import Union, Dict, Any, Optional
from urllib.parse import urlparse


class ExcelFormatter:
    def __init__(self, config: Union[Dict[str, Any], object], logger: logging.Logger = None):
        self.config = config
        self.logger = logger or logging.getLogger(__name__)
        
        # config 객체의 타입 확인
        if hasattr(config, 'sections') and callable(getattr(config, 'get', None)):
            # ConfigParser 객체인 경우 설정값 가져오기
            self._get_config = lambda section, key, default=None: config.get(section, key, fallback=default)
        else:
            # dict 객체인 경우 설정값 가져오기
            self._get_config = lambda section, key, default=None: config.get(section, {}).get(key, default)
        
        # Styling fills for different purposes
        self.price_difference_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )  # Yellow for price differences
        self.header_fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
        )  # Light yellow for headers
        self.alt_row_fill = PatternFill(
            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
        )  # Light gray for alternate rows
        self.group_header_fill = PatternFill(
            start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
        )  # Lavender for group headers
        
        # Font styles
        self.bold_font = Font(bold=True)
        self.header_font = Font(bold=True, size=11)
        self.hyperlink_font = Font(color="0000FF", underline="single")
        self.error_font = Font(color="FF0000", bold=True)
        
        # Border styles
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Alignment
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        self.wrap_text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Height settings
        self.row_height = 18  # Default row height
        self.image_row_height = 110  # Height for rows with images

    def apply_formatting_to_excel(self, excel_file: str) -> None:
        """엑셀 파일에 포맷팅을 적용합니다."""
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            # 1. Format header row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = self.bold_font
                cell.alignment = self.center_align
                cell.fill = self.header_fill
                cell.border = self.thin_border

            # 2. Format image columns - set wider columns and taller rows
            image_col_indices = []
            status_col_indices = []

            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and ("이미지" in str(header)):
                    image_col_indices.append(col)
                    # Make columns with images wider
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = 20
                elif header and (
                    "매칭_상황" in str(header) or "텍스트유사도" in str(header) or "상태" in str(header)
                ):
                    status_col_indices.append(col)
                    # 상태 메시지 컬럼 넓게 설정
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = 30

            # 3. Apply zebra striping and set appropriate row heights
            for row in range(2, ws.max_row + 1):
                # Set row height - taller for rows with images or error messages
                has_image = False
                has_error = False

                for col in image_col_indices:
                    cell_value = ws.cell(row=row, column=col).value
                    if (
                        cell_value
                        and isinstance(cell_value, str)
                        and (
                            cell_value.startswith("http")
                            or cell_value.startswith("=IMAGE")
                        )
                    ):
                        has_image = True
                        break

                for col in status_col_indices:
                    cell_value = ws.cell(row=row, column=col).value
                    if (
                        cell_value
                        and isinstance(cell_value, str)
                        and len(cell_value) > 10
                    ):  # 긴 메시지가 있는 경우
                        has_error = True
                        break

                if has_image:
                    ws.row_dimensions[row].height = self.image_row_height
                elif has_error:
                    ws.row_dimensions[row].height = 60  # 오류 메시지를 위한 높이
                else:
                    ws.row_dimensions[row].height = self.row_height

                # Apply zebra striping (alternate row coloring)
                if row % 2 == 0:  # Even rows get light coloring
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if (
                            not cell.fill or cell.fill.start_color.index == "FFFFFF"
                        ):  # Only if not already colored
                            cell.fill = self.alt_row_fill

                # Apply border to all cells
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.thin_border

                    # 오류 메시지에 빨간색 폰트 적용
                    if col in status_col_indices:
                        cell_value = cell.value
                        if (
                            cell_value
                            and isinstance(cell_value, str)
                            and ("상품이 없음" in cell_value or "오류" in cell_value or "실패" in cell_value)
                        ):
                            cell.font = self.error_font
                            cell.alignment = Alignment(
                                wrap_text=True, vertical="center"
                            )

            # 4. Find price difference columns and apply formatting
            price_diff_cols = []
            price_percent_cols = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    if "가격차이" in str(header) and "%" not in str(header):
                        price_diff_cols.append(col)
                    elif "가격차이" in str(header) and "%" in str(header):
                        price_percent_cols.append(col)

            # 5. Apply highlighting for negative price differences
            for row in range(2, ws.max_row + 1):
                should_highlight_row = False

                # Check if any price difference column has negative value
                for col in price_diff_cols:
                    cell = ws.cell(row=row, column=col)
                    try:
                        value = float(cell.value) if cell.value is not None else 0
                        if value < 0:
                            cell.fill = self.price_difference_fill
                            should_highlight_row = True  # Mark for row highlighting
                    except (ValueError, TypeError):
                        pass  # 숫자가 아닌 셀은 무시

                # Apply number formatting to percentage columns
                for col in price_percent_cols:
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0%"
                    elif isinstance(cell.value, str):
                        # Try to convert string percentage to number format
                        try:
                            if cell.value.endswith("%"):
                                value = float(cell.value.rstrip("%")) / 100
                                cell.value = value
                                cell.number_format = "0.0%"
                        except (ValueError, TypeError):
                            pass

                # 오류 메시지가 있는 텍스트 셀 포맷팅
                for col in status_col_indices:
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    if value and isinstance(value, str):
                        if (
                            "일정 정확도" in value
                            or "범위 내에 없음" in value
                            or "찾을 수 없음" in value
                        ):
                            cell.font = self.error_font
                            cell.alignment = Alignment(
                                wrap_text=True, vertical="center"
                            )

            # 6. Adjust column widths based on content
            for col in range(1, ws.max_column + 1):
                if col in status_col_indices:
                    # 상태 메시지 컬럼은 이미 설정했으므로 건너뜀
                    continue

                if col in image_col_indices:
                    # 이미지 컬럼은 이미 설정했으므로 건너뜀
                    continue

                max_length = 0
                column = get_column_letter(col)

                # Check all values in the column
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except (TypeError, ValueError):
                            pass

                # Adjust column width based on content
                adjusted_width = min(max_length + 4, 50)  # Cap width at 50
                if "상품명" in str(ws.cell(row=1, column=col).value):
                    adjusted_width = max(
                        adjusted_width, 35
                    )  # Product names need more space
                elif "링크" in str(ws.cell(row=1, column=col).value):
                    adjusted_width = max(adjusted_width, 25)  # Links need more space
                else:
                    adjusted_width = max(
                        adjusted_width, 12
                    )  # Minimum width for readability

                ws.column_dimensions[column].width = adjusted_width

            # 7. Freeze panes at A2 (keep header visible while scrolling)
            ws.freeze_panes = "A2"

            # 8. Save the formatted workbook
            wb.save(excel_file)
            self.logger.info(f"Enhanced formatting applied to {excel_file}")

        except Exception as e:
            self.logger.error(
                f"Error applying enhanced formatting: {str(e)}", exc_info=True
            )

    def add_hyperlinks_to_excel(self, file_path: str) -> str:
        """
        엑셀 파일의 URL 필드를 하이퍼링크로 변환합니다.

        Args:
            file_path: 입력 엑셀 파일 경로

        Returns:
            str: 처리된 파일 경로
        """
        try:
            self.logger.info(f"Adding hyperlinks to Excel file: {file_path}")

            # 출력 파일명 생성
            now = datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            output_directory = os.path.dirname(file_path)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_filename = f"{base_name}_result_{timestamp}.xlsx"
            output_file_path = os.path.join(output_directory, output_filename)

            # DataFrame 로드
            df = pd.read_excel(file_path)

            # 워크북 로드
            wb = load_workbook(file_path)
            ws = wb.active

            # 모든 열의 너비 설정 및 자동 줄 바꿈, 테두리 적용
            columns = [get_column_letter(i) for i in range(1, ws.max_column + 1)]
            for col in columns:
                ws.column_dimensions[col].width = 16

                for row in range(1, df.shape[0] + 2):  # 첫 번째 행은 헤더
                    cell = ws[f"{col}{row}"]
                    cell.alignment = Alignment(wrap_text=True)
                    cell.border = self.thin_border

            # 링크 컬럼 처리
            link_columns = [
                "본사상품링크",
                "고려기프트 상품링크",
                "네이버 쇼핑 링크",
                "공급사 상품링크",
                "본사링크",
                "고려 링크",
                "네이버 링크",
            ]
            
            # Process all columns containing 'link' or 'URL' in their name
            for col in df.columns:
                if '링크' in col or 'link' in col.lower() or 'url' in col.lower():
                    if col not in link_columns:
                        link_columns.append(col)
                        
            # 각 링크 컬럼에 대해 하이퍼링크 적용
            for link_column in link_columns:
                if link_column in df.columns:
                    for row_idx, link in enumerate(df[link_column], start=2):  # 첫 번째 행은 헤더
                        if (pd.notna(link) and isinstance(link, str) and 
                            (link.startswith("http") or link.startswith("https") or link.startswith("www"))):
                            
                            # Ensure URL has proper protocol
                            clean_url = link
                            if link.startswith("www."):
                                clean_url = "https://" + link
                                
                            # Add hyperlink
                            cell = ws.cell(row=row_idx, column=df.columns.get_loc(link_column) + 1)
                            cell.value = clean_url
                            cell.hyperlink = clean_url
                            cell.font = self.hyperlink_font
                            cell.alignment = Alignment(vertical='center')

            # 헤더 색상 적용
            for cell in ws["1:1"]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self.thin_border

            # 가격차이가 음수인 행에 노란색 배경 적용
            for row_idx in range(2, df.shape[0] + 2):
                for col_name in df.columns:
                    if "가격차이" in col_name and "%" not in col_name:
                        try:
                            col_idx = df.columns.get_loc(col_name) + 1
                            cell = ws.cell(row=row_idx, column=col_idx)
                            value = cell.value
                            
                            if pd.notna(value):
                                # 숫자 또는 문자열 변환 시도
                                try:
                                    value = float(value)
                                    if value < 0:
                                        # 전체 행 강조
                                        for col in range(1, ws.max_column + 1):
                                            cell = ws.cell(row=row_idx, column=col)
                                            cell.fill = self.price_difference_fill
                                        break
                                except (ValueError, TypeError):
                                    # 문자열이 '-'로 시작하는지 확인
                                    if isinstance(value, str) and value.strip().startswith('-'):
                                        for col in range(1, ws.max_column + 1):
                                            cell = ws.cell(row=row_idx, column=col)
                                            cell.fill = self.price_difference_fill
                                        break
                        except (ValueError, KeyError):
                            continue

            # 결과 저장
            wb.save(output_file_path)
            self.logger.info(f"Excel file with hyperlinks saved to: {output_file_path}")

            return output_file_path

        except Exception as e:
            self.logger.error(f"Error adding hyperlinks to Excel: {str(e)}", exc_info=True)
            return file_path

    def filter_excel_by_price_diff(self, file_path: str) -> str:
        """
        가격차이가 있는 항목들만 필터링하여 업로드용 엑셀 파일을 생성합니다.

        Args:
            file_path: 입력 엑셀 파일 경로

        Returns:
            str: 필터링된 출력 파일 경로
        """
        try:
            self.logger.info(f"Filtering Excel file by price differences: {file_path}")

            # 출력 파일명 생성
            input_filename = os.path.splitext(os.path.basename(file_path))[0]
            current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{input_filename}_upload_{current_datetime}.xlsx"
            output_dir = os.path.join(os.path.dirname(file_path), 'output')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, output_filename)

            # 데이터프레임 로드
            df = pd.read_excel(file_path)

            # 문자열을 숫자로 변환
            def to_float(value):
                try:
                    return float(value)
                except (ValueError, TypeError):
                    return None

            # 가격차이 컬럼 확인 및 변환
            price_diff_columns = [col for col in df.columns if "가격차이" in col and "%" not in col]
            for col in price_diff_columns:
                df[col] = df[col].apply(to_float)

            # 가격차이가 음수인 항목만 필터링
            price_diff_condition = False
            for col in price_diff_columns:
                price_diff_condition |= df[col].notna() & (df[col] < 0)

            if not isinstance(price_diff_condition, bool) and not price_diff_condition.any():
                self.logger.warning("No items with negative price differences found.")
                return file_path

            filtered_df = df[price_diff_condition] if not isinstance(price_diff_condition, bool) else df

            if filtered_df.empty:
                self.logger.warning("No items with negative price differences found.")
                return file_path

            # 원하는 컬럼만 선택
            required_columns = [
                "구분",
                "담당자",
                "업체명",
                "업체코드",
                "Code",
                "상품Code",
                "중분류카테고리",
                "상품명",
                "기본수량(1)",
                "판매단가(V포함)",
                "본사상품링크",
                "기본수량(2)",
                "판매단가(V포함)(2)",
                "가격차이(2)",
                "가격차이(2)(%)",
                "고려기프트 상품링크",
                "기본수량(3)",
                "판매단가(V포함)(3)",
                "가격차이(3)",
                "가격차이(3)(%)",
                "공급사명",
                "공급사 상품링크",
                "본사 이미지",
                "고려기프트 이미지",
                "네이버 이미지",
            ]

            # 존재하는 컬럼만 선택
            existing_columns = [col for col in required_columns if col in filtered_df.columns]
            filtered_df = filtered_df[existing_columns]

            # 컬럼 이름 변경
            column_mapping = {
                "구분": "구분(승인관리:A/가격관리:P)",
                "담당자": "담당자",
                "업체명": "공급사명",
                "업체코드": "공급처코드",
                "Code": "상품코드",
                "상품Code": "상품코드",
                "중분류카테고리": "카테고리(중분류)",
                "상품명": "상품명",
                "기본수량(1)": "본사 기본수량",
                "판매단가(V포함)": "판매단가1(VAT포함)",
                "본사상품링크": "본사링크",
                "기본수량(2)": "고려 기본수량",
                "판매단가(V포함)(2)": "판매단가2(VAT포함)",
                "가격차이(2)": "고려 가격차이",
                "가격차이(2)(%)": "고려 가격차이(%)",
                "고려기프트 상품링크": "고려 링크",
                "기본수량(3)": "네이버 기본수량",
                "판매단가(V포함)(3)": "판매단가3 (VAT포함)",
                "가격차이(3)": "네이버 가격차이",
                "가격차이(3)(%)": "네이버가격차이(%)",
                "공급사명": "네이버 공급사명",
                "공급사 상품링크": "네이버 링크",
                "본사 이미지": "해오름(이미지링크)",
                "고려기프트 이미지": "고려기프트(이미지링크)",
                "네이버 이미지": "네이버쇼핑(이미지링크)",
            }

            # 존재하는 컬럼만 매핑
            rename_mapping = {k: v for k, v in column_mapping.items() if k in filtered_df.columns}
            filtered_df.rename(columns=rename_mapping, inplace=True)

            # 엑셀로 저장
            filtered_df.to_excel(output_path, index=False)

            # 추가 포맷팅 적용
            wb = load_workbook(output_path)
            ws = wb.active

            # 테두리 및 정렬 적용
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.border = self.thin_border

            for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 컬럼 너비 및 행 높이 설정
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, max_row=1):
                for cell in col:
                    ws.column_dimensions[cell.column_letter].width = 15

            for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    ws.row_dimensions[cell.row].height = 16.5

            # 품절 항목 및 특정 조건 처리
            self._apply_filter_rules(ws)

            # 변경사항 저장
            wb.save(output_path)
            self.logger.info(f"Filtered Excel file saved to: {output_path}")

            return output_path

        except Exception as e:
            self.logger.error(f"Error filtering Excel by price differences: {str(e)}", exc_info=True)
            return file_path

    def _apply_filter_rules(self, worksheet):
        """필터링 규칙을 적용합니다."""
        try:
            # 품절 행 식별 및 삭제
            # 컬럼 인덱스 찾기
            goleo_quantity_col = None
            naver_quantity_col = None
            goleo_link_col = None
            naver_link_col = None

            for i, cell in enumerate(worksheet[1]):
                header_value = str(cell.value).lower() if cell.value else ""
                if "고려" in header_value and ("기본수량" in header_value or "수량" in header_value):
                    goleo_quantity_col = i
                elif "네이버" in header_value and ("기본수량" in header_value or "수량" in header_value):
                    naver_quantity_col = i
                elif "고려" in header_value and ("링크" in header_value or "link" in header_value):
                    goleo_link_col = i
                elif "네이버" in header_value and ("링크" in header_value or "link" in header_value):
                    naver_link_col = i

            # 품절 행 삭제 (역순으로 진행하여 인덱스 문제 방지)
            rows_to_delete = []
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                # 고려 기본수량이 품절인지 확인
                if goleo_quantity_col is not None and row[goleo_quantity_col].value == "품절":
                    rows_to_delete.append(row_idx)
                    continue
                
                # 네이버 기본수량이 품절인지 확인
                if naver_quantity_col is not None and row[naver_quantity_col].value == "품절":
                    rows_to_delete.append(row_idx)
                    continue
                
                # 링크 검증
                if goleo_link_col is not None:
                    link_value = row[goleo_link_col].value
                    if (link_value and isinstance(link_value, str) and 
                        not (link_value.startswith("http") or link_value.startswith("https"))):
                        row[goleo_link_col].value = None

                if naver_link_col is not None:
                    link_value = row[naver_link_col].value
                    if (link_value and isinstance(link_value, str) and 
                        not (link_value.startswith("http") or link_value.startswith("https"))):
                        row[naver_link_col].value = None
            
            # 행 삭제 (역순으로)
            for row_idx in sorted(rows_to_delete, reverse=True):
                worksheet.delete_rows(row_idx)
                
            self.logger.info(f"Removed {len(rows_to_delete)} rows with sold-out products")
            
        except Exception as e:
            self.logger.error(f"Error applying filter rules: {str(e)}", exc_info=True) 