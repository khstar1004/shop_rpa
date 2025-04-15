import os
import re
import logging
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from typing import Dict, List, Optional, Union, Any
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelPostProcessor:
    def __init__(self, config: Union[Dict[str, Any], object], logger: Optional[logging.Logger] = None):
        """
        Excel 후처리 모듈 초기화
        
        Args:
            config: 설정 객체
            logger: 로깅 인스턴스
        """
        self.config = config
        self.logger = logger or logging.getLogger(__name__)
        
        # config 객체의 타입 확인
        if hasattr(config, 'sections') and callable(getattr(config, 'get', None)):
            # ConfigParser 객체인 경우 설정값 가져오기
            self._get_config = lambda section, key, default=None: config.get(section, key, fallback=default)
        else:
            # dict 객체인 경우 설정값 가져오기
            self._get_config = lambda section, key, default=None: config.get(section, {}).get(key, default)

    def post_process_excel_file(self, file_path: str) -> str:
        """
        Excel 파일의 후처리 작업을 수행합니다.
        
        후처리 작업:
        1. 하이퍼링크 추가
        2. @ 기호 제거
        3. 상품 품절 상태 및 링크 검증
        4. 필요한 컬럼 추가
        
        Args:
            file_path: 처리할 파일 경로
            
        Returns:
            처리된 파일 경로
        """
        try:
            if not os.path.exists(file_path):
                self.logger.error(f"File not found: {file_path}")
                return file_path
                
            # 출력 파일명 생성
            now = datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            output_directory = os.path.dirname(file_path)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_filename = f"{base_name}_processed_{timestamp}.xlsx"
            output_file_path = os.path.join(output_directory, output_filename)
            
            # DataFrame 로드
            df = pd.read_excel(file_path)
            
            # 문제 파일이면 바로 복사하고 종료
            if df.empty:
                df.to_excel(output_file_path, index=False)
                self.logger.warning(f"Empty DataFrame, copied without processing: {output_file_path}")
                return output_file_path
                
            # @ 기호 제거
            for col in df.columns:
                if df[col].dtype == 'object':  # 문자열 컬럼만 처리
                    df[col] = df[col].apply(
                        lambda x: x.replace('@', '') if isinstance(x, str) else x
                    )
            
            # URL 필드 정리
            url_columns = [col for col in df.columns if '링크' in col or 'url' in col.lower() or 'link' in col.lower()]
            for col in url_columns:
                df[col] = df[col].apply(self._clean_url)
            
            # 이미지 URL 필드 정리
            image_columns = [col for col in df.columns if '이미지' in col or 'image' in col.lower()]
            for col in image_columns:
                df[col] = df[col].apply(
                    lambda x: f'=IMAGE("{self._clean_url(x)}", 2)' if isinstance(x, str) and x.startswith('http') else x
                )
            
            # 필요한 컬럼이 없으면 추가
            required_columns = ["본사 이미지", "고려기프트 이미지", "네이버 이미지"]
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ""
                    self.logger.info(f"Added missing column: {col}")
            
            # 가격 컬럼 숫자로 변환
            numeric_columns = [col for col in df.columns if '가격' in col or '단가' in col]
            for col in numeric_columns:
                try:
                    df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')
                    df[col] = df[col].fillna(0)
                except Exception as e:
                    self.logger.warning(f"Failed to convert {col} to numeric: {e}")
            
            # 저장
            df.to_excel(output_file_path, index=False)
            
            # 서식 적용
            self._apply_formatting(output_file_path)
            
            self.logger.info(f"Post-processed file saved to: {output_file_path}")
            return output_file_path
            
        except Exception as e:
            self.logger.error(f"Error post-processing Excel file: {str(e)}", exc_info=True)
            return file_path

    def _clean_url(self, url):
        """URL 정제 및 유효성 검사"""
        if not url or not isinstance(url, str):
            return ""
            
        url = url.strip().replace('@', '')
        
        # 이미지 함수에서 URL 추출
        if url.startswith('=IMAGE('):
            match = re.search(r'=IMAGE\("([^"]+)"', url)
            if match:
                url = match.group(1)
        
        # URL 프로토콜 정규화
        if url.startswith("http:"):
            url = "https:" + url[5:]
        elif url.startswith("//"):
            url = "https:" + url
        elif not url.startswith("https:"):
            if '.' in url and not url.strip().lower() in ['없음', 'none', 'null', 'na', '-']:
                url = "https://" + url
        
        return url

    def _apply_formatting(self, file_path: str) -> None:
        """엑셀 파일에 서식 적용"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 스타일 정의
            header_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # 헤더 서식 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border
            
            # 이미지 및 URL 컬럼 식별
            image_cols = []
            url_cols = []
            
            for idx, cell in enumerate(ws[1], 1):
                if not cell.value:
                    continue
                header_value = str(cell.value).lower()
                if "이미지" in header_value or "image" in header_value:
                    image_cols.append(idx)
                    ws.column_dimensions[get_column_letter(idx)].width = 20
                elif "링크" in header_value or "link" in header_value or "url" in header_value:
                    url_cols.append(idx)
                    ws.column_dimensions[get_column_letter(idx)].width = 30
            
            # 모든 셀에 테두리 적용 및 이미지 행 높이 조정
            for row_idx in range(2, ws.max_row + 1):
                # 이미지가 있는 행 높이 조정
                has_image = False
                for col_idx in image_cols:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str) and (
                        cell.value.startswith("http") or cell.value.startswith("=IMAGE")
                    ):
                        has_image = True
                        break
                
                if has_image:
                    ws.row_dimensions[row_idx].height = 110
                else:
                    ws.row_dimensions[row_idx].height = 18
                
                # 모든 셀에 테두리 적용
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
            
            # 파일 저장
            wb.save(file_path)
            self.logger.info(f"Formatting applied to {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error applying formatting: {str(e)}", exc_info=True)

    def remove_at_symbol(self, file_path: str) -> str:
        """
        Excel 파일에서 @ 기호 제거
        
        Args:
            file_path: 처리할 파일 경로
            
        Returns:
            처리된 파일 경로
        """
        try:
            if not os.path.exists(file_path):
                self.logger.error(f"File not found: {file_path}")
                return file_path
                
            # DataFrame 로드
            df = pd.read_excel(file_path)
            
            # @ 기호 제거
            replaced = False
            for col in df.columns:
                if df[col].dtype == 'object':  # 문자열 컬럼만 처리
                    original_values = df[col].copy()
                    df[col] = df[col].apply(
                        lambda x: x.replace('@', '') if isinstance(x, str) else x
                    )
                    # 변경 여부 확인
                    if not df[col].equals(original_values):
                        replaced = True
                        self.logger.info(f"Removed @ symbols from column: {col}")
            
            if replaced:
                # 저장
                df.to_excel(file_path, index=False)
                self.logger.info(f"Successfully removed @ symbols from {file_path}")
            else:
                self.logger.info(f"No @ symbols found in {file_path}")
                
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error removing @ symbols: {str(e)}", exc_info=True)
            return file_path

    def _fix_image_formulas(self, file_path: str) -> None:
        """
        Fixes IMAGE formulas in Excel files by ensuring proper format and removing any
        problematic characters.
        """
        try:
            wb = load_workbook(file_path)
            made_changes = False
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Scan all cells for IMAGE formulas
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("=IMAGE("):
                            original = cell.value
                            
                            # Fix common issues in IMAGE formulas
                            fixed = original
                            
                            # 1. Remove @ symbols
                            fixed = fixed.replace('@', '')
                            
                            # 2. Replace backslashes with forward slashes
                            fixed = fixed.replace('\\', '/')
                            
                            # 3. Ensure proper quotation marks
                            if '"' not in fixed:
                                fixed = fixed.replace('=IMAGE(', '=IMAGE("')
                                if fixed.endswith(')'):
                                    fixed = fixed[:-1] + '")'
                                else:
                                    # Add missing closing parenthesis if needed
                                    fixed = fixed + '")'
                            
                            # 4. Fix double quotes
                            fixed = fixed.replace('""', '"')
                            
                            # 5. Ensure proper scaling mode parameter
                            if ',2)' not in fixed and ')' in fixed:
                                fixed = fixed.replace(')', ',2)')
                            
                            # Apply changes if formula was modified
                            if fixed != original:
                                cell.value = fixed
                                made_changes = True
            
            if made_changes:
                wb.save(file_path)
                self.logger.info(f"Image formulas fixed in {file_path}")
                
        except Exception as e:
            self.logger.error(f"Error fixing image formulas: {str(e)}", exc_info=True)
    
    def post_process_excel_file(self, file_path: str) -> str:
        """
        Performs all post-processing steps on an Excel file:
        1. Removes @ symbols
        2. Fixes IMAGE formulas
        3. Other cleanup as needed
        
        Returns the path to the processed file.
        """
        try:
            if not os.path.exists(file_path):
                self.logger.error(f"File not found for post-processing: {file_path}")
                return file_path
            
            # First, remove @ symbols
            cleaned_file = self.remove_at_symbol(file_path)
            
            # Then fix image formulas
            try:
                self._fix_image_formulas(cleaned_file)
            except Exception as img_err:
                self.logger.error(f"Error fixing image formulas: {str(img_err)}")
            
            self.logger.info(f"Post-processing complete for {cleaned_file}")
            return cleaned_file
            
        except Exception as e:
            self.logger.error(f"Error during post-processing: {str(e)}", exc_info=True)
            return file_path 