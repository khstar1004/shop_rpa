import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from utils.excel_utils import insert_image_to_cell, download_image
from ..data_models import Product


class ExcelManager:
    """Excel 파일 읽기, 쓰기, 포맷팅 관련 기능을 담당하는 클래스"""

    def __init__(self, config: Dict, logger: Optional[logging.Logger] = None):
        """
        엑셀 매니저 초기화

        Args:
            config: 애플리케이션 설정
            logger: 로깅 인스턴스
        """
        self.config = config
        self.logger = logger or logging.getLogger(__name__)

        # Extract Excel settings from config or use defaults
        self.excel_settings = config.get("EXCEL", {})
        self._ensure_default_excel_settings()

        # Styling fills for different purposes (enhanced with more colors)
        self.price_difference_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )  # Yellow for price differences
        self.header_fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
        )  # Light yellow for headers
        self.alt_row_fill = PatternFill(
            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
        )  # Light gray for alternate rows
        self.group_header_fill = PatternFill(
            start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
        )  # Lavender for group headers

        # Font styles
        self.bold_font = Font(bold=True)

        # Border styles
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Add border_style property for worksheet creation
        self.border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Alignment
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")

        # Height settings
        self.row_height = 18  # Default row height
        self.image_row_height = 110  # Height for rows with images

    def _ensure_default_excel_settings(self):
        """설정에 필요한 기본값을 설정합니다."""
        defaults = {
            "sheet_name": "Sheet1",
            "start_row": 2,
            "required_columns": [
                "상품명",
                "판매단가(V포함)",
                "상품Code",
                "본사 이미지",
                "본사상품링크",
            ],
            "optional_columns": ["본사 단가", "가격", "상품코드"],
            "max_rows": 10000,
            "enable_formatting": True,
            "date_format": "YYYY-MM-DD",
            "number_format": "#,##0.00",
            "max_file_size_mb": 100,
            "enable_data_quality": True,
            "enable_duplicate_detection": True,
            "enable_auto_correction": True,
            "auto_correction_rules": ["price", "url", "product_code"],
            "report_formatting": True,
            "report_styles": True,
            "report_filters": True,
            "report_sorting": True,
            "report_freeze_panes": True,
            "report_auto_fit": True,
            "validation_rules": {
                "price": {"min": 0, "max": 1000000000},
                "product_code": {"pattern": r"^[A-Za-z0-9-]+$"},
                "url": {"pattern": r"^https?://.*$"},
            },
        }

        # Ensure top-level keys exist
        for key, value in defaults.items():
            if key not in self.excel_settings:
                self.excel_settings[key] = value
            # Ensure nested keys exist (specifically for validation_rules)
            elif key == "validation_rules" and isinstance(value, dict):
                if not isinstance(self.excel_settings[key], dict):
                    self.excel_settings[key] = {}
                for sub_key, sub_value in value.items():
                    if sub_key not in self.excel_settings[key]:
                        self.excel_settings[key][sub_key] = sub_value
                    elif isinstance(sub_value, dict):
                        if not isinstance(self.excel_settings[key][sub_key], dict):
                            self.excel_settings[key][sub_key] = {}
                        for item_key, item_value in sub_value.items():
                            if item_key not in self.excel_settings[key][sub_key]:
                                self.excel_settings[key][sub_key][item_key] = item_value

    def _clean_url(self, url: str) -> str:
        """Normalize and clean a URL string."""
        if not url or not isinstance(url, str):
            return ""
        
        cleaned_url = url.strip()
        
        # Remove potential problematic characters like '@'
        cleaned_url = cleaned_url.replace('@', '')

        # Ensure https protocol
        if cleaned_url.startswith("http:"):
            cleaned_url = "https:" + cleaned_url[5:]
        elif cleaned_url.startswith("//"):
             cleaned_url = "https:" + cleaned_url
        elif not cleaned_url.startswith("https:"):
             # If no protocol, attempt to prepend https://, assuming it's a web URL
             # More robust validation might be needed depending on expected URL types
             if '.' in cleaned_url: # Basic check for domain-like structure
                 cleaned_url = "https://" + cleaned_url
             else:
                 # If it doesn't look like a URL, return original cleaned string or empty
                 return cleaned_url # Or return "" if invalid urls should be blanked

        # Handle backslashes and encode problematic characters for URLs
        cleaned_url = cleaned_url.replace('\\\\', '/') # Corrected backslash replacement
        cleaned_url = cleaned_url.replace('"', '%22')
        cleaned_url = cleaned_url.replace(' ', '%20')

        return cleaned_url

    def _normalize_url(self, url: str) -> str:
        """Normalize and clean URLs for Excel processing. (Deprecated: Use _clean_url)"""
        # Deprecated: Use _clean_url instead
        return self._clean_url(url)

    def _compute_price_metrics(self, base_price, compare_price):
        """Compute price difference and percentage if valid, else return (None, None)."""
        try:
            base_price = float(base_price)
            compare_price = float(compare_price)
        except (ValueError, TypeError):
            return (None, None)
        if base_price > 0 and compare_price > 0 and abs(compare_price - base_price) < base_price * 5:
            diff = compare_price - base_price
            percent = round((diff / base_price) * 100, 2)
            return (diff, percent)
        return (None, None)

    def read_excel_file(self, file_path: str) -> pd.DataFrame:
        """Excel 파일을 읽고 검증하여 DataFrame을 반환합니다."""
        try:
            # 파일 존재 확인
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                self.logger.error(f"Excel file not found: {file_path}")
                return self._create_minimal_error_dataframe("File not found")

            self.logger.info(f"Reading Excel file: {file_path}")

            # 시트 확인
            sheet_name = self.excel_settings["sheet_name"]
            sheet_names = []

            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()  # Close workbook to free resources

                if sheet_name not in sheet_names and sheet_names:
                    self.logger.warning(
                        f"Sheet '{sheet_name}' not found. Using first sheet: '{sheet_names[0]}'"
                    )
                    sheet_name = sheet_names[0]
            except Exception as e:
                self.logger.warning(
                    f"Could not inspect Excel structure: {str(e)}. Using default sheet name."
                )

            # 다양한 방법으로 읽기 시도
            all_dataframes = []

            # 모든 시트 읽기 시도
            try:
                all_sheets = pd.read_excel(
                    file_path, sheet_name=None, engine="openpyxl"
                )
                for sheet, df in all_sheets.items():
                    if not df.empty:
                        self.logger.info(
                            f"Found data in sheet '{sheet}' with {len(df)} rows"
                        )
                        all_dataframes.append((df, sheet, 0))
            except Exception as e:
                self.logger.warning(f"Failed to read all sheets: {str(e)}")

            # 첫 행 스킵 시도 (헤더가 있는 경우)
            if not all_dataframes:
                for skip_rows in range(6):  # 최대 5개 행 스킵 시도
                    try:
                        df = pd.read_excel(
                            file_path,
                            sheet_name=sheet_name,
                            skiprows=skip_rows,
                            engine="openpyxl",
                        )
                        if not df.empty:
                            self.logger.info(f"Found data with skiprows={skip_rows}")
                            all_dataframes.append((df, sheet_name, skip_rows))
                    except Exception as e:
                        self.logger.debug(f"Failed with skiprows={skip_rows}: {str(e)}")

            # 대체 엔진 시도
            if not all_dataframes:
                try:
                    # xlrd 엔진 시도 (오래된 Excel 형식)
                    df = pd.read_excel(file_path, engine="xlrd")
                    if not df.empty:
                        all_dataframes.append((df, "Unknown", 0))
                except Exception:
                    pass

                try:
                    # CSV 형식 시도
                    df = pd.read_csv(file_path)
                    if not df.empty:
                        all_dataframes.append((df, "CSV", 0))
                except Exception:
                    pass

            # 데이터프레임이 없으면 기본 형식 반환
            if not all_dataframes:
                self.logger.error(f"Could not read any valid data from {file_path}")
                return self._create_minimal_error_dataframe("No data found")

            # 가장 많은 행을 가진 데이터프레임 선택
            best_df, sheet, skiprows = max(all_dataframes, key=lambda x: len(x[0]))
            self.logger.info(
                f"Selected dataframe from sheet '{sheet}' with {len(best_df)} rows"
            )

            # 필수 컬럼 확인 및 추가
            best_df = self._ensure_required_columns(best_df)

            return best_df

        except Exception as e:
            self.logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
            return self._create_minimal_error_dataframe(f"Error: {str(e)}")

    def _create_minimal_error_dataframe(self, error_message: str) -> pd.DataFrame:
        """오류 상황에서 기본 데이터프레임을 생성합니다."""
        required_columns = self.excel_settings["required_columns"]
        error_data = {col: [""] for col in required_columns}

        # 첫 번째 행에 오류 메시지 표시
        if "상품명" in required_columns:
            error_data["상품명"] = [f"Error: {error_message}"]

        if "판매단가(V포함)" in required_columns:
            error_data["판매단가(V포함)"] = [0]

        if "상품Code" in required_columns:
            error_data["상품Code"] = ["ERROR"]

        return pd.DataFrame(error_data)

    def _ensure_required_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """데이터프레임에 필수 컬럼이 있는지 확인하고 없으면 추가합니다."""
        df_copy = df.copy()
        required_columns = self.excel_settings["required_columns"]

        # 해오름기프트 엑셀 형식에 맞게 컬럼 형태 표준화
        # 샘플: 구분 담당자 업체명 업체코드 상품Code 중분류카테고리 상품명 기본수량(1) 판매단가(V포함) 본사상품링크
        haeoreum_columns = [
            "구분",
            "담당자",
            "업체명",
            "업체코드",
            "상품Code",
            "중분류카테고리",
            "상품명",
            "기본수량(1)",
            "판매단가(V포함)",
            "본사상품링크",
        ]

        # 컬럼명 공백 제거 및 표준화
        renamed_columns = {}
        for col in df_copy.columns:
            cleaned_col = str(col).strip()
            if cleaned_col != col:
                renamed_columns[col] = cleaned_col

        if renamed_columns:
            df_copy = df_copy.rename(columns=renamed_columns)
            self.logger.info(f"Cleaned column names: {renamed_columns}")

        # 컬럼명 매핑 (해오름 형식에 맞게)
        column_mapping = {
            # 기본 필드 매핑
            "Code": "상품Code",
            "상품코드": "상품Code",
            "상품 코드": "상품Code",
            "상품번호": "상품Code",
            "상품이름": "상품명",
            "상품 이름": "상품명",
            "제품명": "상품명",
            "판매가(V포함)": "판매단가(V포함)",
            "판매가(VAT포함)": "판매단가(V포함)",
            "판매가": "판매단가(V포함)",
            "가격": "판매단가(V포함)",
            "가격(V포함)": "판매단가(V포함)",
            "본사링크": "본사상품링크",
            "상품링크": "본사상품링크",
            "URL": "본사상품링크",
            "수량": "기본수량(1)",
            "기본수량": "기본수량(1)",
            "이미지": "본사 이미지",
            "이미지URL": "본사 이미지",
            "제품이미지": "본사 이미지",
            "상품이미지": "본사 이미지",
        }

        # 해오름 형식에 맞게 컬럼 리매핑
        for src_col, target_col in column_mapping.items():
            if src_col in df_copy.columns and target_col not in df_copy.columns:
                df_copy[target_col] = df_copy[src_col]
                self.logger.info(f"Mapped column '{src_col}' to '{target_col}'")

        # 필수 컬럼 확인 및 생성
        for col in required_columns:
            if col not in df_copy.columns:
                # 유사한 컬럼 찾기
                similar_col = self._find_similar_column(df_copy, col)

                if similar_col:
                    self.logger.info(
                        f"Mapped '{similar_col}' to required column '{col}'"
                    )
                    df_copy[col] = df_copy[similar_col]
                else:
                    self.logger.warning(
                        f"Creating default values for missing column '{col}'"
                    )
                    # 컬럼 유형에 따른 기본값 설정
                    if "단가" in col or "가격" in col:
                        df_copy[col] = 0
                    elif "Code" in col or "코드" in col:
                        df_copy[col] = df_copy.index.map(lambda i: f"GEN-{i}")
                    elif "이미지" in col:
                        df_copy[col] = ""
                    elif "링크" in col:
                        df_copy[col] = ""
                    else:
                        df_copy[col] = ""

        # 소스 컬럼 추가 (해오름기프트 소스 명시)
        if "source" not in df_copy.columns:
            df_copy["source"] = "haeoreum"
            self.logger.info("Added 'source' column with value 'haeoreum'")

        return df_copy

    def _find_similar_column(
        self, df: pd.DataFrame, target_column: str
    ) -> Optional[str]:
        """데이터프레임에서 타겟 컬럼과 유사한 컬럼을 찾습니다."""
        # 한국어 컬럼명 매핑
        column_mapping = {
            "상품명": [
                "품명",
                "제품명",
                "상품",
                "product",
                "name",
                "item",
                "품목",
                "상품이름",
            ],
            "판매단가(V포함)": [
                "단가",
                "판매가",
                "가격",
                "price",
                "가격(v포함)",
                "단가(vat)",
                "판매단가",
            ],
            "상품Code": [
                "코드",
                "code",
                "item code",
                "product code",
                "품목코드",
                "제품코드",
                "상품코드",
            ],
            "본사 이미지": [
                "이미지",
                "image",
                "상품이미지",
                "제품이미지",
                "이미지주소",
                "image url",
            ],
            "본사상품링크": [
                "링크",
                "link",
                "url",
                "상품링크",
                "제품링크",
                "상품url",
                "제품url",
                "홈페이지",
            ],
        }

        # 대소문자 무시, 유사성 검사
        for col in df.columns:
            if col.lower() == target_column.lower():
                return col

        # 매핑 검사
        if target_column in column_mapping:
            for similar in column_mapping[target_column]:
                for col in df.columns:
                    if similar.lower() in col.lower():
                        return col

        return None

    def apply_formatting_to_excel(self, excel_file: str) -> None:
        """엑셀 파일에 포맷팅을 적용합니다."""
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            # 1. Format header row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = self.bold_font
                cell.alignment = self.center_align
                cell.fill = self.header_fill
                cell.border = self.thin_border

            # 2. Format image columns - set wider columns and taller rows
            image_col_indices = []
            status_col_indices = []

            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and ("이미지" in str(header)):
                    image_col_indices.append(col)
                    # Make columns with images wider
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = 20
                elif header and (
                    "매칭_상황" in str(header) or "텍스트유사도" in str(header)
                ):
                    status_col_indices.append(col)
                    # 상태 메시지 컬럼 넓게 설정
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = 30

            # 3. Apply zebra striping and set appropriate row heights
            for row in range(2, ws.max_row + 1):
                # Set row height - taller for rows with images or error messages
                has_image = False
                has_error = False

                for col in image_col_indices:
                    cell_value = ws.cell(row=row, column=col).value
                    if (
                        cell_value
                        and isinstance(cell_value, str)
                        and (
                            cell_value.startswith("http")
                            or cell_value.startswith("=IMAGE")
                        )
                    ):
                        has_image = True
                        break

                for col in status_col_indices:
                    cell_value = ws.cell(row=row, column=col).value
                    if (
                        cell_value
                        and isinstance(cell_value, str)
                        and len(cell_value) > 10
                    ):  # 긴 메시지가 있는 경우
                        has_error = True
                        break

                if has_image:
                    ws.row_dimensions[row].height = self.image_row_height
                elif has_error:
                    ws.row_dimensions[row].height = 60  # 오류 메시지를 위한 높이
                else:
                    ws.row_dimensions[row].height = self.row_height

                # Apply zebra striping (alternate row coloring)
                if row % 2 == 0:  # Even rows get light coloring
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if (
                            not cell.fill or cell.fill.start_color.index == "FFFFFF"
                        ):  # Only if not already colored
                            cell.fill = self.alt_row_fill

                # Apply border to all cells
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.thin_border

                    # 오류 메시지에 노란색 폰트 적용
                    if col in status_col_indices:
                        cell_value = cell.value
                        if (
                            cell_value
                            and isinstance(cell_value, str)
                            and "상품이 없음" in cell_value
                        ):
                            cell.font = Font(color="FF0000", bold=True)
                            cell.alignment = Alignment(
                                wrap_text=True, vertical="center"
                            )

            # 4. Find price difference columns and apply formatting
            price_diff_cols = []
            price_percent_cols = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    if "가격차이" in str(header) and "%" not in str(header):
                        price_diff_cols.append(col)
                    elif "가격차이" in str(header) and "%" in str(header):
                        price_percent_cols.append(col)

            # 5. Apply highlighting for negative price differences
            for row in range(2, ws.max_row + 1):
                should_highlight_row = False

                # Check if any price difference column has negative value
                for col in price_diff_cols:
                    cell = ws.cell(row=row, column=col)
                    try:
                        value = float(cell.value) if cell.value is not None else 0
                        if value < 0:
                            cell.fill = self.price_difference_fill
                            should_highlight_row = True  # Mark for row highlighting
                    except:
                        pass  # 숫자가 아닌 셀은 무시

                # Apply number formatting to percentage columns
                for col in price_percent_cols:
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0%"
                    elif isinstance(cell.value, str):
                        # Try to convert string percentage to number format
                        try:
                            if cell.value.endswith("%"):
                                value = float(cell.value.rstrip("%")) / 100
                                cell.value = value
                                cell.number_format = "0.0%"
                        except:
                            pass

                # 오류 메시지가 있는 텍스트 셀 포맷팅
                for col in status_col_indices:
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    if value and isinstance(value, str):
                        if (
                            "일정 정확도" in value
                            or "범위 내에 없음" in value
                            or "찾을 수 없음" in value
                        ):
                            cell.font = Font(color="FF0000", bold=True)
                            cell.alignment = Alignment(
                                wrap_text=True, vertical="center"
                            )

            # 6. Adjust column widths based on content
            for col in range(1, ws.max_column + 1):
                if col in status_col_indices:
                    # 상태 메시지 컬럼은 이미 설정했으므로 건너뜀
                    continue

                if col in image_col_indices:
                    # 이미지 컬럼은 이미 설정했으므로 건너뜀
                    continue

                max_length = 0
                column = get_column_letter(col)

                # Check all values in the column
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass

                # Adjust column width based on content
                adjusted_width = min(max_length + 4, 50)  # Cap width at 50
                if "상품명" in str(ws.cell(row=1, column=col).value):
                    adjusted_width = max(
                        adjusted_width, 35
                    )  # Product names need more space
                elif "링크" in str(ws.cell(row=1, column=col).value):
                    adjusted_width = max(adjusted_width, 25)  # Links need more space
                else:
                    adjusted_width = max(
                        adjusted_width, 12
                    )  # Minimum width for readability

                ws.column_dimensions[column].width = adjusted_width

            # 7. Freeze panes at A2 (keep header visible while scrolling)
            ws.freeze_panes = "A2"

            # 8. Save the formatted workbook
            wb.save(excel_file)
            self.logger.info(f"Enhanced formatting applied to {excel_file}")

        except Exception as e:
            self.logger.error(
                f"Error applying enhanced formatting: {str(e)}", exc_info=True
            )

    def generate_enhanced_output(self, results: List, input_file: str, output_dir: Optional[str] = None) -> str:
        """처리 결과를 엑셀로 저장하고 포맷팅을 적용합니다."""
        try:
            # 결과 데이터 준비
            report_data = []
            processed_count = 0

            for result in results:
                if hasattr(result, "source_product") and result.source_product:
                    processed_count += 1
                    row = {}
                    
                    # <<< Logging >>> Check the structure of the result object
                    self.logger.debug(f"Processing result object: {result}")
                    if hasattr(result, "best_koryo_match"):
                        self.logger.debug(f"  -> Found best_koryo_match: {result.best_koryo_match}")
                    else:
                        self.logger.debug("  -> best_koryo_match attribute NOT found.")
                    if hasattr(result, "best_naver_match"):
                        self.logger.debug(f"  -> Found best_naver_match: {result.best_naver_match}")
                    elif hasattr(result, "naver_matches") and result.naver_matches:
                         self.logger.debug(f"  -> Found naver_matches (using first): {result.naver_matches[0]}")
                    else:
                        self.logger.debug("  -> best_naver_match / naver_matches attributes NOT found.")
                    # <<< End Logging >>>

                    # 원본 제품 데이터 추출
                    source_data = result.source_product.original_input_data
                    if isinstance(source_data, dict):
                        for key, value in source_data.items():
                            row[key] = value

                    # 필수 필드 확인
                    self._ensure_required_fields(row, result)

                    # 중요 필드가 비어있지 않은지 재확인
                    self._validate_critical_fields(row, result)

                    # 고려기프트 매칭 데이터 추가
                    if hasattr(result, "best_koryo_match") and result.best_koryo_match:
                        self._add_koryo_match_data(row, result.best_koryo_match)

                    # 네이버 매칭 데이터 추가
                    if hasattr(result, "best_naver_match") and result.best_naver_match:
                        self._add_naver_match_data(row, result.best_naver_match)
                    elif hasattr(result, "naver_matches") and result.naver_matches:
                        # 최적 매칭이 없지만 다른 후보가 있는 경우 첫 번째 후보 사용
                        self._add_naver_match_data(row, result.naver_matches[0])
                        self.logger.info(
                            f"최적 네이버 매칭이 없어 첫 번째 후보({result.naver_matches[0].matched_product.name})를 사용합니다."
                        )

                    # 빈 필드에 기본값 설정
                    self._set_default_values(row)

                    # URL에서 @ 기호를 제거하여 Excel 저장 오류 방지 (모든 필드에 적용)
                    for key, value in row.items():
                        if isinstance(value, str) and '@' in value:
                            row[key] = value.replace('@', '')
                    
                    report_data.append(row)
                    
                    # <<< Logging >>> Check the row added to report_data
                    self.logger.debug(f"  -> Row added to report_data: {row}")
                    # <<< End Logging >>>

            # 로깅 - 처리된 결과 수
            self.logger.info(
                f"총 {processed_count}개 제품 처리됨, 엑셀 파일에 {len(report_data)}행 작성"
            )

            # 결과 데이터가 비어있는지 확인
            if not report_data:
                self.logger.warning(
                    "엑셀 보고서에 작성할 데이터 없음! 기본 데이터 생성"
                )
                # 최소한의 데이터 생성
                empty_row = {
                    "상품명": "데이터 처리 중 오류 발생",
                    "판매단가(V포함)": 0,
                    "상품Code": "ERROR",
                    "구분": "ERROR",
                }
                report_data.append(empty_row)

            # DataFrame 생성
            result_df = pd.DataFrame(report_data)

            # 가격 차이 데이터 정확하게 계산하고 기록 (필요한 경우 다시 계산)
            self._calculate_price_differences(result_df)

            # 컬럼 순서 정렬
            result_df = self._reorder_columns(result_df)

            # 파일명 생성
            base_name = os.path.splitext(input_file)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 출력 디렉토리 처리: 지정된 디렉토리가 없으면 'output' 디렉토리 사용
            # output_dir이 지정된 경우 해당 디렉토리 사용
            if output_dir:
                # 출력 디렉토리가 존재하는지 확인하고 없으면 생성
                os.makedirs(output_dir, exist_ok=True)
                output_file = os.path.join(output_dir, f'{os.path.basename(base_name)}-result_result_{timestamp}_upload_{timestamp}.xlsx')
            else:
                # 기본 output 디렉토리 생성 확인
                os.makedirs('output', exist_ok=True)
                output_file = os.path.join('output', f'{os.path.basename(base_name)}-result_result_{timestamp}_upload_{timestamp}.xlsx')

            # 엑셀로 저장 전 데이터 클리닝 최종 확인
            # IMAGE 함수에서 URL에 @ 기호가 포함되면 오류가 발생할 수 있으므로 제거
            for col in result_df.columns:
                if col in ['본사 이미지', '고려기프트 이미지', '네이버 이미지']:
                    result_df[col] = result_df[col].apply(lambda x: str(x).replace('@', '') if isinstance(x, str) else x)

            # 엑셀로 저장
            self.logger.info(f"Saving Excel file to: {output_file}")
            result_df.to_excel(output_file, index=False)
            self.logger.info(f"Excel file saved successfully")

            # 후처리 적용 (@ 기호 제거, 포맷팅, 하이퍼링크 등)
            try:
                output_file = self.post_process_excel_file(output_file)
            except Exception as post_err:
                self.logger.error(f"Error in post-processing: {str(post_err)}")
                # 포맷팅 직접 적용 시도
                try:
                    self.apply_formatting_to_excel(output_file)
                except Exception as format_err:
                    self.logger.error(f"Formatting also failed: {str(format_err)}")

            self.logger.info(
                f"결과 파일 생성 완료: {output_file} (총 {len(report_data)}행)"
            )

            return output_file

        except Exception as e:
            self.logger.error(f"결과 파일 생성 중 오류 발생: {str(e)}", exc_info=True)

            # 오류가 발생해도 파일은 생성 시도
            try:
                # 기본 데이터로 빈 파일 생성
                error_data = [
                    {
                        "오류": str(e),
                        "상품명": "데이터 처리 중 오류 발생",
                        "시간": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    }
                ]

                error_df = pd.DataFrame(error_data)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # output_dir이 지정된 경우 해당 디렉토리 사용, 없으면 생성
                if output_dir:
                    os.makedirs(output_dir, exist_ok=True)
                    error_output_file = os.path.join(output_dir, f'error-result_{timestamp}.xlsx')
                else:
                    os.makedirs('output', exist_ok=True)
                    error_output_file = os.path.join('output', f'error-result_{timestamp}.xlsx')
                
                # 에러 파일 저장
                error_df.to_excel(error_output_file, index=False)

                self.logger.warning(
                    f"오류 정보가 포함된 파일을 생성했습니다: {error_output_file}"
                )
                return error_output_file

            except Exception as inner_e:
                self.logger.critical(
                    f"오류 파일 생성 중 추가 오류 발생: {str(inner_e)}"
                )
                return ""

    def _calculate_price_differences(self, df: pd.DataFrame) -> None:
        """가격 차이 및 백분율 계산을 확인하고 필요한 경우 다시 계산"""
        try:
            # 1. 고려기프트 가격 차이 계산
            if "판매단가(V포함)" in df.columns and "판매단가(V포함)(2)" in df.columns:
                # 기존 열이 있으면 삭제
                if "가격차이(2)" in df.columns:
                    del df["가격차이(2)"]
                if "가격차이(2)(%)" in df.columns:
                    del df["가격차이(2)(%)"]

                # 가격 차이 계산 (고려기프트 가격 - 본사 가격)
                df["가격차이(2)"], df["가격차이(2)(%)"] = zip(*df.apply(
                    lambda row: self._compute_price_metrics(row.get("판매단가(V포함)"), row.get("판매단가(V포함)(2)")), axis=1))

            # 2. 네이버 가격 차이 계산
            if "판매단가(V포함)" in df.columns and "판매단가(V포함)(3)" in df.columns:
                # 기존 열이 있으면 삭제
                if "가격차이(3)" in df.columns:
                    del df["가격차이(3)"]
                if "가격차이(3)(%)" in df.columns:
                    del df["가격차이(3)(%)"]

                # 가격 차이 계산 (네이버 가격 - 본사 가격)
                df["가격차이(3)"], df["가격차이(3)(%)"] = zip(*df.apply(
                    lambda row: self._compute_price_metrics(row.get("판매단가(V포함)"), row.get("판매단가(V포함)(3)")), axis=1))
        except Exception as e:
            self.logger.error(f"가격 차이 계산 중 오류 발생: {str(e)}", exc_info=True)

    def _reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """컬럼 순서를 예시와 일치하도록 정렬합니다."""
        # 참고 컬럼 순서 (예시 엑셀과 일치)
        column_order = [
            "구분",
            "담당자",
            "업체명",
            "업체코드",
            "Code",
            "중분류카테고리",
            "상품명",
            "기본수량(1)",
            "판매단가(V포함)",
            "본사상품링크",
            "기본수량(2)",
            "판매단가(V포함)(2)",
            "판매가(V포함)(2)",
            "가격차이(2)",
            "가격차이(2)(%)",
            "매칭_상황(2)",
            "텍스트유사도(2)",
            "고려기프트 상품링크",
            "기본수량(3)",
            "판매단가(V포함)(3)",
            "가격차이(3)",
            "가격차이(3)(%)",
            "매칭_상황(3)",
            "텍스트유사도(3)",
            "공급사명",
            "네이버 쇼핑 링크",
            "공급사 상품링크",
            "본사 이미지",
            "고려기프트 이미지",
            "네이버 이미지",
        ]

        # 실제 존재하는 컬럼만 정렬에 포함
        existing_columns = [col for col in column_order if col in df.columns]

        # 누락된 컬럼은 원래 위치에 보존
        missing_columns = [col for col in df.columns if col not in column_order]

        # 누락된 컬럼이 있을 경우 기본 순서 뒤에 추가
        ordered_columns = existing_columns + missing_columns

        # 재정렬된 DataFrame 반환
        return df[ordered_columns]

    def _ensure_required_fields(self, row: Dict, result: object) -> None:
        """필수 필드가 존재하는지 확인하고 없으면 생성"""
        # 기본 필드 확인
        required_fields = [
            "구분",
            "담당자",
            "업체명",
            "업체코드",
            "Code",
            "상품Code",
            "중분류카테고리",
            "상품명",
            "기본수량(1)",
            "판매단가(V포함)",
            "본사상품링크",
        ]

        for field in required_fields:
            if field not in row:
                # Code와 상품Code는 상호 대체 가능
                if field == "Code" and "상품Code" in row:
                    row["Code"] = row["상품Code"]
                elif field == "상품Code" and "Code" in row:
                    row["상품Code"] = row["Code"]
                else:
                    row[field] = ""

        # 이미지 URL 처리 함수
        def process_image_url(url):
            cleaned_url = self._clean_url(url) # Use centralized cleaner
            if not cleaned_url:
                return ""
            try:
                # HTTPS로 시작하는 경우에만 IMAGE 함수 적용
                if cleaned_url.startswith('https:'):
                    # 이미지 크기 제한을 위한 파라미터 추가
                    return f'=IMAGE("{cleaned_url}", 2)'
                return cleaned_url # Return cleaned URL if not https (might be local path etc.)
            except Exception as e:
                self.logger.warning(f"이미지 URL 처리 중 오류 발생: {str(e)}")
                return "" # Return empty on error

        # 기본 이미지 URL 설정
        if "본사 이미지" not in row and hasattr(result.source_product, "image_url") and result.source_product.image_url:
            row["본사 이미지"] = process_image_url(result.source_product.image_url)
        elif "본사 이미지" in row and isinstance(row["본사 이미지"], str) and row["본사 이미지"].startswith('http'):
            row["본사 이미지"] = process_image_url(row["본사 이미지"])
        # Make sure existing image urls are also cleaned if they are just strings
        elif "본사 이미지" in row and isinstance(row["본사 이미지"], str):
             cleaned_existing_url = self._clean_url(row["본사 이미지"])
             if cleaned_existing_url.startswith('https:'):
                 row["본사 이미지"] = f'=IMAGE("{cleaned_existing_url}", 2)'
             else:
                 row["본사 이미지"] = cleaned_existing_url

        # 기본 URL 설정
        if "본사상품링크" not in row and hasattr(result.source_product, "url"):
            row["본사상품링크"] = self._clean_url(result.source_product.url) # Use cleaner
        elif "본사상품링크" in row and isinstance(row["본사상품링크"], str):
             row["본사상품링크"] = self._clean_url(row["본사상품링크"]) # Clean existing links too

    def _validate_critical_fields(self, row: Dict, result: object) -> None:
        """중요 필드 검증 및 수정"""
        # 상품명 필드 확인
        if not row.get("상품명") and hasattr(result.source_product, "name"):
            row["상품명"] = result.source_product.name

        # 가격 필드 확인
        if not row.get("판매단가(V포함)") and hasattr(result.source_product, "price"):
            row["판매단가(V포함)"] = result.source_product.price

        # 상품코드 확인
        if (
            not row.get("상품Code")
            and not row.get("Code")
            and hasattr(result.source_product, "id")
        ):
            row["상품Code"] = result.source_product.id

    def _add_koryo_match_data(self, row: Dict, koryo_match: object) -> None:
        """고려기프트 매칭 데이터 추가"""
        try:
            if hasattr(koryo_match, "matched_product"):
                match_product = koryo_match.matched_product

                # 매칭 성공 여부 확인 (유사도 기준을 0.85로 상향)
                match_success = getattr(
                    koryo_match, "text_similarity", 0
                ) >= self.config.get("MATCHING", {}).get("TEXT_SIMILARITY_THRESHOLD", 0.85)

                # 카테고리 매칭 확인
                if hasattr(match_product, "category") and hasattr(koryo_match.source_product, "category"):
                    category_match = match_product.category == koryo_match.source_product.category
                    if not category_match:
                        match_success = False

                # 브랜드 매칭 확인
                if hasattr(match_product, "brand") and hasattr(koryo_match.source_product, "brand"):
                    brand_match = match_product.brand == koryo_match.source_product.brand
                    if not brand_match:
                        match_success = False

                # 가격 정보
                if hasattr(match_product, "price"):
                    price_in_range = self._is_price_in_range(match_product.price)
                    if not price_in_range:
                        match_success = False
                    row["판매단가(V포함)(2)"] = match_product.price

                # 가격 차이 정보
                if hasattr(koryo_match, "price_difference"):
                    row["가격차이(2)"] = koryo_match.price_difference

                if hasattr(koryo_match, "price_difference_percent"):
                    row["가격차이(2)(%)"] = koryo_match.price_difference_percent

                # 텍스트 유사도
                if hasattr(koryo_match, "text_similarity"):
                    row["텍스트유사도(2)"] = round(koryo_match.text_similarity, 2)  # 소수점 2자리까지만 표시

                # 이미지 및 링크
                if hasattr(match_product, "image_url") and match_product.image_url:
                    image_url = self._clean_url(match_product.image_url) # Use cleaner
                    if image_url.startswith('https:'):
                        # 이미지 크기 제한을 위한 파라미터 추가
                        row["고려기프트 이미지"] = f'=IMAGE("{image_url}", 2)'
                    else:
                        row["고려기프트 이미지"] = image_url
                else:
                    # 이미지 URL이 없는 경우 메시지 설정
                    row["고려기프트 이미지"] = "이미지를 찾을 수 없음"

                if hasattr(match_product, "url") and match_product.url:
                    row["고려기프트 상품링크"] = self._clean_url(match_product.url) # Use cleaner
                else:
                    # 링크가 없는 경우 메시지 설정
                    row["고려기프트 상품링크"] = "상품 링크를 찾을 수 없음"

                # 매칭 실패 시 메시지 설정
                if not match_success:
                    text_similarity = getattr(koryo_match, "text_similarity", 0)
                    category_match = getattr(koryo_match, "category_match", True)
                    brand_match = getattr(koryo_match, "brand_match", True)
                    price_in_range = getattr(koryo_match, "price_in_range", True)
                    
                    error_messages = []
                    if text_similarity < 0.85:
                        error_messages.append(f"유사도: {text_similarity:.2f}")
                    if not category_match:
                        error_messages.append("카테고리 불일치")
                    if not brand_match:
                        error_messages.append("브랜드 불일치")
                    if not price_in_range:
                        error_messages.append("가격 범위 초과")
                    
                    row["매칭_상황(2)"] = f"매칭 실패 ({', '.join(error_messages)})"
            else:
                # 매칭된 상품이 없는 경우
                row["매칭_상황(2)"] = "고려기프트에서 상품을 찾지 못했습니다"
                row["판매단가(V포함)(2)"] = 0
                row["가격차이(2)"] = 0
                row["가격차이(2)(%)"] = 0
                row["고려기프트 이미지"] = "상품을 찾을 수 없음"
                row["고려기프트 상품링크"] = "상품을 찾을 수 없음"
        except Exception as e:
            self.logger.error(f"고려기프트 매칭 데이터 추가 중 오류 발생: {str(e)}", exc_info=True)
            row["매칭_상황(2)"] = f"오류 발생: {str(e)}"

    def _add_naver_match_data(self, row: Dict, naver_match: object) -> None:
        """네이버 매칭 데이터 추가"""
        try:
            if hasattr(naver_match, "matched_product"):
                match_product = naver_match.matched_product

                # 매칭 성공 여부 확인 (유사도 기준을 0.85로 상향)
                match_success = getattr(
                    naver_match, "text_similarity", 0
                ) >= self.config.get("MATCHING", {}).get("TEXT_SIMILARITY_THRESHOLD", 0.85)

                # 카테고리 매칭 확인
                if hasattr(match_product, "category") and hasattr(naver_match.source_product, "category"):
                    category_match = match_product.category == naver_match.source_product.category
                    if not category_match:
                        match_success = False

                # 브랜드 매칭 확인
                if hasattr(match_product, "brand") and hasattr(naver_match.source_product, "brand"):
                    brand_match = match_product.brand == naver_match.source_product.brand
                    if not brand_match:
                        match_success = False

                # 공급사 정보
                if hasattr(match_product, "brand") and match_product.brand:
                    row["공급사명"] = match_product.brand
                else:
                    row["공급사명"] = "정보 없음"

                # 가격 정보
                if hasattr(match_product, "price"):
                    price_in_range = self._is_price_in_range(match_product.price)
                    if not price_in_range:
                        match_success = False
                    row["판매단가(V포함)(3)"] = match_product.price

                # 가격 차이 정보
                if hasattr(naver_match, "price_difference"):
                    row["가격차이(3)"] = naver_match.price_difference

                if hasattr(naver_match, "price_difference_percent"):
                    row["가격차이(3)(%)"] = naver_match.price_difference_percent

                # 텍스트 유사도
                if hasattr(naver_match, "text_similarity"):
                    row["텍스트유사도(3)"] = round(naver_match.text_similarity, 2)  # 소수점 2자리까지만 표시

                # 이미지 및 링크
                if hasattr(match_product, "image_url") and match_product.image_url:
                    image_url = self._clean_url(match_product.image_url) # Use cleaner
                    if image_url.startswith('https:'):
                        # 이미지 크기 제한을 위한 파라미터 추가
                        row["네이버 이미지"] = f'=IMAGE("{image_url}", 2)'
                    else:
                        row["네이버 이미지"] = image_url
                else:
                    # 이미지 URL이 없는 경우 메시지 설정
                    row["네이버 이미지"] = "이미지를 찾을 수 없음"

                if hasattr(match_product, "url") and match_product.url:
                    row["네이버 쇼핑 링크"] = self._clean_url(match_product.url) # Use cleaner
                    row["공급사 상품링크"] = self._clean_url(match_product.url) # Use cleaner
                else:
                    # 링크가 없는 경우 메시지 설정
                    row["네이버 쇼핑 링크"] = "상품 링크를 찾을 수 없음"
                    row["공급사 상품링크"] = "상품 링크를 찾을 수 없음"

                # 매칭 실패 시 메시지 설정
                if not match_success and not row.get("매칭_상황(3)"):
                    text_similarity = getattr(naver_match, "text_similarity", 0)
                    category_match = getattr(naver_match, "category_match", True)
                    brand_match = getattr(naver_match, "brand_match", True)
                    price_in_range = getattr(naver_match, "price_in_range", True)
                    
                    error_messages = []
                    if text_similarity < 0.85:
                        error_messages.append(f"유사도: {text_similarity:.2f}")
                    if not category_match:
                        error_messages.append("카테고리 불일치")
                    if not brand_match:
                        error_messages.append("브랜드 불일치")
                    if not price_in_range:
                        error_messages.append("가격 범위 초과")
                    
                    row["매칭_상황(3)"] = f"매칭 실패 ({', '.join(error_messages)})"
            else:
                # 매칭된 상품이 없는 경우
                row["매칭_상황(3)"] = "네이버에서 상품을 찾지 못했습니다"
                row["판매단가(V포함)(3)"] = 0
                row["가격차이(3)"] = 0
                row["가격차이(3)(%)"] = 0
                row["공급사명"] = "상품을 찾을 수 없음"
                row["네이버 이미지"] = "상품을 찾을 수 없음"
                row["네이버 쇼핑 링크"] = "상품을 찾을 수 없음"
                row["공급사 상품링크"] = "상품을 찾을 수 없음"
        except Exception as e:
            self.logger.error(f"네이버 매칭 데이터 추가 중 오류 발생: {str(e)}", exc_info=True)
            row["매칭_상황(3)"] = f"오류 발생: {str(e)}"

    def _is_price_in_range(self, price) -> bool:
        """Check if price is within the valid range."""
        try:
            price_value = float(price)
            min_price = float(
                self.excel_settings.get("validation_rules", {})
                .get("price", {})
                .get("min", 0)
            )
            max_price = float(
                self.excel_settings.get("validation_rules", {})
                .get("price", {})
                .get("max", 1000000000)
            )
            return min_price <= price_value <= max_price
        except (ValueError, TypeError):
            return False

    def _set_default_values(self, row: Dict) -> None:
        """Set default values for empty fields in the row dictionary."""
        for field in ["상품명", "상품Code", "판매단가(V포함)", "본사상품링크"]:
            if field not in row or pd.isna(row[field]) or row[field] == "":
                if field == "판매단가(V포함)":
                    row[field] = 0
                elif field == "상품Code":
                    row[field] = "DEFAULT-CODE"
                else:
                    row[field] = ""

    def check_excel_file(self, file_path: str) -> None:
        """
        엑셀 파일의 필요한 컬럼이 있는지 확인하고 없으면 추가합니다.

        Args:
            file_path: 엑셀 파일 경로
        """
        try:
            self.logger.info(f"Checking Excel file: {file_path}")

            # 파일 존재 확인
            if not os.path.exists(file_path):
                self.logger.error(f"Excel file not found: {file_path}")
                return

            # 엑셀 파일 읽기
            df = pd.read_excel(file_path)

            # 추가할 컬럼 정의
            columns_to_add = ["본사 이미지", "고려기프트 이미지", "네이버 이미지"]
            need_to_modify = False

            # 컬럼 확인 및 추가
            for column in columns_to_add:
                if column not in df.columns:
                    df[column] = ""
                    need_to_modify = True
                    self.logger.info(f"Added missing column: {column}")

            if not need_to_modify:
                self.logger.info("All required columns exist. No modifications needed.")
                return

            # 컬럼명의 앞뒤 공백 제거
            df.columns = [col.strip() for col in df.columns]

            # 엑셀 파일 저장
            df.to_excel(file_path, index=False)
            self.logger.info(f"Updated Excel file with required columns: {file_path}")

        except Exception as e:
            self.logger.error(f"Error checking Excel file: {str(e)}", exc_info=True)

    def convert_xls_to_xlsx(self, input_directory: str) -> str:
        """
        XLS 파일을 XLSX 형식으로 변환합니다.

        Args:
            input_directory: 입력 디렉토리 경로

        Returns:
            str: 변환된 XLSX 파일 경로 또는 빈 문자열
        """
        try:
            self.logger.info(f"Looking for XLS files in: {input_directory}")

            # XLS 파일 찾기
            xls_files = [
                f for f in os.listdir(input_directory) if f.lower().endswith(".xls")
            ]

            if not xls_files:
                self.logger.warning("No XLS files found in the directory.")
                return ""

            # 첫 번째 XLS 파일 처리
            file_name = xls_files[0]
            file_path = os.path.join(input_directory, file_name)

            self.logger.info(f"Converting XLS file: {file_path}")

            # XLS 파일 로드
            tables = pd.read_html(file_path, encoding="cp949")
            df = tables[0]

            # 첫 번째 행을 헤더로 사용
            df.columns = df.iloc[0].str.strip()
            df = df.drop(0)

            # 상품명 전처리 (// 이후 제거 및 특수 문자 정리)
            pattern = r"(\d{4}_[A-Z]\.)|(\d+\+\d+)|[^a-zA-Z0-9가-힣\s]|\s+"

            def preprocess_product_name(product_name):
                if not isinstance(product_name, str):
                    return product_name

                # // 이후 제거
                if "//" in product_name:
                    product_name = product_name.split("//")[0]

                # 특수 문자 및 패턴 제거
                product_name = re.sub(pattern, " ", product_name)
                product_name = (
                    product_name.replace("정품", "")
                    .replace("NEW", "")
                    .replace("특가", "")
                )
                product_name = product_name.replace("주문제작타올", "").replace(
                    "주문제작수건", ""
                )
                product_name = product_name.replace("결혼답례품 수건", "").replace(
                    "답례품수건", ""
                )
                product_name = product_name.replace("주문제작 수건", "").replace(
                    "돌답례품수건", ""
                )
                product_name = (
                    product_name.replace("명절선물세트", "")
                    .replace("각종행사수건", "")
                    .strip()
                )
                product_name = re.sub(" +", " ", product_name)
                return product_name

            df["상품명"] = df["상품명"].apply(preprocess_product_name)

            # 필요한 컬럼 추가
            for column in ["본사 이미지", "고려기프트 이미지", "네이버 이미지"]:
                if column not in df.columns:
                    df[column] = ""

            # 출력 파일명 설정
            output_file_name = file_name.replace(".xls", ".xlsx")
            output_file_path = os.path.join(input_directory, output_file_name)

            # XLSX로 저장
            df.to_excel(output_file_path, index=False)
            self.logger.info(f"Converted file saved to: {output_file_path}")

            return output_file_path

        except Exception as e:
            self.logger.error(f"Error converting XLS to XLSX: {str(e)}", exc_info=True)
            return ""

    def add_hyperlinks_to_excel(self, file_path: str) -> str:
        """
        엑셀 파일의 URL 필드를 하이퍼링크로 변환합니다.

        Args:
            file_path: 입력 엑셀 파일 경로

        Returns:
            str: 처리된 파일 경로
        """
        try:
            self.logger.info(f"Adding hyperlinks to Excel file: {file_path}")

            # 출력 파일명 생성 제거 (기존 파일 수정)
            # now = datetime.now()
            # timestamp = now.strftime("%Y%m%d_%H%M%S")
            # output_directory = os.path.dirname(file_path)
            # base_name = os.path.splitext(os.path.basename(file_path))[0]
            # output_filename = f"{base_name}_result_{timestamp}.xlsx"
            # output_file_path = os.path.join(output_directory, output_filename)

            # 워크북 로드 (기존 파일 사용)
            try:
                df = pd.read_excel(file_path)
                wb = load_workbook(file_path)
                ws = wb.active
            except FileNotFoundError:
                self.logger.error(f"File not found for adding hyperlinks: {file_path}")
                return file_path
            except Exception as load_err:
                 self.logger.error(f"Error loading workbook for hyperlinks: {str(load_err)}")
                 return file_path

            # 테두리 스타일 정의
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 모든 열의 너비 설정 및 자동 줄 바꿈, 테두리 적용
            columns = [get_column_letter(i) for i in range(1, ws.max_column + 1)]
            for col in columns:
                ws.column_dimensions[col].width = 16

                for row in range(1, ws.max_row + 1): # 헤더 포함 모든 행
                    cell = ws[f"{col}{row}"]
                    cell.alignment = Alignment(wrap_text=True, vertical="center") # 중앙 정렬 추가
                    cell.border = thin_border

            # 링크 컬럼 처리
            link_columns = [
                "본사상품링크",
                "고려기프트 상품링크",
                "네이버 쇼핑 링크",
                "공급사 상품링크",
            ]
            for link_column in link_columns:
                if link_column in df.columns:
                    col_idx = df.columns.get_loc(link_column) + 1
                    for row_idx, link in enumerate(
                        df[link_column], start=2
                    ):  # 첫 번째 행은 헤더
                        if (
                            pd.notna(link)
                            and isinstance(link, str)
                        ):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cleaned_link = self._clean_url(link)
                            cell.value = cleaned_link # Display cleaned link text
                            if cleaned_link.startswith("https"):
                                cell.hyperlink = cleaned_link
                                cell.font = Font(color="0000FF", underline="single") # Style as hyperlink
                            else:
                                cell.hyperlink = None

            # 헤더 색상 적용
            gray_fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            for cell in ws["1:1"]:
                cell.fill = gray_fill

            # 가격차이가 음수인 행에 노란색 배경 적용
            yellow_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            for row_idx in range(2, df.shape[0] + 2):
                for col_name in ["가격차이(2)", "가격차이(3)"]:
                    if col_name in df.columns:
                        try:
                            value = df.loc[row_idx - 2, col_name]
                            if pd.notna(value):
                                value = float(value)
                                if value < 0:
                                    for col in ws.iter_cols(
                                        min_row=row_idx,
                                        max_row=row_idx,
                                        min_col=1,
                                        max_col=ws.max_column,
                                    ):
                                        for cell in col:
                                            cell.fill = yellow_fill
                                    break
                        except (ValueError, TypeError):
                            continue

            # 결과 저장 (기존 파일 덮어쓰기)
            wb.save(file_path)
            self.logger.info(f"Excel file with hyperlinks updated: {file_path}")

            return file_path

        except Exception as e:
            self.logger.error(
                f"Error adding hyperlinks to Excel: {str(e)}", exc_info=True
            )
            return file_path

    def filter_excel_by_price_diff(self, file_path: str) -> str:
        """
        가격차이가 있는 항목들만 필터링하여 업로드용 엑셀 파일을 생성합니다.

        Args:
            file_path: 입력 엑셀 파일 경로

        Returns:
            str: 필터링된 출력 파일 경로
        """
        try:
            self.logger.info(f"Filtering Excel file by price differences: {file_path}")

            # 출력 파일명 생성
            input_filename = os.path.splitext(os.path.basename(file_path))[0]
            current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{input_filename}_upload_{current_datetime}.xlsx"
            output_path = os.path.join('output', output_filename)

            # 데이터프레임 로드
            df = pd.read_excel(file_path)

            # 문자열을 숫자로 변환
            def to_float(value):
                try:
                    return float(value)
                except (ValueError, TypeError):
                    return None

            if "가격차이(2)" in df.columns:
                df["가격차이(2)"] = df["가격차이(2)"].apply(to_float)
            if "가격차이(3)" in df.columns:
                df["가격차이(3)"] = df["가격차이(3)"].apply(to_float)

            # 가격차이가 음수인 항목만 필터링
            price_diff_condition = False
            if "가격차이(2)" in df.columns:
                price_diff_condition |= df["가격차이(2)"].notna() & (
                    df["가격차이(2)"] < 0
                )
            if "가격차이(3)" in df.columns:
                price_diff_condition |= df["가격차이(3)"].notna() & (
                    df["가격차이(3)"] < 0
                )

            if isinstance(price_diff_condition, bool) and not price_diff_condition:
                self.logger.warning(
                    "No price difference columns found or no conditions specified."
                )
                return file_path

            filtered_df = df[price_diff_condition]

            if filtered_df.empty:
                self.logger.warning("No items with negative price differences found.")
                return file_path

            # 원하는 컬럼만 선택
            required_columns = [
                "구분",
                "담당자",
                "업체명",
                "업체코드",
                "Code",
                "중분류카테고리",
                "상품명",
                "기본수량(1)",
                "판매단가(V포함)",
                "본사상품링크",
                "기본수량(2)",
                "판매단가(V포함)(2)",
                "가격차이(2)",
                "가격차이(2)(%)",
                "고려기프트 상품링크",
                "기본수량(3)",
                "판매단가(V포함)(3)",
                "가격차이(3)",
                "가격차이(3)(%)",
                "공급사명",
                "공급사 상품링크",
                "본사 이미지",
                "고려기프트 이미지",
                "네이버 이미지",
            ]

            # 존재하는 컬럼만 선택
            existing_columns = [
                col for col in required_columns if col in filtered_df.columns
            ]
            filtered_df = filtered_df[existing_columns]

            # 컬럼 이름 변경
            column_mapping = {
                "구분": "구분(승인관리:A/가격관리:P)",
                "담당자": "담당자",
                "업체명": "공급사명",
                "업체코드": "공급처코드",
                "Code": "상품코드",
                "중분류카테고리": "카테고리(중분류)",
                "상품명": "상품명",
                "기본수량(1)": "본사 기본수량",
                "판매단가(V포함)": "판매단가1(VAT포함)",
                "본사상품링크": "본사링크",
                "기본수량(2)": "고려 기본수량",
                "판매단가(V포함)(2)": "판매단가2(VAT포함)",
                "가격차이(2)": "고려 가격차이",
                "가격차이(2)(%)": "고려 가격차이(%)",
                "고려기프트 상품링크": "고려 링크",
                "기본수량(3)": "네이버 기본수량",
                "판매단가(V포함)(3)": "판매단가3 (VAT포함)",
                "가격차이(3)": "네이버 가격차이",
                "가격차이(3)(%)": "네이버가격차이(%)",
                "공급사명": "네이버 공급사명",
                "공급사 상품링크": "네이버 링크",
                "본사 이미지": "해오름(이미지링크)",
                "고려기프트 이미지": "고려기프트(이미지링크)",
                "네이버 이미지": "네이버쇼핑(이미지링크)",
            }

            # 존재하는 컬럼만 매핑
            rename_mapping = {
                k: v for k, v in column_mapping.items() if k in filtered_df.columns
            }
            filtered_df.rename(columns=rename_mapping, inplace=True)

            # 엑셀로 저장
            filtered_df.to_excel(output_path, index=False)

            # 추가 포맷팅 적용
            wb = load_workbook(output_path)
            ws = wb.active

            # 스타일 설정
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 테두리 및 정렬 적용
            for row in ws.iter_rows(
                min_row=2, max_col=ws.max_column, max_row=ws.max_row
            ):
                for cell in row:
                    cell.border = thin_border

            for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
                for cell in col:
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

            # 컬럼 너비 및 행 높이 설정
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, max_row=1):
                for cell in col:
                    ws.column_dimensions[cell.column_letter].width = 15

            for row in ws.iter_rows(
                min_row=1, max_col=ws.max_column, max_row=ws.max_row
            ):
                for cell in row:
                    ws.row_dimensions[cell.row].height = 16.5

            # 품절 항목 및 특정 조건 처리
            self._apply_filter_rules(ws)

            # 변경사항 저장
            wb.save(output_path)
            self.logger.info(f"Filtered Excel file saved to: {output_path}")

            return output_path

        except Exception as e:
            self.logger.error(
                f"Error filtering Excel by price differences: {str(e)}", exc_info=True
            )
            return file_path

    def _apply_filter_rules(self, worksheet):
        """필터링 규칙을 적용합니다. (filter_excel_file.py 에서 추출)"""
        from urllib.parse import urlparse

        # 품절 행 삭제 및 데이터 정리 로직
        for row in reversed(
            list(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row))
        ):
            # 고려 기본수량이 품절일 경우 처리
            goleo_quantity_col = None
            naver_quantity_col = None

            # 컬럼 인덱스 찾기
            for i, cell in enumerate(worksheet[1]):
                if cell.value == "고려 기본수량":
                    goleo_quantity_col = i
                elif cell.value == "네이버 기본수량":
                    naver_quantity_col = i

            # 품절 항목 삭제
            if (
                goleo_quantity_col is not None
                and row[goleo_quantity_col].value == "품절"
            ):
                worksheet.delete_rows(row[0].row)
                continue

            if (
                naver_quantity_col is not None
                and row[naver_quantity_col].value == "품절"
            ):
                worksheet.delete_rows(row[0].row)
                continue

            # 링크 검증
            goleo_link_col = None
            naver_link_col = None

            for i, cell in enumerate(worksheet[1]):
                if cell.value == "고려 링크":
                    goleo_link_col = i
                elif cell.value == "네이버 링크":
                    naver_link_col = i

            # 유효하지 않은 링크 제거
            if goleo_link_col is not None:
                link_value = row[goleo_link_col].value
                if (
                    link_value
                    and isinstance(link_value, str)
                    and not bool(urlparse(link_value).scheme)
                ):
                    row[goleo_link_col].value = None

            if naver_link_col is not None:
                link_value = row[naver_link_col].value
                if (
                    link_value
                    and isinstance(link_value, str)
                    and not bool(urlparse(link_value).scheme)
                ):
                    row[naver_link_col].value = None

    def _write_product_data(self, worksheet, row_idx: int, product: Product) -> int:
        """Write product data to worksheet and return the next row index"""
        try:
            # 기본 데이터 쓰기
            worksheet.cell(row=row_idx, column=1, value=product.name or "제품명 없음")
            worksheet.cell(row=row_idx, column=2, value=product.product_code or product.id or "코드 없음")
            worksheet.cell(row=row_idx, column=3, value=product.price or 0)
            
            # 이미지 URL 처리
            main_image = product.image_url or "이미지 없음"
            cleaned_main_image = self._clean_url(main_image) # Use cleaner
            
            if cleaned_main_image.startswith('https:'):
                # 이미지 URL이 있으면 Excel IMAGE 함수 사용
                # 두 번째 매개변수 2 = 셀에 맞게 크기 조정
                image_formula = f'=IMAGE("{cleaned_main_image}", 2)'
                worksheet.cell(row=row_idx, column=4, value=image_formula)
                
                # 이미지 URL도 별도 열에 저장 (디버깅용)
                worksheet.cell(row=row_idx, column=13, value=cleaned_main_image)
            else:
                # 이미지 URL이 없거나 유효하지 않은 경우 기본 네이버 이미지 사용
                default_image = "https://ssl.pstatic.net/static/shop/front/techreview/web/resource/images/naver.png"
                image_formula = f'=IMAGE("{default_image}", 2)'
                worksheet.cell(row=row_idx, column=4, value=image_formula)
                worksheet.cell(row=row_idx, column=13, value="기본 이미지 사용: " + default_image)
                if main_image != "이미지 없음": # Log only if there was an attempt at a URL
                    self.logger.warning(f"Invalid or non-HTTPS image URL for product {product.name}: {main_image}. Using default.")
            
            # 이미지 갤러리 처리
            gallery_urls = product.image_gallery or []
            if gallery_urls and isinstance(gallery_urls, list):
                # URL 정규화 및 @ 기호 제거 후 갤러리 이미지 URL들을 쉼표로 구분하여 저장
                normalized_urls = []
                for url in gallery_urls:
                    cleaned_gallery_url = self._clean_url(url) # Use cleaner
                    if cleaned_gallery_url.startswith('https:'): # Only include valid https urls
                        normalized_urls.append(cleaned_gallery_url)
                
                gallery_str = " | ".join(normalized_urls) if normalized_urls else "추가 이미지 없음"
                worksheet.cell(row=row_idx, column=5, value=gallery_str)
            else:
                worksheet.cell(row=row_idx, column=5, value="추가 이미지 없음")
            
            # 제품 URL
            product_url = product.url or "URL 없음"
            cleaned_product_url = self._clean_url(product_url) # Use cleaner
            worksheet.cell(row=row_idx, column=6, value=cleaned_product_url)
            
            # 소스 정보
            source = product.source or "알 수 없음"
            worksheet.cell(row=row_idx, column=7, value=source)
            
            # 브랜드 정보 
            brand = product.brand or "알 수 없음"
            worksheet.cell(row=row_idx, column=8, value=brand)
            
            # 카테고리 정보
            category = product.category or "알 수 없음"
            worksheet.cell(row=row_idx, column=9, value=category)
            
            # 스크래핑 시간
            fetched_at = product.fetched_at or datetime.now().isoformat()
            worksheet.cell(row=row_idx, column=10, value=fetched_at)
            
            # 추가 제품 정보 (설명, 상세정보 등)
            description = product.description or ""
            worksheet.cell(row=row_idx, column=11, value=description)
            
            # 원본 데이터 (JSON으로 변환하여 저장)
            if hasattr(product, 'original_input_data') and product.original_input_data:
                try:
                    import json
                    original_data = json.dumps(product.original_input_data, ensure_ascii=False)
                    worksheet.cell(row=row_idx, column=12, value=str(original_data)[:32000])  # Excel 셀 최대 크기 제한
                except Exception as json_err:
                    self.logger.warning(f"Error converting original data to JSON: {str(json_err)}")
            
            # 행 높이 조정 (이미지가 잘 표시되도록)
            worksheet.row_dimensions[row_idx].height = 120
            
            # 다음 행 인덱스 반환
            return row_idx + 1
        
        except Exception as e:
            self.logger.error(f"Error writing product data to Excel: {str(e)}", exc_info=True)
            # 오류가 발생해도 다음 행으로 진행
            return row_idx + 1

    def create_worksheet(self, workbook, sheet_name: str) -> Worksheet:
        """Create a new worksheet with headers and styling"""
        try:
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # 기본 헤더 설정
            headers = [
                "제품명", 
                "상품코드", 
                "가격", 
                "메인 이미지", 
                "추가 이미지", 
                "상품 URL", 
                "소스", 
                "브랜드", 
                "카테고리", 
                "스크래핑 시간", 
                "상세 설명",
                "원본 데이터",
                "이미지 URL"  # 디버깅용 이미지 URL 열 추가
            ]
            
            # 헤더 쓰기
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                # 헤더 스타일 적용
                if hasattr(self, 'header_fill') and hasattr(self, 'border_style'):
                    cell.fill = self.header_fill
                    cell.border = self.border_style
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 열 너비 초기 설정
            column_widths = {
                1: 40,  # 제품명
                2: 15,  # 상품코드
                3: 15,  # 가격
                4: 25,  # 메인 이미지
                5: 25,  # 추가 이미지
                6: 30,  # 상품 URL
                7: 15,  # 소스
                8: 20,  # 브랜드
                9: 20,  # 카테고리
                10: 20, # 스크래핑 시간
                11: 40, # 상세 설명
                12: 50, # 원본 데이터
                13: 50, # 이미지 URL
            }
            
            # 열 너비 설정
            for col_idx, width in column_widths.items():
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = width
            
            # 필터 추가
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
            
            # 첫 행 고정
            worksheet.freeze_panes = "A2"
            
            return worksheet
            
        except Exception as e:
            self.logger.error(f"Error creating worksheet: {str(e)}", exc_info=True)
            # 에러 시 기본 워크시트 반환
            worksheet = workbook.create_sheet(title=sheet_name)
            worksheet.cell(row=1, column=1, value="워크시트 생성 중 오류 발생")
            return worksheet

    def save_products(self, products: List[Product], output_path: str, sheet_name: str = None, naver_results: List[Product] = None):
        """Save products to Excel file with proper formatting and error handling"""
        try:
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"Created output directory: {output_dir}")
            
            workbook = Workbook()
            
            # 시트 이름이 없으면 현재 시간으로 생성
            if not sheet_name:
                sheet_name = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 기본 시트 제거 및 새 시트 생성
            workbook.remove(workbook.active)
            worksheet = self.create_worksheet(workbook, sheet_name)
            
            # 제품 데이터 쓰기
            current_row = 2  # 헤더 다음 행부터 시작
            
            # 고려기프트 제품 쓰기
            if products:
                self.logger.info(f"Writing {len(products)} products to Excel")
                for product in products:
                    if hasattr(product, 'source') and product.source == 'koryo':  # 고려기프트 제품 확인
                        self.logger.debug(f"Processing Koryo product: {product.name}")
                    current_row = self._write_product_data(worksheet, current_row, product)
            
            # 네이버 검색 결과 쓰기
            if naver_results:
                self.logger.info(f"Writing {len(naver_results)} Naver results to Excel")
                for product in naver_results:
                    if hasattr(product, 'id') and product.id != "no_match": # no_match 제품은 건너뜁니다
                        self.logger.debug(f"Processing Naver product: {product.name}")
                        current_row = self._write_product_data(worksheet, current_row, product)
            
            # 결과 요약 추가
            total_products = (len(products) if products else 0) + (len([p for p in naver_results if hasattr(p, 'id') and p.id != "no_match"]) if naver_results else 0)
            summary_row = current_row + 1
            worksheet.cell(row=summary_row, column=1, value=f"총 제품 수: {total_products}")
            worksheet.cell(row=summary_row, column=2, value=f"처리 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # 파일 저장
            try:
                self.logger.info(f"Saving Excel file to: {output_path}")
                workbook.save(output_path)
                self.logger.info(f"Successfully saved {total_products} products to {output_path}")
                
                # 파일 후처리 (포맷팅, @ 기호 제거 등) - 직접 호출
                self.logger.info(f"Starting post-processing for {output_path}")
                
                # 1. @ 기호 제거 (remove_at_symbol은 파일을 직접 수정)
                try:
                    self.remove_at_symbol(output_path)
                    self.logger.info(f"Removed @ symbols from {output_path}")
                except Exception as remove_err:
                     self.logger.error(f"Error removing @ symbols during save: {str(remove_err)}")
                     # Continue processing even if @ removal fails
                
                # 2. 이미지 수식 개선 (_fix_image_formulas는 파일을 직접 수정)
                try:
                    self._fix_image_formulas(output_path)
                    self.logger.info(f"Fixed image formulas in {output_path}")
                except Exception as img_err:
                    self.logger.error(f"Error fixing image formulas during save: {str(img_err)}")
                    # Continue processing even if image fixing fails
                
                # 3. 포맷팅 적용 (apply_formatting_to_excel은 파일을 직접 수정)
                try:
                    self.apply_formatting_to_excel(output_path)
                    self.logger.info(f"Applied formatting to {output_path}")
                except Exception as format_err:
                    self.logger.error(f"Error applying formatting during save: {str(format_err)}")
                    # Continue processing even if formatting fails
                    
                # 4. 하이퍼링크 추가 (add_hyperlinks_to_excel은 파일을 직접 수정)
                try:
                    self.add_hyperlinks_to_excel(output_path)
                    self.logger.info(f"Added hyperlinks to {output_path}")
                except Exception as link_err:
                    self.logger.error(f"Error adding hyperlinks during save: {str(link_err)}")
                    # Continue processing even if linking fails

                # processed_path = self.post_process_excel_file(output_path)
                # if processed_path != output_path:
                #     self.logger.info(f"Post-processed Excel file: {processed_path}")
                #     return processed_path
                
                self.logger.info(f"Post-processing completed for {output_path}")
                return output_path
            except Exception as save_err:
                self.logger.error(f"Error saving Excel file: {str(save_err)}")
                raise
            
        except Exception as e:
            self.logger.error(f"Error saving products to Excel: {str(e)}", exc_info=True)
            raise

    def remove_at_symbol(self, file_path: str) -> str:
        """
        엑셀 파일에서 @ 기호를 제거합니다.

        Args:
            file_path: 입력 엑셀 파일 경로

        Returns:
            str: 처리된 파일 경로
        """
        try:
            self.logger.info(f"Removing @ symbols from Excel file: {file_path}")

            # 파일이 존재하는지 확인
            if not os.path.exists(file_path):
                self.logger.error(f"File not found: {file_path}")
                return file_path

            # 파일이 열려있는지 확인
            try:
                with open(file_path, 'a'):
                    pass
            except PermissionError:
                self.logger.error(f"File is currently in use: {file_path}")
                return file_path

            # 임시 파일 경로 생성
            temp_dir = os.path.dirname(file_path)
            temp_file = os.path.join(temp_dir, f"temp_{os.path.basename(file_path)}")

            # DataFrame 로드
            df = pd.read_excel(file_path)

            # 모든 문자열 컬럼에서 @ 기호 제거
            for column in df.columns:
                if df[column].dtype == 'object':  # 문자열 타입 컬럼만 처리
                    df[column] = df[column].astype(str).str.replace('@', '', regex=False)

            # 특별히 IMAGE 함수가 포함된 셀 처리 
            image_columns = ['본사 이미지', '고려기프트 이미지', '네이버 이미지']
            for col in image_columns:
                if col in df.columns:
                    # IMAGE 함수 내의 URL에서도 @ 제거
                    df[col] = df[col].apply(
                        lambda x: x.replace('@', '') if isinstance(x, str) else x
                    )

            # 임시 파일에 저장
            self.logger.info(f"Saving cleaned Excel file to: {temp_file}")
            df.to_excel(temp_file, index=False)

            # 원본 파일 삭제 및 임시 파일 이동
            try:
                os.remove(file_path)
                os.rename(temp_file, file_path)
                self.logger.info(f"Successfully replaced original file with cleaned version")
            except Exception as e:
                # 실패 시 임시 파일 삭제
                self.logger.error(f"Error replacing original file: {str(e)}")
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                return file_path

            self.logger.info(f"Successfully removed @ symbols from: {file_path}")
            return file_path

        except Exception as e:
            self.logger.error(f"Error removing @ symbols from Excel: {str(e)}", exc_info=True)
            return file_path

    def post_process_excel_file(self, file_path: str) -> str:
        """
        엑셀 파일에 대한 후처리를 수행합니다.
        1. @ 기호 제거
        2. 이미지 수식 개선
        3. 포맷팅 적용
        4. 하이퍼링크 추가
        
        Args:
            file_path: 처리할 엑셀 파일 경로
            
        Returns:
            str: 처리된 파일 경로
        """
        try:
            if not os.path.exists(file_path):
                self.logger.error(f"File not found for post-processing: {file_path}")
                return file_path
                
            self.logger.info(f"Post-processing Excel file: {file_path}")
            
            # 1. @ 기호 제거
            cleaned_file = self.remove_at_symbol(file_path)
            
            # 2. 이미지 수식 개선
            try:
                self._fix_image_formulas(cleaned_file)
                self.logger.info(f"Fixed image formulas in: {cleaned_file}")
            except Exception as img_err:
                self.logger.error(f"Error fixing image formulas: {str(img_err)}")
            
            # 3. 포맷팅 적용
            try:
                self.apply_formatting_to_excel(cleaned_file)
                self.logger.info(f"Applied formatting to: {cleaned_file}")
            except Exception as format_err:
                self.logger.error(f"Error applying formatting: {str(format_err)}")
                
            # 4. 하이퍼링크 추가
            try:
                linked_file = self.add_hyperlinks_to_excel(cleaned_file)
                if linked_file != cleaned_file:
                    self.logger.info(f"Added hyperlinks, new file: {linked_file}")
                    return linked_file
            except Exception as link_err:
                self.logger.error(f"Error adding hyperlinks: {str(link_err)}")
            
            return cleaned_file
        except Exception as e:
            self.logger.error(f"Error in post-processing Excel file: {str(e)}")
            return file_path
            
    def _fix_image_formulas(self, file_path: str) -> None:
        """이미지 표시 관련 수식을 수정합니다."""
        try:
            # openpyxl을 사용하여 엑셀 파일 로드
            wb = load_workbook(file_path)
            made_changes = False
            
            # 모든 시트 검사
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 이미지 URL이 있는 열 찾기 (보통 13번 열)
                image_url_col_idx = None # Use index (1-based)
                image_formula_col_idx = None # Main image formula column (usually 4)
                header_row = ws[1]

                for i, cell in enumerate(header_row, 1):
                    if cell.value == "이미지 URL":
                        image_url_col_idx = i
                    elif cell.value == "메인 이미지": # Assuming the formula is in "메인 이미지"
                         image_formula_col_idx = i
                
                # Ensure both columns were found
                if image_url_col_idx is None:
                    self.logger.warning(f"'이미지 URL' column not found in sheet '{sheet_name}'. Skipping formula fix for this sheet.")
                    continue
                if image_formula_col_idx is None:
                     self.logger.warning(f"'메인 이미지' column not found in sheet '{sheet_name}'. Skipping formula fix for this sheet.")
                     continue
                
                # 모든 셀을 검사하여 이미지 수식 찾기
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  # 헤더 제외
                    # 1. IMAGE 수식이 있는 셀 확인 ("메인 이미지" 열)
                    image_cell = row[image_formula_col_idx - 1]  # 0-based index
                    
                    # 2. 원본 이미지 URL 가져오기 ("이미지 URL" 열)
                    original_url = None
                    if image_url_col_idx <= len(row):
                        url_cell = row[image_url_col_idx - 1]  # 0-based index
                        if url_cell.value and isinstance(url_cell.value, str) and (
                            url_cell.value.startswith("http") or
                            url_cell.value.startswith("기본 이미지 사용:")
                        ):
                            # URL 추출
                            if url_cell.value.startswith("기본 이미지 사용:"):
                                parts = url_cell.value.split(":", 1)
                                if len(parts) > 1:
                                    original_url = parts[1].strip()
                            else:
                                original_url = url_cell.value
                    
                    # 3. 이미지 수식 처리
                    if isinstance(image_cell.value, str) and image_cell.value.startswith('=IMAGE('):
                        # 원래 수식 저장
                        original_formula = image_cell.value
                        
                        # 수식에서 URL 추출
                        url_match = re.search(r'=IMAGE\("([^"]+)"', original_formula)
                        url = None
                        
                        if url_match:
                            url = url_match.group(1)
                        elif original_url:  # 수식에서 URL을 추출할 수 없지만 원본 URL이 있는 경우
                            url = original_url
                        
                        if url:
                            # URL 정상화
                            if url.startswith('http:'):
                                url = 'https:' + url[5:]
                            elif url.startswith('//'):
                                url = 'https:' + url
                            
                            # @ 기호 제거
                            if '@' in url:
                                url = url.replace('@', '')
                            
                            # 특수 문자 처리
                            url = url.replace('\\', '/')
                            url = url.replace('"', '%22')
                            url = url.replace(' ', '%20')
                            
                            # 수정된 수식 적용 (2 = 셀에 맞게 크기 조정)
                            new_formula = f'=IMAGE("{url}", 2)'
                            if new_formula != original_formula:
                                image_cell.value = new_formula
                                made_changes = True
                                self.logger.debug(f"이미지 수식 수정: {sheet_name} 시트 {row_idx}행")
                    
                    # 4. 이미지 수식이 없지만 원본 URL이 있는 경우 수식 생성
                    elif original_url and (
                        not image_cell.value or 
                        not isinstance(image_cell.value, str) or 
                        not image_cell.value.startswith('=IMAGE(')
                    ):
                        # URL 정상화
                        if original_url.startswith('http:'):
                            original_url = 'https:' + original_url[5:]
                        elif original_url.startswith('//'):
                            original_url = 'https:' + original_url
                        
                        # @ 기호 제거
                        if '@' in original_url:
                            original_url = original_url.replace('@', '')
                        
                        # 특수 문자 처리
                        original_url = original_url.replace('\\', '/')
                        original_url = original_url.replace('"', '%22')
                        original_url = original_url.replace(' ', '%20')
                        
                        # 새 수식 생성
                        new_formula = f'=IMAGE("{original_url}", 2)'
                        image_cell.value = new_formula
                        made_changes = True
                        self.logger.debug(f"이미지 수식 새로 추가: {sheet_name} 시트 {row_idx}행")
            
            # 변경 사항이 있으면 저장
            if made_changes:
                wb.save(file_path)
                self.logger.info(f"이미지 수식 수정 완료: {file_path}")
        except Exception as e:
            self.logger.error(f"이미지 수식 수정 중 오류: {str(e)}", exc_info=True)

    def save_excel_file(self, df: pd.DataFrame, file_path: str) -> None:
        """Saves a DataFrame to an Excel file."""
        try:
            self.logger.info(f"Saving DataFrame to Excel file: {file_path}")
            # Ensure output directory exists
            output_dir = os.path.dirname(file_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"Created output directory: {output_dir}")

            df.to_excel(file_path, index=False)
            self.logger.info(f"Successfully saved DataFrame to {file_path}")
        except Exception as e:
            self.logger.error(f"Error saving DataFrame to Excel: {str(e)}", exc_info=True)
            raise # Re-raise the exception to be handled by the caller
