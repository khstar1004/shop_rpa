import logging
import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger("test_excel_import")

def create_sample_data():
    """Create sample product data"""
    products = [
        {
            "name": "사무용 볼펜 세트",
            "price": 12000,
            "id": "PEN001",
            "source": "test",
            "url": "https://example.com/pen001",
            "image_url": "https://example.com/images/pen001.jpg",
            "brand": "스테들러",
            "category": "사무용품",
            "description": "고급 사무용 볼펜 세트입니다."
        },
        {
            "name": "노트북 파우치",
            "price": 25000,
            "id": "POUCH001",
            "source": "test",
            "url": "https://example.com/pouch001",
            "image_url": "https://example.com/images/pouch001.jpg",
            "brand": "삼성",
            "category": "컴퓨터 액세서리",
            "description": "15인치 노트북 수납 가능한 패딩 파우치입니다."
        },
        {
            "name": "메모지 스티커",
            "price": 5000,
            "id": "MEMO001",
            "source": "test",
            "url": "https://example.com/memo001",
            "image_url": "https://example.com/images/memo001.jpg",
            "brand": "포스트잇",
            "category": "사무용품",
            "description": "다양한 색상의 메모지 스티커 세트입니다."
        }
    ]
    return products

def save_to_excel(products, output_file):
    """Save products to Excel file"""
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "제품 목록"
    
    # Define headers
    headers = [
        "제품명", 
        "상품코드", 
        "가격", 
        "메인 이미지", 
        "상품 URL", 
        "소스", 
        "브랜드", 
        "카테고리", 
        "설명"
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
    
    # Write product data
    for row_idx, product in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=product["name"])
        ws.cell(row=row_idx, column=2, value=product["id"])
        ws.cell(row=row_idx, column=3, value=product["price"])
        
        # Add image formula
        image_url = product["image_url"]
        image_formula = f'=IMAGE("{image_url}", 2)'
        ws.cell(row=row_idx, column=4, value=image_formula)
        
        ws.cell(row=row_idx, column=5, value=product["url"])
        ws.cell(row=row_idx, column=6, value=product["source"])
        ws.cell(row=row_idx, column=7, value=product["brand"])
        ws.cell(row=row_idx, column=8, value=product["category"])
        ws.cell(row=row_idx, column=9, value=product["description"])
    
    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 30  # 제품명
    ws.column_dimensions[get_column_letter(4)].width = 30  # 이미지
    ws.column_dimensions[get_column_letter(5)].width = 40  # URL
    ws.column_dimensions[get_column_letter(9)].width = 40  # 설명
    
    # Save workbook
    wb.save(output_file)
    logger.info(f"Excel file saved successfully: {output_file}")
    
    return output_file

def main():
    """Test Excel functionality"""
    logger.info("Starting Excel functionality test")
    
    # Create sample data
    products = create_sample_data()
    logger.info(f"Created {len(products)} sample products")
    
    # Define output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"excel_test_{timestamp}.xlsx")
    
    # Save to Excel
    try:
        saved_file = save_to_excel(products, output_file)
        logger.info(f"Test completed successfully. Output file: {saved_file}")
    except Exception as e:
        logger.error(f"Test failed: {str(e)}", exc_info=True)
        
if __name__ == "__main__":
    main() 