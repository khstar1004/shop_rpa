"""
Excel 파일 처리를 위한 유틸리티 함수들

PythonScript 폴더의 Excel 관련 기능을 활용하여 일반적인 Excel 작업을 쉽게 수행할 수 있는
유틸리티 함수들을 제공합니다.
"""

import logging
import os
import re
import zipfile # Added for exception handling
from datetime import datetime
from typing import Optional, Any
from urllib.parse import urlparse # Moved import
import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font, NamedStyle
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image as OpenpyxlImage # Renamed to avoid conflict
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage # Added Pillow import
from tenacity import retry, stop_after_attempt, wait_exponential

try:
    from PIL import Image
    _PIL_AVAILABLE = True
except ImportError:
    _PIL_AVAILABLE = False
    IMAGE = None

# 기본 로거 설정
logger = logging.getLogger(__name__)

# --- Formatting and Styling Functions ---

DEFAULT_FONT = Font(name="맑은 고딕", size=10)
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER_SIDE = Side(style="thin")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")


# Define NamedStyles
header_style = NamedStyle(
    name="header_style",
    font=HEADER_FONT,
    border=THIN_BORDER,
    alignment=CENTER_ALIGNMENT,
    fill=GRAY_FILL,
)
default_style = NamedStyle(
    name="default_style", font=DEFAULT_FONT, border=THIN_BORDER, alignment=LEFT_ALIGNMENT
)
center_style = NamedStyle(
    name="center_style",
    font=DEFAULT_FONT,
    border=THIN_BORDER,
    alignment=CENTER_ALIGNMENT,
)


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
def download_image(url: str, timeout: int = 10) -> Optional[bytes]:
    """Downloads an image from a URL."""
    # (Keep existing download_image function)
    # ... existing code ...


def insert_image_to_cell(ws: Worksheet, image_path: str, cell_address: str, size: tuple = (100, 100)):
    """
    Inserts a resized image into the specified cell using openpyxl.

    Args:
        ws: The openpyxl worksheet object.
        image_path: Path to the local image file.
        cell_address: The target cell (e.g., 'A1').
        size: Tuple of desired (width, height) in pixels.
    """
    try:
        if not os.path.exists(image_path):
            logging.warning(f"Image path does not exist: {image_path}")
            return

        # Resize image using Pillow
        img_pil = PILImage.open(image_path)
        img_pil = img_pil.convert("RGB")  # Ensure compatibility, remove alpha
        img_pil.thumbnail(size) # Resize while maintaining aspect ratio

        # Prepare image for openpyxl
        img_byte_arr = io.BytesIO()
        img_pil.save(img_byte_arr, format='PNG') # Save resized image to byte stream
        img_byte_arr.seek(0)

        img_openpyxl = OpenpyxlImage(img_byte_arr)

        # Set image dimensions (optional, as thumbnail maintains aspect ratio)
        # img_openpyxl.width, img_openpyxl.height = size

        # Add image to worksheet anchored to the cell
        ws.add_image(img_openpyxl, cell_address)

        # Adjust row height and column width (approximations based on Excelsaveimage.py)
        target_cell = ws[cell_address]
        # Excel row height is in points (1 point = 1/72 inch). Pillow size is in pixels.
        # Approximation: 1 pixel ~ 0.75 points (at 96 DPI)
        row_height_points = size[1] * 0.75
        ws.row_dimensions[target_cell.row].height = max(ws.row_dimensions[target_cell.row].height or 0, row_height_points)

        # Excel column width is in characters. Very approximate conversion.
        # Using the ratio from Excelsaveimage.py: width_pixels / 6.25
        col_width_chars = size[0] / 7.0 # Adjusted divisor slightly based on common observations
        col_letter = get_column_letter(target_cell.column)
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, col_width_chars)

        logging.debug(f"Inserted image {os.path.basename(image_path)} into cell {cell_address}, adjusted row/col size.")

    except FileNotFoundError:
        logging.warning(f"Image file not found: {image_path}")
    except OSError as e:
        logging.error(f"Failed to insert image {image_path} into cell {cell_address}: {e}", exc_info=True)


def clean_value(value: Any) -> Any:
    """Cleans value for Excel insertion."""
    # (Keep existing clean_value function)
    # ... existing code ...
