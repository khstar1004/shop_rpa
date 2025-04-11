"""
데이터 전처리 및 파일 관련 유틸리티 함수 모음.

이 모듈은 로깅 설정, 입력 파일 유효성 검사, 데이터 정제 (예: 상품명 정리),
파일 분할 및 병합 등 RPA 프로세스의 다양한 단계에서 사용될 수 있는
전처리 및 파일 관리 함수들을 제공합니다.
"""
import logging
import os
import shutil
import sys
import traceback
from datetime import datetime
from logging.handlers import RotatingFileHandler, TimedRotatingFileHandler
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


def setup_logging(log_dir: Optional[str] = None) -> None:
    """Setup logging configuration with improved error handling and features."""
    try:
        # Create log directory if specified
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)

        # Get current timestamp for log file names
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Create formatters
        detailed_formatter = logging.Formatter(
            "%(asctime)s - %(name)s - %(levelname)s - %(message)s\n"
            "File: %(pathname)s:%(lineno)d\n"
            "Function: %(funcName)s\n"
            "%(message)s"
        )

        simple_formatter = logging.Formatter(
            "%(asctime)s - %(levelname)s - %(message)s"
        )

        # Setup root logger
        root_logger = logging.getLogger()
        root_logger.setLevel(logging.DEBUG)

        # Clear any existing handlers
        root_logger.handlers.clear()

        # Console handler (INFO level)
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(simple_formatter)
        root_logger.addHandler(console_handler)

        # Error console handler (ERROR level)
        error_console_handler = logging.StreamHandler(sys.stderr)
        error_console_handler.setLevel(logging.ERROR)
        error_console_handler.setFormatter(detailed_formatter)
        root_logger.addHandler(error_console_handler)

        if log_dir:
            # File handler for all logs (DEBUG level)
            all_log_file = os.path.join(log_dir, f"all_{timestamp}.log")
            file_handler = RotatingFileHandler(
                all_log_file,
                maxBytes=10 * 1024 * 1024,  # 10MB
                backupCount=5,
                encoding="utf-8",
            )
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(detailed_formatter)
            root_logger.addHandler(file_handler)

            # Error log file handler (ERROR level)
            error_log_file = os.path.join(log_dir, f"error_{timestamp}.log")
            error_file_handler = TimedRotatingFileHandler(
                error_log_file,
                when="midnight",
                interval=1,
                backupCount=30,  # Keep 30 days of error logs
                encoding="utf-8",
            )
            error_file_handler.setLevel(logging.ERROR)
            error_file_handler.setFormatter(detailed_formatter)
            root_logger.addHandler(error_file_handler)

            # Create a copy of the latest log file instead of symlink
            latest_log = os.path.join(log_dir, "latest.log")
            if os.path.exists(latest_log):
                os.remove(latest_log)
            shutil.copy2(all_log_file, latest_log)

        # Add unhandled exception handler
        def handle_exception(exc_type, exc_value, exc_traceback):
            if issubclass(exc_type, KeyboardInterrupt):
                # Don't log keyboard interrupts
                sys.__excepthook__(exc_type, exc_value, exc_traceback)
            else:
                root_logger.error(
                    "Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback)
                )

        sys.excepthook = handle_exception

        # Log successful setup
        root_logger.info("Logging system initialized successfully")

    except Exception as e:
        # If logging setup fails, try to log to console as last resort
        print(f"Failed to setup logging: {str(e)}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        raise


class LogCapture:
    """
    특정 코드 블록 실행 동안 발생하는 로그 메시지를 캡처하는 컨텍스트 관리자.

    테스트 또는 특정 작업의 로깅 출력을 분리하여 확인하고 싶을 때 유용합니다.

    사용 예시:
        with LogCapture('my_logger') as lc:
            # 로그를 발생시키는 코드 실행
            logger.info("This message will be captured")
        messages = lc.get_messages()
        print(messages)
    """

    def __init__(self, logger_name: str, level: int = logging.DEBUG):
        self.logger_name = logger_name
        self.level = level
        self.logger = logging.getLogger(logger_name)
        self.handler = None
        self.records = []

    def __enter__(self):
        class RecordListHandler(logging.Handler):
            def __init__(self, records):
                super().__init__()
                self.records = records

            def emit(self, record):
                self.records.append(record)

        self.handler = RecordListHandler(self.records)
        self.handler.setLevel(self.level)
        self.logger.addHandler(self.handler)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.handler:
            self.logger.removeHandler(self.handler)

    def get_records(self) -> list:
        """Get captured log records."""
        return self.records

    def get_messages(self) -> list:
        """Get formatted log messages."""
        return [self.handler.format(record) for record in self.records]


def validate_input_file(file_path: str, required_columns: List[str]) -> bool:
    """
    Validate input file format and required columns.

    Args:
        file_path: Path to the input Excel file
        required_columns: List of required column names

    Returns:
        bool: True if file is valid, False otherwise
    """
    try:
        # Verify file exists
        if not os.path.exists(file_path):
            logger.error("Input file not found: %s", file_path)
            return False

        # Check file extension
        _, ext = os.path.splitext(file_path)
        if ext.lower() not in [".xls", ".xlsx", ".xlsm"]:
            logger.error(
                "Invalid file format: %s. Expected Excel file (.xls, .xlsx, .xlsm)", ext
            )
            return False

        # Verify file can be opened and read headers
        try:
            df_headers = pd.read_excel(file_path, nrows=0)
        except FileNotFoundError:
            logger.error("Input file not found during read: %s", file_path)
            return False
        except pd.errors.EmptyDataError:
            logger.error("Input file is empty: %s", file_path)
            return False
        except ValueError as ve:
            logger.error("Error parsing Excel file structure: %s", ve)
            return False
        except Exception as e:  # Catch other potential read errors (e.g., corrupted file)
            logger.error("Failed to open or read header from Excel file: %s", e)
            return False

        # Check for required columns
        missing_cols = [col for col in required_columns if col not in df_headers.columns]
        if missing_cols:
            logger.error("Missing required columns: %s", ', '.join(missing_cols))
            return False

        return True

    except OSError as e:
        logger.error("OS error during file validation: %s", e)
        return False
    except Exception as e:
        logger.error("Unexpected error validating input file: %s", e)
        return False


def clean_product_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean product names according to manual requirements:
    - Remove '1-' prefix and numbers/hyphens before it
    - Remove special characters like '/' and '()'
    - Remove numbers in parentheses

    Args:
        df: Input DataFrame with '상품명' column

    Returns:
        DataFrame with cleaned product names
    """
    if "상품명" not in df.columns:
        logger.warning("Column '상품명' not found, skipping product name cleaning")
        return df

    logger.info("Cleaning product names")

    try:
        # Create a copy to avoid modifying the original
        cleaned_df = df.copy()

        # Ensure '상품명' is string type before applying string methods
        if pd.api.types.is_string_dtype(cleaned_df["상품명"]):
            # Remove '1-' and numbers/hyphens before it
            cleaned_df["상품명"] = cleaned_df["상품명"].str.replace(
                r"^\d+-", "", regex=True
            )

            # Remove special characters and brackets
            cleaned_df["상품명"] = cleaned_df["상품명"].str.replace(
                r"[/()]", "", regex=True
            )

            # Remove numbers in parentheses
            cleaned_df["상품명"] = cleaned_df["상품명"].str.replace(
                r"\(\d+\)", "", regex=True
            )

            # Trim extra whitespace
            cleaned_df["상품명"] = cleaned_df["상품명"].str.strip()
        else:
            logger.warning("Skipping product name cleaning as the column is not string type.")

        return cleaned_df

    except (AttributeError, TypeError, ValueError) as e:
        logger.error("Error during product name cleaning: %s", e)
        return df  # Return original if specific error occurs
    except Exception as e:
        logger.error("Unexpected error cleaning product names: %s", e)
        return df


def process_input_file(
    file_path: str, config: Dict[str, Any]
) -> Optional[pd.DataFrame]:
    """
    Process and validate input Excel file.

    Args:
        file_path: Path to the input Excel file
        config: Configuration dictionary

    Returns:
        DataFrame or None if validation fails
    """
    logger.info("Processing input file: %s", file_path)

    try:
        # Extract required columns from config
        required_columns = config["EXCEL"].get(
            "REQUIRED_COLUMNS",
            ["상품명", "판매단가(V포함)", "상품Code", "본사 이미지", "본사상품링크"],
        )

        # Validate input file
        if not validate_input_file(file_path, required_columns):
            return None

        # Read the Excel file
        df = pd.read_excel(file_path)

        # Automatically clean product names if enabled in config
        if config.get("PROCESSING", {}).get("AUTO_CLEAN_PRODUCT_NAMES", True):
            df = clean_product_names(df)
            logger.info("Product names cleaned automatically")

        # Add other preprocessing steps here
        df = normalize_column_types(df)
        df = handle_missing_values(df, required_columns)

        # Check for duplicate products
        if config.get("EXCEL", {}).get("ENABLE_DUPLICATE_DETECTION", True):
            if "상품Code" in df.columns:
                duplicates = df.duplicated(subset=["상품Code"], keep="first")
                if duplicates.any():
                    logger.warning("Found %d duplicate product codes", duplicates.sum())
                    # Keep first occurrence
                    df = df[~duplicates]
            else:
                logger.warning("'상품Code' column not found, skipping duplicate detection.")

        return df

    except FileNotFoundError:
        logger.error("Input file not found during processing: %s", file_path)
        return None
    except (pd.errors.ParserError, ValueError) as e:
        logger.error("Error parsing Excel file during processing: %s", e)
        return None
    except KeyError as e:
        logger.error("Configuration key error during processing: Missing key %s", e)
        return None
    except (AttributeError, TypeError) as e:
        logger.error("Data type or attribute error during preprocessing steps: %s", e)
        return None
    except Exception as e:
        logger.error("Unexpected error processing input file: %s", e)
        return None


def normalize_column_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize column data types for consistent processing.

    Args:
        df: Input DataFrame

    Returns:
        DataFrame with normalized data types
    """
    try:
        # Create a copy to avoid modifying the original
        normalized_df = df.copy()

        # Normalize string columns
        string_cols = [
            "상품명",
            "담당자",
            "업체명",
            "업체코드",
            "중분류카테고리",
            "상품Code",
            "본사상품링크",
            "구분",
        ]
        for col in string_cols:
            if col in normalized_df.columns:
                # Ensure column is actually string before replace
                if not pd.api.types.is_string_dtype(normalized_df[col]):
                    normalized_df[col] = normalized_df[col].astype(str)
                normalized_df[col] = normalized_df[col].replace("nan", "")

        # Ensure price/numeric columns are floats
        numeric_cols = ["판매단가(V포함)", "기본수량(1)"]
        for col in numeric_cols:
            if col in normalized_df.columns:
                normalized_df[col] = pd.to_numeric(normalized_df[col], errors="coerce")
                # Fill NaN resulting from coercion with 0
                normalized_df[col] = normalized_df[col].fillna(0)

        return normalized_df

    except (TypeError, ValueError, KeyError) as e:
        logger.error("Error normalizing column types: %s", e)
        return df  # Return original if error occurs
    except Exception as e:
        logger.error("Unexpected error normalizing column types: %s", e)
        return df


def handle_missing_values(
    df: pd.DataFrame, required_columns: List[str]
) -> pd.DataFrame:
    """
    Handle missing values in the DataFrame.

    Args:
        df: Input DataFrame
        required_columns: List of columns that must have values

    Returns:
        DataFrame with missing values handled
    """
    try:
        # Create a copy to avoid modifying the original
        handled_df = df.copy()

        # Define column types for appropriate filling
        string_like_cols = [
            "상품명", "담당자", "업체명", "업체코드", "중분류카테고리",
            "상품Code", "본사상품링크", "구분"
        ]
        numeric_like_cols = ["판매단가(V포함)", "기본수량(1)"]

        # Fill missing values based on type
        for col in handled_df.columns:
            if col in string_like_cols:
                handled_df[col] = handled_df[col].fillna("")
            elif col in numeric_like_cols:
                handled_df[col] = handled_df[col].fillna(0)
            # Add handling for other types if necessary

        # Ensure required columns (after potential fillna) exist and handle if not
        # Note: normalize_column_types should ideally ensure these exist
        # This is an extra safety check
        for col in required_columns:
            if col not in handled_df.columns:
                logger.warning("Required column '%s' still missing after handling NaNs. Adding empty.", col)
                if col in string_like_cols:
                    handled_df[col] = ""
                elif col in numeric_like_cols:
                    handled_df[col] = 0
                # Add default for other types if needed

        # Special handling for '구분' column
        if "구분" in handled_df.columns:
            # Default to 'A' (approval management) if invalid or missing
            handled_df["구분"] = handled_df["구분"].apply(
                lambda x: "A" if not isinstance(x, str) or x not in ["A", "P"] else x
            ).astype(str) # Ensure it's string type

        return handled_df

    except (KeyError, TypeError, ValueError) as e:
        logger.error("Error handling missing values: %s", e)
        return df  # Return original if error occurs
    except Exception as e:
        logger.error("Unexpected error handling missing values: %s", e)
        return df


def split_large_file(
    file_path: str, threshold: int = 300, clean_names: bool = True
) -> List[str]:
    """
    Split large Excel file into smaller chunks according to the manual requirements.

    Args:
        file_path: Path to the input Excel file
        threshold: Maximum rows per output file
        clean_names: Whether to clean product names

    Returns:
        List of paths to the split files (or original path if no split needed or error)
    """
    split_files = [file_path] # Default return value
    try:
        logger.info("Attempting to split file: %s (threshold: %d)", file_path, threshold)

        # Read the input file
        df = pd.read_excel(file_path)

        total_rows = len(df)
        if total_rows <= threshold:
            logger.info("File size is below threshold, no splitting needed.")
            return split_files

        # Clean product names if requested
        if clean_names:
            df = clean_product_names(df)

        # Get processing type ('A' for approval, 'P' for price)
        processing_type = "승인관리"
        if "구분" in df.columns and not df.empty and df["구분"].iloc[0].upper() == "P":
            processing_type = "가격관리"

        # Get current date for filename
        current_date = datetime.now().strftime("%Y%m%d")

        # Get base directory and filename
        base_dir = os.path.dirname(file_path)
        base_name, ext = os.path.splitext(os.path.basename(file_path))

        # Calculate number of files needed
        num_files = (total_rows + threshold - 1) // threshold  # Ceiling division

        split_files = []
        for i in range(num_files):
            start_idx = i * threshold
            end_idx = min((i + 1) * threshold, total_rows)
            chunk_df = df.iloc[start_idx:end_idx]

            # Create filename as per manual: 승인관리1(300)-날짜
            file_count = i + 1
            file_size = len(chunk_df)
            split_filename = (
                f"{processing_type}{file_count}({file_size})-{current_date}{ext}"
            )
            split_path = os.path.join(base_dir, split_filename)

            # Save the chunk
            chunk_df.to_excel(split_path, index=False)
            split_files.append(split_path)

            logger.info("Created split file: %s with %d rows", split_path, file_size)

        logger.info("Successfully split file into %d parts.", len(split_files))
        return split_files

    except FileNotFoundError:
        logger.error("Input file not found for splitting: %s", file_path)
        return [file_path] # Return original path
    except (pd.errors.ParserError, ValueError) as e:
        logger.error("Error parsing Excel file during split: %s", e)
        return [file_path] # Return original path
    except (OSError, IOError) as e:
        logger.error("File system error during splitting (read/write): %s", e)
        return [file_path] # Return original path
    except (KeyError, IndexError) as e:
        logger.error("Data access error (e.g., column '구분' missing or empty df): %s", e)
        return [file_path] # Return original path
    except Exception as e:
        logger.error("Unexpected error splitting file: %s", e)
        return [file_path]  # Return original file path if any error occurs


def merge_result_files(file_paths: List[str], original_input: str) -> Optional[str]:
    """
    Merge multiple result files into a single file according to the manual requirements.

    Args:
        file_paths: List of paths to the result files
        original_input: Path to the original input file

    Returns:
        Path to the merged file or None if error
    """
    if not file_paths:
        logger.warning("No result files provided for merging.")
        return None

    logger.info("Attempting to merge %d result files", len(file_paths))

    merged_filename = None
    try:
        # Verify files exist
        for file_path in file_paths:
            if not os.path.exists(file_path):
                logger.error("Result file not found for merging: %s", file_path)
                return None

        # Read all files into DataFrames
        dfs = []
        for file_path in file_paths:
            df = pd.read_excel(file_path)
            dfs.append(df)

        # Concatenate DataFrames
        if not dfs:
            logger.error("No dataframes could be read from result files.")
            return None

        merged_df = pd.concat(dfs, ignore_index=True)

        # Generate output filename
        base_dir = os.path.dirname(original_input)
        base_name, ext = os.path.splitext(os.path.basename(original_input))
        # Add timestamp to avoid overwriting potentially existing merged file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_filename = os.path.join(base_dir, f"{base_name}-merged-result_{timestamp}{ext}")

        # Save merged file
        merged_df.to_excel(merged_filename, index=False)

        logger.info(
            "Created merged file (pre-formatting): %s with %d rows", merged_filename, len(merged_df)
        )

        # Apply formatting (in a separate try-except block)
        try:
            wb = load_workbook(merged_filename)
            ws = wb.active

            # Format header
            header_font = ws.cell(row=1, column=1).font # Get default font
            header_font = header_font.copy(bold=True) # Create bold copy
            for col in range(1, ws.max_column + 1):
                ws.cell(row=1, column=col).font = header_font

            # Apply yellow highlighting for price differences
            yellow_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            price_diff_cols = ["가격차이(2)", "가격차이(3)"]
            col_indices = {ws.cell(row=1, column=j+1).value: j for j in range(ws.max_column)}

            for row_idx in range(2, ws.max_row + 1):
                highlight_row = False
                for col_name in price_diff_cols:
                    col_idx = col_indices.get(col_name)
                    if col_idx is not None:
                        cell = ws.cell(row=row_idx, column=col_idx + 1)
                        if isinstance(cell.value, (int, float)) and cell.value < 0:
                            highlight_row = True
                            break
                if highlight_row:
                    for col_idx in range(ws.max_column):
                        ws.cell(row=row_idx, column=col_idx + 1).fill = yellow_fill

            wb.save(merged_filename)
            logger.info("Applied formatting to merged file: %s", merged_filename)

        except ImportError:
            logger.error("openpyxl or PatternFill not found. Cannot apply formatting.")
            # Return unformatted file path
        except Exception as e:
            logger.error("Error applying formatting to merged file: %s. Returning unformatted file.", e)
            # Continue with unformatted file path

        return merged_filename

    except FileNotFoundError as e:
        logger.error("Result file not found during merging read: %s", e)
        return None
    except pd.errors.EmptyDataError:
        logger.error("One of the result files is empty.")
        return None
    except (pd.errors.ParserError, ValueError) as e:
        logger.error("Error parsing an Excel file during merge: %s", e)
        return None
    except (OSError, IOError) as e:
        logger.error("File system error during merging (read/write): %s", e)
        # Clean up partially created merged file if it exists
        if merged_filename and os.path.exists(merged_filename):
            try:
                os.remove(merged_filename)
                logger.info("Removed partially created file: %s", merged_filename)
            except OSError as rm_err:
                logger.error("Failed to remove partially created file %s: %s", merged_filename, rm_err)
        return None
    except Exception as e:
        logger.error("Unexpected error merging result files: %s", e)
        # Clean up partially created merged file if it exists
        if merged_filename and os.path.exists(merged_filename):
            try:
                os.remove(merged_filename)
                logger.info("Removed partially created file: %s", merged_filename)
            except OSError as rm_err:
                logger.error("Failed to remove partially created file %s: %s", merged_filename, rm_err)
        return None
