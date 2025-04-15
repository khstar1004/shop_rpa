"""
Excel 파일 처리를 위한 통합 유틸리티 모듈

Excel 파일 읽기, 쓰기, 변환 및 데이터 처리에 관련된 기능을 제공합니다.
특히 1차 파일과 2차 파일 변환 프로세스에 최적화되어 있습니다.
"""

import logging
import os
import re
import time
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional, Union, Tuple, Any
from pathlib import Path
import asyncio
import nest_asyncio

# Add the project root directory to the Python path
project_root = str(Path(__file__).parent.parent.absolute())
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# Import scrapers
from core.scraping.haeoeum_scraper import HaeoeumScraper
from core.scraping.koryo_scraper import KoryoScraper
from core.scraping.naver_crawler import NaverShoppingAPI

# 기본 로거 설정
logger = logging.getLogger(__name__)

# --- 스타일 상수 정의 ---
DEFAULT_FONT = Font(name="맑은 고딕", size=10)
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER_SIDE = Side(style="thin")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
RED_FONT = Font(name="맑은 고딕", size=10, color="FF0000")
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
RED_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # 연한 빨강
BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 연한 파랑
PURPLE_FILL = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # 연한 보라
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 연한 초록
LIGHT_GRAY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 연한 회색

# --- 데이터 변환 유틸리티 ---

def safe_to_numeric(series, errors='coerce'):
    """
    판다스 Series를 숫자로 변환, 일반적인 비숫자 값 처리
    
    Args:
        series: 변환할 Series
        errors: 에러 처리 방법 ('coerce', 'raise', 'ignore')
    
    Returns:
        변환된 Series
    """
    non_numeric_placeholders = ['#REF!', '', '-', '동일상품 없음', '(공백)', '(공백 or 없음)', 
                                '가격이 범위 내에 없거나 검색된 상품이 없음']
    series = series.replace(non_numeric_placeholders, np.nan)
    
    # 쉼표 제거 (천 단위 구분자)
    if series.dtype == 'object':
        series = series.str.replace(',', '', regex=False)
        
    return pd.to_numeric(series, errors=errors)

def clean_url(url):
    """
    URL 정제 및 이미지 URL 유효성 검사
    
    Args:
        url: 정제할 URL 문자열
    
    Returns:
        정제된 URL 문자열
    """
    if pd.isna(url) or not url:
        return ""
        
    url_str = str(url).strip()
    
    # 이미 IMAGE 함수인 경우 URL 추출
    if url_str.startswith('=IMAGE('):
        img_match = re.search(r'=IMAGE\("([^"]+)"', url_str)
        if img_match:
            url_str = img_match.group(1)
    
    # Excel 이미지 수식에서 URL 추출 ('=IMAGE("URL", 2)' 형식)
    image_formula_match = re.search(r'=IMAGE\("([^"]+)",\s*\d+\)', url_str)
    if image_formula_match:
        url_str = image_formula_match.group(1)
        
    # 따옴표나 공백 제거
    url_str = url_str.strip('"\'').strip()
    
    # 빈 문자열이면 반환
    if not url_str or url_str.lower() in ['없음', 'none', 'null', 'na', '-']:
        return ""
    
    # http:// 또는 https:// 로 시작하는지 확인
    if url_str and not url_str.lower().startswith(("http://", "https://")):
        url_str = "https://" + url_str
    
    # 이미지 확장자 확인
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
    has_extension = any(url_str.lower().endswith(ext) for ext in image_extensions)
    
    # 일반적인 이미지 호스팅 도메인 확인
    image_domains = ['i.jclgift.com', 'koreagift.com', 'img.danawa.com', 'shopping-phinf.pstatic.net', 
                     'image.gmarket.co.kr', 'image.auction.co.kr', 'thumbnail.image.rakuten.co.jp', 
                     'cdn.pixabay.com', 'images.unsplash.com']
    
    has_image_domain = any(domain in url_str.lower() for domain in image_domains)
    
    # 이미지 파라미터 확인
    has_image_param = any(param in url_str.lower() for param in ['image', 'img', 'photo', 'picture', 'thumbnail', 'thumb'])
    
    # 이미지 URL로 판단되지 않는 경우 로그 남기기 (디버깅용)
    if not (has_extension or has_image_domain or has_image_param) and len(url_str) > 10:
        logger.debug(f"이미지 URL로 판단되지 않음: {url_str}")
    
    return url_str

# --- Excel 파일 입출력 ---

def read_excel_file(file_path: str) -> pd.DataFrame:
    """
    Excel 파일을 읽어 DataFrame으로 반환
    
    Args:
        file_path: Excel 파일 경로
    
    Returns:
        DataFrame
    """
    try:
        df = pd.read_excel(file_path)
        logger.info(f"파일 읽기 완료: {file_path}, {len(df)}행 로드됨")
        return df
    except Exception as e:
        logger.error(f"파일 읽기 오류: {file_path}, {str(e)}", exc_info=True)
        # 빈 DataFrame 반환 (오류 발생 시)
        return pd.DataFrame()

def save_excel_file(df: pd.DataFrame, file_path: str, apply_formatting: bool = True) -> str:
    """
    DataFrame을 Excel 파일로 저장
    
    Args:
        df: 저장할 DataFrame
        file_path: 저장할 파일 경로
        apply_formatting: 기본 서식 적용 여부
    
    Returns:
        저장된 파일 경로
    """
    try:
        # 출력 디렉토리 확인 및 생성
        output_dir = os.path.dirname(file_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        # 이미지 컬럼 처리 (URL을 IMAGE 함수로 변환)
        df_processed = df.copy()
        for col in df_processed.columns:
            col_lower = col.lower()
            # 이미지 관련 컬럼 식별
            if '이미지' in col_lower or 'image' in col_lower:
                # 해당 컬럼의 값을 IMAGE 함수 형식으로 변환
                try:
                    df_processed[col] = df_processed[col].apply(
                        lambda x: f'=IMAGE("{clean_url(x)}", 2)' if pd.notna(x) and clean_url(x) else ""
                    )
                    logger.info(f"이미지 컬럼 '{col}' IMAGE 함수 적용 완료")
                except Exception as img_err:
                    logger.warning(f"이미지 컬럼 '{col}' 처리 중 오류: {img_err}")
        
        # 파일 저장
        df_processed.to_excel(file_path, index=False, engine='openpyxl')
        logger.info(f"파일 저장 완료: {file_path}")
        
        # 서식 적용 (선택적)
        if apply_formatting:
            apply_excel_formatting(file_path)
            
        return file_path
    except Exception as e:
        logger.error(f"파일 저장 오류: {file_path}, {str(e)}", exc_info=True)
        return ""

def apply_excel_formatting(file_path: str) -> bool:
    """
    Excel 파일에 기본 서식 적용
    
    Args:
        file_path: 서식을 적용할 Excel 파일 경로
    
    Returns:
        성공 여부
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 헤더 행 서식 적용
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
            cell.fill = GRAY_FILL
        
        # 이미지 컬럼 식별
        image_cols = []
        for idx, cell in enumerate(ws[1]):
            cell_value = str(cell.value).lower() if cell.value else ""
            if "이미지" in cell_value or "image" in cell_value:
                image_cols.append(idx + 1)  # 1-based 인덱스
        
        # 열 너비 자동 조정 (최소/최대 너비 지정)
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        # IMAGE 함수가 있는 경우 길이 계산 제외
                        if isinstance(cell.value, str) and cell.value.startswith('=IMAGE('):
                            continue
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            col_letter = get_column_letter(col_idx)
            
            # 이미지 컬럼인 경우 넓게 설정
            if col_idx in image_cols:
                ws.column_dimensions[col_letter].width = 20  # 이미지 컬럼 너비 고정
            else:
                adjusted_width = min(max(max_length + 2, 8), 50)  # 최소 8, 최대 50 
                ws.column_dimensions[col_letter].width = adjusted_width
        
        # 행 높이 설정 (이미지를 위해)
        for row_idx in range(2, ws.max_row + 1):  # 헤더 행 제외
            has_image = False
            for col_idx in image_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('=IMAGE('):
                    has_image = True
                    break
            
            if has_image:
                ws.row_dimensions[row_idx].height = 100  # 이미지 행 높이 설정
            else:
                ws.row_dimensions[row_idx].height = 20  # 일반 행 높이
        
        # 특정 컬럼 식별 (가격차이 컬럼)
        price_diff_cols = []
        for idx, cell in enumerate(ws[1]):
            cell_value = str(cell.value).lower() if cell.value else ""
            if "가격차이" in cell_value and "%" not in cell_value:
                price_diff_cols.append(idx + 1)  # 1-based 인덱스
        
        # 상태 컬럼 식별
        status_cols = []
        for idx, cell in enumerate(ws[1]):
            cell_value = str(cell.value).lower() if cell.value else ""
            if "상태" in cell_value:
                status_cols.append(idx + 1)  # 1-based 인덱스
        
        # 각 행의 셀에 서식 적용
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  # 2부터 시작 (헤더 제외)
            for col_idx, cell in enumerate(row, 1):
                # 이미지 셀이 아닌 경우에만 테두리 적용
                if col_idx not in image_cols or not (isinstance(cell.value, str) and cell.value.startswith('=IMAGE(')):
                    cell.border = THIN_BORDER
                
                # 이미지 셀인 경우 중앙 정렬
                if col_idx in image_cols and isinstance(cell.value, str) and cell.value.startswith('=IMAGE('):
                    cell.alignment = CENTER_ALIGNMENT
                    continue
                
                # 가격차이 컬럼이고 음수 값이면 노란색 배경
                if col_idx in price_diff_cols:
                    try:
                        value = cell.value
                        if isinstance(value, (int, float)) and value < 0:
                            cell.fill = YELLOW_FILL
                            cell.font = BOLD_FONT
                        elif isinstance(value, str) and ('-' in value or '－' in value):
                            # 문자열 형태의 음수 처리 ('-123', '－456' 등)
                            cell.fill = YELLOW_FILL
                            cell.font = BOLD_FONT
                    except:
                        pass
                
                # 상태 컬럼의 특정 값에 따른 배경색 설정
                if col_idx in status_cols and cell.value:
                    status_value = str(cell.value).lower()
                    # 상품을 찾을 수 없음 - 빨간색 배경
                    if "찾을 수 없" in status_value or "없음" in status_value:
                        cell.fill = RED_FILL
                    # 유사 이미지가 없음 - 연한 파란색 배경
                    elif "이미지" in status_value and "없음" in status_value:
                        cell.fill = BLUE_FILL
                    # 텍스트 유사도 낮음 - 연한 보라색 배경
                    elif "텍스트" in status_value and "낮음" in status_value:
                        cell.fill = PURPLE_FILL
                    # 오류 - 회색 배경
                    elif "오류" in status_value:
                        cell.fill = LIGHT_GRAY_FILL
                    # 정상 - 연한 초록색 배경
                    elif "정상" in status_value:
                        cell.fill = GREEN_FILL
            
        # 파일 저장
        wb.save(file_path)
        logger.info(f"서식 적용 완료: {file_path}")
        return True
    except Exception as e:
        logger.error(f"서식 적용 오류: {file_path}, {str(e)}", exc_info=True)
        return False

# --- 1차 파일 처리 함수 ---

def process_first_stage(input_file: str, output_dir: Optional[str] = None) -> str:
    """
    1차 엑셀 파일 처리
    
    - 기본 데이터 정리
    - 필요한 열 추가 및 계산 수행
    - 결과를 intermediate 파일로 저장
    
    Args:
        input_file: 입력 파일 경로
        output_dir: 출력 디렉토리 (기본값: output/intermediate)
        
    Returns:
        생성된 1차 처리 파일 경로
    """
    start_time = time.time()
    
    # 출력 디렉토리 설정
    if not output_dir:
        output_dir = os.path.join("output", "intermediate")
    
    # 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # 출력 파일 경로 생성
    base_name = os.path.basename(input_file)
    name, ext = os.path.splitext(base_name)
    output_file = os.path.join(output_dir, f"{name}_intermediate{ext}")
    
    logger.info(f"1차 파일 처리 시작: {input_file}")
    
    try:
        # 엑셀 파일 로드
        df = read_excel_file(input_file)
        
        if df.empty:
            logger.warning(f"파일이 비어 있습니다: {input_file}")
            return input_file
            
        logger.info(f"파일 읽기 완료: {input_file}, {len(df)}행 로드됨")
        
        # === 1차 처리 로직 적용 ===
        
        # 스크래핑 단계 추가
        df = run_scraping_for_excel(df)
        
        # 기존 로직
        df = apply_first_stage_rules(df)
        
        # 결과 저장
        save_excel_file(df, output_file)
        
        # 서식 적용
        apply_excel_formatting(output_file)
        logger.info(f"서식 적용 완료: {output_file}")
        
        # 처리 완료 로그
        elapsed_time = time.time() - start_time
        logger.info(f"1차 파일 처리 완료: {output_file} ({elapsed_time:.2f}초, {len(df)}행 x {len(df.columns)}열)")
        
        return output_file
        
    except Exception as e:
        logger.error(f"1차 파일 처리 중 오류 발생: {str(e)}", exc_info=True)
        return input_file  # 오류 발생 시 원본 파일 반환

def run_scraping_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    데이터프레임에 스크래핑 결과를 적용하는 함수
    
    Args:
        df: 원본 데이터프레임
        
    Returns:
        스크래핑 결과가 적용된 데이터프레임
    """
    logger.info("스크래핑 시작...")
    
    # 스크래퍼 초기화
    haeoeum_scraper = None
    koryo_scraper = None
    naver_scraper = None
    
    try:
        from core.scraping.haeoeum_scraper import HaeoeumScraper
        haeoeum_scraper = HaeoeumScraper(debug=True)
        logger.info("해오름 스크래퍼 초기화 완료")
    except Exception as e:
        logger.error(f"해오름 스크래퍼 초기화 중 오류: {e}")
    
    try:
        from core.scraping.koryo_scraper import KoryoScraper
        koryo_scraper = KoryoScraper(debug=True)
        logger.info("고려 스크래퍼 초기화 완료")
    except Exception as e:
        logger.error(f"고려 스크래퍼 초기화 중 오류: {e}")
    
    try:
        from core.scraping.naver_scraper import NaverScraper
        naver_scraper = NaverScraper(debug=True)
        logger.info("네이버 스크래퍼 초기화 완료")
    except Exception as e:
        logger.error(f"네이버 스크래퍼 초기화 중 오류: {e}")
    
    # 스크래퍼가 하나도 초기화되지 않았으면 종료
    if not haeoeum_scraper and not koryo_scraper and not naver_scraper:
        logger.error("스크래퍼를 초기화할 수 없습니다. 스크래핑을 건너뜁니다.")
        return df
    
    # 스크래핑할 데이터가 없으면 종료
    if df.empty:
        logger.warning("데이터프레임이 비어 있어 스크래핑을 건너뜁니다.")
        return df
    
    # 열 추가
    if haeoeum_scraper:
        if '본사 이미지' not in df.columns:
            df['본사 이미지'] = ''
        if '본사 상태' not in df.columns:
            df['본사 상태'] = ''
    
    if koryo_scraper:
        if '고려기프트 이미지' not in df.columns:
            df['고려기프트 이미지'] = ''
        if '고려_가격' not in df.columns:
            df['고려_가격'] = None
        if '고려기프트 상태' not in df.columns:
            df['고려기프트 상태'] = ''
    
    if naver_scraper:
        if '네이버 이미지' not in df.columns:
            df['네이버 이미지'] = ''
        if '네이버_가격' not in df.columns:
            df['네이버_가격'] = None
        if '네이버_기본수량' not in df.columns:
            df['네이버_기본수량'] = None
        if '네이버 상태' not in df.columns:
            df['네이버 상태'] = ''
    
    # URL 컬럼 식별
    url_cols = {
        'haeoeum': None,
        'koryo': None,
        'naver': None
    }
    
    for col in df.columns:
        col_lower = col.lower()
        if '해오름' in col_lower or 'haeoeum' in col_lower or '본사' in col_lower:
            if 'url' in col_lower or '링크' in col_lower or 'link' in col_lower:
                url_cols['haeoeum'] = col
                break
    
    for col in df.columns:
        col_lower = col.lower()
        if '고려' in col_lower or 'koryo' in col_lower:
            if 'url' in col_lower or '링크' in col_lower or 'link' in col_lower:
                url_cols['koryo'] = col
                break
    
    for col in df.columns:
        col_lower = col.lower()
        if '네이버' in col_lower or 'naver' in col_lower:
            if 'url' in col_lower or '링크' in col_lower or 'link' in col_lower:
                url_cols['naver'] = col
                break
    
    # 로깅
    for site, col in url_cols.items():
        if col:
            logger.info(f"{site} URL 컬럼: {col}")
        else:
            logger.warning(f"{site} URL 컬럼을 찾을 수 없습니다.")
    
    # 각 행에 대해 스크래핑 수행
    for idx, row in df.iterrows():
        # 해오름 스크래핑
        if haeoeum_scraper and url_cols['haeoeum'] in df.columns:
            try:
                url = str(row[url_cols['haeoeum']])
                if pd.notna(url):
                    try:
                        # URL에서 product_idx 추출
                        product_idx = haeoeum_scraper._extract_product_idx(url)
                        if product_idx:
                            # 해오름 기프트에는 상품이 무조건 있다는 대전제에 따라 항상 결과 기대
                            product = haeoeum_scraper.get_product(product_idx)
                            
                            if product:
                                # 항상 이미지가 있음을 보장
                                if product.image_gallery:
                                    df.at[idx, '본사 이미지'] = product.image_gallery[0]
                                    # 상태가 정상임을 표시
                                    df.at[idx, '본사 상태'] = "정상"
                                else:
                                    df.at[idx, '본사 이미지'] = f"{haeoeum_scraper.BASE_URL}/images/no_image.jpg"
                                    logger.warning(f"해오름 상품 {product_idx}의 이미지가 없어 기본 이미지 사용")
                                    df.at[idx, '본사 상태'] = "유사 이미지가 없음"
                                
                                # 상품 상태 추가 정보 확인
                                if hasattr(product, 'status') and product.status == "Fallback":
                                    df.at[idx, '본사 상태'] = "상품 정보 추출 실패"
                                    logger.warning(f"해오름 상품 {product_idx}의 정보 추출 실패, 대체 정보 사용")
                            else:
                                # 상품 객체가 없더라도 대전제에 따라 기본 이미지 설정
                                df.at[idx, '본사 이미지'] = f"{haeoeum_scraper.BASE_URL}/images/no_image.jpg"
                                logger.warning(f"해오름 상품 {product_idx}을(를) 찾을 수 없어 기본 이미지 사용")
                                df.at[idx, '본사 상태'] = "상품을 찾을 수 없음"
                        else:
                            df.at[idx, '본사 이미지'] = f"{haeoeum_scraper.BASE_URL}/images/no_image.jpg"
                            logger.warning(f"해오름 상품 ID 추출 실패: {url}, 기본 이미지 사용")
                            df.at[idx, '본사 상태'] = "상품 ID 추출 실패"
                    except Exception as e:
                        logger.error(f"해오름 product_idx 추출 오류: {e}, URL: {url}")
                        df.at[idx, '본사 이미지'] = f"{haeoeum_scraper.BASE_URL}/images/no_image.jpg"
                        df.at[idx, '본사 상태'] = "상품 정보 추출 오류"
            except Exception as e:
                logger.error(f"해오름 스크래핑 오류 (행 {idx}): {e}")
                df.at[idx, '본사 이미지'] = f"{haeoeum_scraper.BASE_URL}/images/no_image.jpg"
                df.at[idx, '본사 상태'] = "처리 오류 발생"
                
        # 고려 스크래핑
        if koryo_scraper and url_cols['koryo'] in df.columns:
            try:
                url = str(row[url_cols['koryo']])
                if pd.notna(url):
                    # 상품명 가져오기
                    product_name = ""
                    if '상품명' in df.columns and pd.notna(row['상품명']):
                        product_name = str(row['상품명'])
                    
                    if product_name:
                        logger.info(f"고려 상품 검색: {product_name}")
                        try:
                            # 비동기 메서드 실행 - 고려 상품은 없을 수 있음
                            products = run_async_func(koryo_scraper.search_product, product_name, max_items=1)
                            
                            if products and len(products) > 0:
                                product = products[0]
                                
                                # 안전하게 이미지 URL 추출
                                try:
                                    if hasattr(product, 'image_gallery') and product.image_gallery:
                                        df.at[idx, '고려기프트 이미지'] = product.image_gallery[0] if isinstance(product.image_gallery, (list, tuple)) else ""
                                    else:
                                        df.at[idx, '고려기프트 이미지'] = ""
                                        # 이미지가 없는 경우 메시지 추가
                                        df.at[idx, '고려기프트 상태'] = "유사 이미지가 없음"
                                        logger.warning(f"고려 상품 {product_name}의 이미지가 없음")
                                except (IndexError, TypeError) as e:
                                    logger.warning(f"고려 이미지 추출 오류: {e}")
                                    df.at[idx, '고려기프트 이미지'] = ""
                                    df.at[idx, '고려기프트 상태'] = "유사 이미지가 없음"
                                
                                # 가격 정보 추출
                                if hasattr(product, 'price') and pd.notna(product.price):
                                    df.at[idx, '고려_가격'] = product.price
                                    logger.info(f"고려 상품 추출 성공: {product.name}, 가격: {product.price}")
                                
                                # 유사도 정보가 있으면 체크
                                if hasattr(product, 'similarity_score') and pd.notna(product.similarity_score):
                                    similarity = product.similarity_score
                                    if similarity < 0.6:  # 낮은 텍스트 유사도 기준값
                                        df.at[idx, '고려기프트 상태'] = "텍스트 유사도 낮음"
                                        logger.warning(f"고려 상품 {product_name}의 텍스트 유사도가 낮음: {similarity:.2f}")
                            else:
                                logger.warning(f"고려 상품 검색 결과 없음: {product_name}")
                                # 상품이 없을 수 있음을 허용 - 상태 메시지 추가
                                df.at[idx, '고려기프트 상태'] = "상품을 찾을 수 없음"
                        except Exception as e:
                            logger.error(f"고려 상품 검색 중 오류: {e}")
                            df.at[idx, '고려기프트 상태'] = "검색 오류 발생"
                    else:
                        logger.warning("고려 스크래핑: 상품명이 비어 있어 검색 불가")
                        df.at[idx, '고려기프트 상태'] = "상품명 정보 없음"
            except Exception as e:
                logger.error(f"고려 스크래핑 오류 (행 {idx}): {e}")
                df.at[idx, '고려기프트 상태'] = "처리 오류 발생"
                
        # 네이버 스크래핑
        if naver_scraper and url_cols['naver'] in df.columns:
            try:
                url = str(row[url_cols['naver']])
                if pd.notna(url):
                    # 상품명 가져오기
                    product_name = ""
                    if '상품명' in df.columns and pd.notna(row['상품명']):
                        product_name = str(row['상품명'])
                    
                    if product_name:
                        logger.info(f"네이버 상품 검색: {product_name}")
                        try:
                            # 네이버 스크래퍼의 search_product가 비동기인 경우를 대비해 같은 방식으로 처리
                            if asyncio.iscoroutinefunction(naver_scraper.search_product):
                                products = run_async_func(naver_scraper.search_product, product_name, max_items=1)
                            else:
                                products = naver_scraper.search_product(product_name, max_items=1)
                            
                            if products and len(products) > 0:
                                product = products[0]
                                
                                # 안전하게 이미지 URL 추출
                                try:
                                    if hasattr(product, 'image_gallery') and product.image_gallery:
                                        df.at[idx, '네이버 이미지'] = product.image_gallery[0] if isinstance(product.image_gallery, (list, tuple)) else ""
                                    else:
                                        df.at[idx, '네이버 이미지'] = ""
                                        # 이미지가 없는 경우 메시지 추가
                                        df.at[idx, '네이버 상태'] = "유사 이미지가 없음"
                                        logger.warning(f"네이버 상품 {product_name}의 이미지가 없음")
                                except (IndexError, TypeError) as e:
                                    logger.warning(f"네이버 이미지 추출 오류: {e}")
                                    df.at[idx, '네이버 이미지'] = ""
                                    df.at[idx, '네이버 상태'] = "유사 이미지가 없음"
                                
                                # 가격 정보 추출
                                if hasattr(product, 'price') and pd.notna(product.price):
                                    df.at[idx, '네이버_가격'] = product.price
                                
                                # 기본수량 정보 추출
                                if hasattr(product, 'min_order_quantity') and pd.notna(product.min_order_quantity):
                                    df.at[idx, '네이버_기본수량'] = product.min_order_quantity
                                
                                logger.info(f"네이버 상품 추출 성공: {product.name}, 가격: {product.price}")
                                
                                # 유사도 정보가 있으면 체크
                                if hasattr(product, 'similarity_score') and pd.notna(product.similarity_score):
                                    similarity = product.similarity_score
                                    if similarity < 0.6:  # 낮은 텍스트 유사도 기준값
                                        df.at[idx, '네이버 상태'] = "텍스트 유사도 낮음"
                                        logger.warning(f"네이버 상품 {product_name}의 텍스트 유사도가 낮음: {similarity:.2f}")
                            else:
                                logger.warning(f"네이버 상품 검색 결과 없음: {product_name}")
                                # 상품이 없을 수 있음을 허용 - 상태 메시지 추가
                                df.at[idx, '네이버 상태'] = "상품을 찾을 수 없음"
                        except Exception as e:
                            logger.error(f"네이버 상품 검색 중 오류: {e}")
                            df.at[idx, '네이버 상태'] = "검색 오류 발생"
                    else:
                        logger.warning("네이버 스크래핑: 상품명이 비어 있어 검색 불가")
                        df.at[idx, '네이버 상태'] = "상품명 정보 없음"
            except Exception as e:
                logger.error(f"네이버 스크래핑 오류 (행 {idx}): {e}")
                df.at[idx, '네이버 상태'] = "처리 오류 발생"
    
    # 데이터 타입 변환 (숫자형 컬럼)
    numeric_cols = ['고려_가격', '네이버_가격']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    logger.info("데이터 스크래핑 완료")
    return df

# 1차 처리 규칙 적용 함수 추가
def apply_first_stage_rules(df: pd.DataFrame) -> pd.DataFrame:
    """
    1차 처리 규칙을 데이터프레임에 적용합니다.
    
    Args:
        df: 원본 데이터프레임
        
    Returns:
        규칙이 적용된 데이터프레임
    """
    logger.info("1차 처리 규칙 적용 중...")
    
    # 1. 가격차이 컬럼 생성
    if '고려_가격' in df.columns and '네이버_가격' in df.columns:
        df['고려_가격차이'] = df['고려_가격'] - df['네이버_가격']
        df['네이버_가격차이'] = df['네이버_가격'] - df['고려_가격']
        
        # 가격차이 퍼센트 계산
        if '고려_가격차이' in df.columns and '네이버_가격' in df.columns:
            df['고려_가격차이_퍼센트'] = (df['고려_가격차이'] * 100 / df['네이버_가격']).round(2)
        
        if '네이버_가격차이' in df.columns and '고려_가격' in df.columns:
            df['네이버_가격차이_퍼센트'] = (df['네이버_가격차이'] * 100 / df['고려_가격']).round(2)
        
        logger.info("가격차이 및 퍼센트 계산 완료")
    else:
        missing_cols = []
        if '고려_가격' not in df.columns:
            missing_cols.append('고려_가격')
        if '네이버_가격' not in df.columns:
            missing_cols.append('네이버_가격')
        logger.warning(f"가격차이 계산에 필요한 컬럼이 없습니다: {', '.join(missing_cols)} (현재 컬럼: {', '.join(df.columns)})")
    
    # 2. 기본수량 컬럼 확인
    if '네이버_기본수량' not in df.columns:
        df['네이버_기본수량'] = pd.NA
        logger.info("네이버_기본수량 컬럼 생성 완료")
    
    return df

# --- 2차 파일 처리 함수 ---

def process_second_stage(input_file: str, output_dir: Optional[str] = None) -> str:
    """
    1차 처리 파일을 입력으로 2차 처리 실행
    
    Args:
        input_file: 1차 처리 파일 경로
        output_dir: 출력 디렉토리 (기본값: output/final)
        
    Returns:
        생성된 2차 처리 파일 경로
    """
    start_time = time.time()
    
    # 출력 디렉토리 설정
    if not output_dir:
        output_dir = os.path.join("output", "final")
    
    # 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # 출력 파일 경로 생성
    base_name = os.path.basename(input_file)
    name, ext = os.path.splitext(base_name)
    # intermediate 접미사 제거
    name = name.replace("_intermediate", "")
    output_file = os.path.join(output_dir, f"{name}_final{ext}")
    
    logger.info(f"2차 파일 처리 시작: {input_file}")
    
    try:
        # 1차 파일 읽기
        df = read_excel_file(input_file)
        
        if df.empty:
            logger.warning(f"1차 파일이 비어 있습니다: {input_file}")
            # 비어있는 경우 기본 구조만 있는 파일 생성
            default_df = pd.DataFrame({
                '상품명': ['처리할 데이터가 없습니다'],
                '처리 상태': ['입력 파일에 유효한 데이터가 없습니다'],
                '처리 시간': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            })
            save_excel_file(default_df, output_file)
            logger.info(f"대체 파일 생성 완료: {output_file}")
            return output_file
        
        # 2차 처리 규칙 적용
        df_processed = apply_second_stage_rules(df)
        
        if df_processed.empty:
            logger.warning("2차 처리 후 데이터가 비어 있습니다. 기본 행 추가")
            # 원본 데이터에서 기본 컬럼만 선택
            if len(df.columns) > 0:
                # 처음 10개 컬럼 선택 
                first_cols = list(df.columns[:min(10, len(df.columns))])
                df_processed = df[first_cols].head(1).copy()
                # 처리 상태 컬럼 추가
                df_processed['처리 상태'] = '필터링 후 데이터 없음'
            else:
                # 완전히 빈 경우 기본 구조만 있는 행 추가
                df_processed = pd.DataFrame({
                    '상품명': ['필터링 후 데이터가 없습니다'],
                    '처리 상태': ['필터 기준을 충족하는 데이터가 없습니다'],
                    '처리 시간': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                })
        
        # 2차 파일 저장
        save_excel_file(df_processed, output_file)
        
        # 서식 적용
        apply_excel_formatting(output_file)
        
        # 처리 완료 로그
        elapsed_time = time.time() - start_time
        logger.info(f"2차 파일 처리 완료: {output_file} ({elapsed_time:.2f}초, {len(df_processed)}행 x {len(df_processed.columns)}열)")
        
        return output_file
        
    except Exception as e:
        logger.error(f"2차 파일 처리 중 오류 발생: {str(e)}", exc_info=True)
        
        # 오류 발생 시 기본 파일 생성
        try:
            error_df = pd.DataFrame({
                '오류 발생': ['2차 처리 중 오류가 발생했습니다'],
                '오류 메시지': [str(e)],
                '처리 시간': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            })
            save_excel_file(error_df, output_file)
            logger.info(f"오류 보고 파일 생성 완료: {output_file}")
            return output_file
        except Exception as inner_e:
            logger.error(f"오류 보고 파일 생성 실패: {str(inner_e)}")
            return ""

def apply_second_stage_rules(df):
    """
    2차 처리 규칙 적용
    
    규칙:
    1. 노란색 표시 상품 이동: 가격차이가 음수인 상품만 선택
    2. 네이버쇼핑 상품 중 기본수량이 없고 가격차이가 10% 이하인 경우 제외
    3. 고려기프트 상품 중 가격차이가 1% 이하인 경우 제외
    4. 가격차이가 양수(+)인 상품 제외 
    5. 고려기프트와 네이버쇼핑에 가격불량상품이 전혀 없는 경우 해당 줄 제외
    
    Args:
        df: 입력 데이터프레임
        
    Returns:
        규칙이 적용된 데이터프레임
    """
    logger.info("2차 파일 처리 규칙 적용 중...")
    
    # 비어있으면 그대로 반환
    if df.empty:
        logger.warning("입력 데이터프레임이 비었습니다.")
        return df
    
    # 원본 데이터 백업
    original_df = df.copy()
    start_count = len(df)
    logger.info(f"초기 행 수: {start_count}")
    
    try:
        # 1. 가격차이가 음수인 상품 선택 (노란색 셀로 표시된 상품)
        price_diff_found = False
        
        # 가격차이 컬럼 식별
        koryo_diff_col = None  # 고려 가격차이
        naver_diff_col = None  # 네이버 가격차이
        koryo_diff_pct_col = None  # 고려 가격차이 퍼센트
        naver_diff_pct_col = None  # 네이버 가격차이 퍼센트
        
        for col in df.columns:
            col_lower = col.lower()
            if ('고려' in col_lower or '기프트' in col_lower or 'koryo' in col_lower) and '가격차이' in col_lower and '%' not in col_lower:
                koryo_diff_col = col
            elif ('네이버' in col_lower or 'naver' in col_lower) and '가격차이' in col_lower and '%' not in col_lower:
                naver_diff_col = col
            elif ('고려' in col_lower or '기프트' in col_lower or 'koryo' in col_lower) and '가격차이' in col_lower and '%' in col_lower:
                koryo_diff_pct_col = col
            elif ('네이버' in col_lower or 'naver' in col_lower) and '가격차이' in col_lower and '%' in col_lower:
                naver_diff_pct_col = col
        
        # 고려기프트 기본수량 컬럼 식별
        koryo_base_qty_col = None
        for col in df.columns:
            col_lower = col.lower()
            if ('고려' in col_lower or '기프트' in col_lower or 'koryo' in col_lower) and ('기본수량' in col_lower or 'base' in col_lower):
                koryo_base_qty_col = col
                break
                
        # 네이버 기본수량 컬럼 식별
        naver_base_qty_col = None
        for col in df.columns:
            col_lower = col.lower()
            if ('네이버' in col_lower or 'naver' in col_lower) and ('기본수량' in col_lower or 'base' in col_lower):
                naver_base_qty_col = col
                break
                
        if koryo_diff_col:
            logger.info(f"고려 가격차이 컬럼 식별: {koryo_diff_col}")
            price_diff_found = True
        else:
            logger.warning("고려 가격차이 컬럼을 찾을 수 없습니다.")
            
        if naver_diff_col:
            logger.info(f"네이버 가격차이 컬럼 식별: {naver_diff_col}")
            price_diff_found = True
        else:
            logger.warning("네이버 가격차이 컬럼을 찾을 수 없습니다.")
            
        if price_diff_found:
            # 가격차이가 음수인 상품 선택
            price_diff_mask = pd.Series(False, index=df.index)
            
            if koryo_diff_col:
                # 숫자 형태의 가격차이 처리
                if df[koryo_diff_col].dtype in (int, float):
                    price_diff_mask = price_diff_mask | (df[koryo_diff_col] < 0)
                else:
                    # 문자열에서 숫자 추출 시도
                    try:
                        numeric_values = pd.to_numeric(df[koryo_diff_col], errors='coerce')
                        price_diff_mask = price_diff_mask | (numeric_values < 0)
                    except:
                        # '-' 문자로 시작하는 값 처리
                        price_diff_mask = price_diff_mask | df[koryo_diff_col].astype(str).str.startswith('-')
            
            if naver_diff_col:
                # 숫자 형태의 가격차이 처리
                if df[naver_diff_col].dtype in (int, float):
                    price_diff_mask = price_diff_mask | (df[naver_diff_col] < 0)
                else:
                    # 문자열에서 숫자 추출 시도
                    try:
                        numeric_values = pd.to_numeric(df[naver_diff_col], errors='coerce')
                        price_diff_mask = price_diff_mask | (numeric_values < 0)
                    except:
                        # '-' 문자로 시작하는 값 처리
                        price_diff_mask = price_diff_mask | df[naver_diff_col].astype(str).str.startswith('-')
            
            # 가격차이가 음수인 상품만 선택
            before_count = len(df)
            df = df[price_diff_mask].copy()
            filtered_count = before_count - len(df)
            logger.info(f"가격차이가 음수인 상품 선택: {len(df)}개 (제외: {filtered_count}개)")
            
            # 모든 필터링 후 데이터가 없으면, 다른 규칙을 적용하기 전에 처리
            if df.empty:
                logger.warning("가격차이가 음수인 상품이 없습니다. 원본에서 일부 행 선택")
                # 원본 데이터에서 처음 몇개 행을 선택하고 상태 표시
                if not original_df.empty:
                    df = original_df.head(3).copy()
                    logger.info(f"원본에서 선택한 행: {len(df)}개")
                    if '처리 상태' not in df.columns:
                        df['처리 상태'] = '가격차이 음수 상품 없음'
                return df
        else:
            logger.warning("가격차이 컬럼을 찾을 수 없어 원본 데이터를 그대로 사용합니다.")
        
        # 2. 네이버쇼핑 상품 중 기본수량이 없고 가격차이가 10% 이하인 경우 제외
        if naver_base_qty_col and naver_diff_pct_col:
            before_count = len(df)
            
            # 네이버 기본수량이 없고 가격차이 퍼센트가 10% 이하인 상품 필터링
            # 기본수량 NULL 여부 확인 (문자열 'null', '없음', '-', 빈 문자열 포함)
            is_null_qty = df[naver_base_qty_col].isna() | \
                         (df[naver_base_qty_col].astype(str).str.lower().isin(['null', 'none', 'nan', '없음', '-', '']))
            
            # 가격차이 % 를 숫자로 변환해 10% 이하 확인
            try:
                naver_pct_values = pd.to_numeric(df[naver_diff_pct_col].astype(str).str.replace('%', ''), errors='coerce')
                is_small_diff = (naver_pct_values.abs() <= 10)
                
                # 둘 다 해당하는 행 제외
                filter_mask = ~(is_null_qty & is_small_diff)
                df = df[filter_mask].copy()
                
                filtered_count = before_count - len(df)
                logger.info(f"네이버 기본수량 없고 가격차이 10% 이하 상품 제외: {filtered_count}개")
            except Exception as e:
                logger.error(f"네이버 가격차이 퍼센트 처리 중 오류: {e}")
        
        # 3. 고려기프트 상품 중 가격차이가 1% 이하인 경우 제외
        if koryo_diff_pct_col:
            before_count = len(df)
            
            try:
                # 가격차이 % 를 숫자로 변환해 1% 이하 확인
                koryo_pct_values = pd.to_numeric(df[koryo_diff_pct_col].astype(str).str.replace('%', ''), errors='coerce')
                is_small_diff = (koryo_pct_values.abs() <= 1)
                
                # 해당하는 행 제외
                df = df[~is_small_diff].copy()
                
                filtered_count = before_count - len(df)
                logger.info(f"고려기프트 가격차이 1% 이하 상품 제외: {filtered_count}개")
            except Exception as e:
                logger.error(f"고려기프트 가격차이 퍼센트 처리 중 오류: {e}")
        
        # 4. 가격차이가 양수(+)인 상품 제외 (혹시 남아있다면)
        if koryo_diff_col or naver_diff_col:
            before_count = len(df)
            positive_diff_mask = pd.Series(False, index=df.index)
            
            if koryo_diff_col:
                # 숫자 형태의 가격차이 처리
                try:
                    koryo_values = pd.to_numeric(df[koryo_diff_col], errors='coerce')
                    positive_diff_mask = positive_diff_mask | (koryo_values > 0)
                except:
                    pass
            
            if naver_diff_col:
                # 숫자 형태의 가격차이 처리
                try:
                    naver_values = pd.to_numeric(df[naver_diff_col], errors='coerce')
                    positive_diff_mask = positive_diff_mask | (naver_values > 0)
                except:
                    pass
            
            # 가격차이가 양수인 상품 제외
            df = df[~positive_diff_mask].copy()
            filtered_count = before_count - len(df)
            logger.info(f"가격차이가 양수인 상품 제외: {filtered_count}개")
        
        # 5. 고려기프트와 네이버쇼핑 둘 다에 가격불량 상품이 없는 행(줄) 삭제
        if koryo_diff_col and naver_diff_col:
            before_count = len(df)
            
            # 고려기프트와 네이버 모두 가격불량(음수)인지 확인
            has_koryo_issue = pd.Series(False, index=df.index)
            has_naver_issue = pd.Series(False, index=df.index)
            
            try:
                # 고려기프트 가격불량 확인
                koryo_values = pd.to_numeric(df[koryo_diff_col], errors='coerce')
                has_koryo_issue = koryo_values < 0
            except:
                pass
                
            try:
                # 네이버 가격불량 확인
                naver_values = pd.to_numeric(df[naver_diff_col], errors='coerce')
                has_naver_issue = naver_values < 0
            except:
                pass
            
            # 둘 중 하나라도 가격불량이 있는 행만 선택
            df = df[has_koryo_issue | has_naver_issue].copy()
            filtered_count = before_count - len(df)
            logger.info(f"고려/네이버 모두 가격불량 없는 상품 제외: {filtered_count}개")
        
        # 6. 선택된 컬럼만 남기기
        columns_to_keep = [
            '구 분', '담당자', '업체명', '상품명',
            '판매단가(V포함)', '기본수량(1)', '본사상품링크', '본사 이미지', '본사 상태',
            '고려_가격', '고려_가격차이', '고려_가격차이_퍼센트', '고려기프트 상품링크', '고려기프트 이미지', '고려기프트 상태',
            '네이버_가격', '네이버_가격차이', '네이버_가격차이_퍼센트', '네이버_기본수량', '네이버 쇼핑 링크', '네이버 이미지', '네이버 상태'
        ]
        
        # 존재하는 컬럼만 선택
        columns_to_keep = [col for col in columns_to_keep if col in df.columns]
        
        # 만약 필수 컬럼이 없다면 원본 데이터의 컬럼으로 대체 후 로그 출력
        if not columns_to_keep:
            logger.warning("필수 컬럼을 찾을 수 없습니다. 원본 데이터의 모든 컬럼을 유지합니다.")
            columns_to_keep = list(df.columns)
        
        # 최소한 몇 개의 컬럼은 필요함
        if len(columns_to_keep) < 5 and len(original_df.columns) > 0:
            logger.warning("선택된 컬럼이 5개 미만입니다. 원본 데이터에서 추가 컬럼을 선택합니다.")
            # 기본 컬럼이 부족하면 원본 데이터에서 처음 15개 컬럼 선택
            columns_to_keep = list(original_df.columns[:min(15, len(original_df.columns))])
        
        # 선택된 컬럼만 유지
        df = df[columns_to_keep].copy()
        
        # 7. 결과가 비어있는 경우 처리
        if df.empty:
            logger.warning("모든 필터링 후 데이터가 비어 있습니다. 결과 표시용 행 추가")
            
            # 비어있는 결과 표시용 행 추가
            result_row = pd.DataFrame({
                col: ["가격차이 음수 상품이 없거나 모든 필터링 규칙으로 인해 제외됨"] 
                     if col == "상품명" else [""] for col in columns_to_keep
            })
            
            # 처리 상태 추가 (있을 경우)
            if '처리 상태' in columns_to_keep:
                result_row['처리 상태'] = ["모든 필터링 후 데이터 없음"]
                
            # 처리 시간 추가 (있을 경우)
            if '처리 시간' in columns_to_keep:
                result_row['처리 시간'] = [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                
            df = result_row
        
        logger.info(f"최종 2차 파일 출력 결과: {len(df)}행, {len(df.columns)}열")
        logger.info(f"초기 {start_count}행에서 {len(df)}행으로 필터링 ({start_count - len(df)}행 제외)")
        
        return df
        
    except Exception as e:
        logger.error(f"2차 파일 처리 규칙 적용 중 오류 발생: {str(e)}", exc_info=True)
        # 오류 발생 시 원본 최소 처리 후 반환 (최대 15개 컬럼)
        logger.warning("오류로 인해 원본 데이터에서 일부 행과 컬럼만 선택합니다.")
        
        if not original_df.empty:
            try:
                # 오류 발생 표시 데이터프레임 생성
                error_cols = ["상품명", "처리 상태", "처리 시간"]
                error_df = pd.DataFrame({
                    "상품명": ["오류 발생으로 필터링 실패"],
                    "처리 상태": [f"오류: {str(e)}"],
                    "처리 시간": [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                })
                
                # 원본에서 추가 컬럼 가져오기
                for col in list(original_df.columns[:min(12, len(original_df.columns))]):
                    if col not in error_df.columns:
                        error_df[col] = [""]
                
                return error_df
            except:
                # 최후의 방어: 완전히 실패한 경우 원본에서 최소한 행만 가져오기
                columns_to_keep = list(original_df.columns[:min(15, len(original_df.columns))])
                return original_df[columns_to_keep].head(3).copy()
                
        return pd.DataFrame()  # 완전히 실패한 경우에만 빈 데이터프레임 반환

# --- 전체 프로세스 실행 함수 ---

def run_complete_workflow(input_file: str, output_root_dir: Optional[str] = None, 
                          run_first_stage: bool = True, run_second_stage: bool = True,
                          send_email: bool = False) -> Dict[str, str]:
    """
    전체 워크플로우를 한번에 실행하는 함수
    
    Args:
        input_file: 입력 파일 경로
        output_root_dir: 출력 루트 디렉토리 (기본값: 자동 생성)
        run_first_stage: 1차 처리를 실행할지 여부
        run_second_stage: 2차 처리를 실행할지 여부
        send_email: 이메일 전송 여부
        
    Returns:
        처리 결과를 담은 딕셔너리
    """
    logger.info(f"전체 워크플로우 시작: {input_file}")
    
    # 결과 딕셔너리 초기화
    results = {
        "input_file": input_file,
        "first_stage_file": "",
        "second_stage_file": "",
        "email_sent": False,
        "status": "pending",
        "error": "",
        "start_time": datetime.now().isoformat()
    }
    
    start_time = datetime.now()
    
    # 경로 설정
    if not output_root_dir:
        # 입력 파일과 같은 디렉토리에 output 폴더 생성
        parent_dir = os.path.dirname(input_file)
        output_root_dir = os.path.join(parent_dir, "output")
    
    # 출력 디렉토리 구조 생성
    first_stage_dir = os.path.join(output_root_dir, "intermediate")
    second_stage_dir = os.path.join(output_root_dir, "final")
    os.makedirs(first_stage_dir, exist_ok=True)
    os.makedirs(second_stage_dir, exist_ok=True)
    
    # 입력 파일 존재 확인
    if not os.path.exists(input_file):
        error_msg = f"입력 파일이 존재하지 않음: {input_file}"
        logger.error(error_msg)
        results["status"] = "error"
        results["error"] = error_msg
        print(f"\n[오류] {error_msg}")
        return results
    
    # 1. 1차 파일 처리
    try:
        if run_first_stage:
            print(f"1차 처리 시작: {input_file}")
            logger.info(f"1차 처리 시작: {input_file}")
            
            first_stage_file = process_first_stage(input_file, first_stage_dir)
            
            if not first_stage_file or not os.path.exists(first_stage_file):
                error_msg = f"1차 파일 처리 실패: {input_file}"
                logger.error(error_msg)
                results["status"] = "error"
                results["error"] = error_msg
                
                # 원본 파일을 결과로 설정 (최소한의 결과라도 반환)
                if os.path.exists(input_file):
                    results["first_stage_file"] = input_file
                
                print(f"\n[오류] {error_msg}")
                return results
                
            results["first_stage_file"] = first_stage_file
            logger.info(f"1차 처리 완료: {first_stage_file}")
            print(f"1차 처리 완료: {first_stage_file}")
        else:
            # 1차 처리 생략 시 기존 파일 경로 추정
            base_name = os.path.basename(input_file)
            name, ext = os.path.splitext(base_name)
            first_stage_file = os.path.join(first_stage_dir, f"{name}_intermediate{ext}")
            
            if not os.path.exists(first_stage_file):
                # 파일이 없으면 원본 파일 사용
                logger.warning(f"1차 파일이 존재하지 않음: {first_stage_file}. 원본 파일을 사용합니다.")
                first_stage_file = input_file
                
            results["first_stage_file"] = first_stage_file
    except Exception as e:
        error_msg = f"1차 처리 중 예외 발생: {str(e)}"
        logger.error(error_msg, exc_info=True)
        results["status"] = "error"
        results["error"] = error_msg
        
        # 원본 파일을 결과로 설정
        results["first_stage_file"] = input_file
        
        print(f"\n[오류] {error_msg}")
        return results
    
    # 이메일 전송 (필요시)
    if send_email and results["first_stage_file"]:
        try:
            # 이메일 전송 로직은 utils/preprocessing.py의 send_report_email 함수 호출
            from utils.preprocessing import send_report_email
            email_sent = send_report_email(results["first_stage_file"], recipient_email='dasomas@kakao.com')
            results["email_sent"] = email_sent
            logger.info(f"이메일 전송 결과: {'성공' if email_sent else '실패'}")
            print(f"이메일 전송: {'성공' if email_sent else '실패'}")
        except Exception as e:
            logger.error(f"이메일 전송 중 오류: {str(e)}", exc_info=True)
            results["email_sent"] = False
            print(f"이메일 전송 오류: {str(e)}")
    
    # 2차 파일 처리
    if run_second_stage and results["first_stage_file"]:
        try:
            print(f"2차 처리 시작: {results['first_stage_file']}")
            logger.info(f"2차 처리 시작: {results['first_stage_file']}")
            
            second_stage_file = process_second_stage(results["first_stage_file"], second_stage_dir)
            
            if not second_stage_file or not os.path.exists(second_stage_file):
                error_msg = "2차 파일 처리 실패: final 파일이 생성되지 않았습니다."
                logger.error(error_msg)
                results["status"] = "error"
                results["error"] = error_msg
                print(f"\n[오류] {error_msg}")
                
                # 오류 발생 시 1차 파일을 복사하여 2차 파일로 사용
                try:
                    import shutil
                    base_name = os.path.basename(results["first_stage_file"])
                    name, ext = os.path.splitext(base_name)
                    fallback_file = os.path.join(second_stage_dir, f"{name.replace('_intermediate', '')}_final_fallback{ext}")
                    shutil.copy2(results["first_stage_file"], fallback_file)
                    logger.info(f"대체 2차 파일 생성: {fallback_file}")
                    results["second_stage_file"] = fallback_file
                    print(f"대체 파일 생성됨: {fallback_file}")
                except Exception as copy_error:
                    logger.error(f"대체 파일 생성 실패: {str(copy_error)}")
            else:
                results["second_stage_file"] = second_stage_file
                results["status"] = "success"
                logger.info(f"2차 처리 완료: {second_stage_file}")
                print(f"\n[성공] 최종 보고서가 생성되었습니다: {second_stage_file}")
        except Exception as e:
            error_msg = f"2차 처리 중 예외 발생: {str(e)}"
            logger.error(error_msg, exc_info=True)
            results["status"] = "error" 
            results["error"] = error_msg
            print(f"\n[오류] {error_msg}")
            
            # 오류 발생 시 1차 파일을 결과로 설정
            if not results.get("second_stage_file") and results.get("first_stage_file"):
                try:
                    import shutil
                    base_name = os.path.basename(results["first_stage_file"])
                    name, ext = os.path.splitext(base_name)
                    error_file = os.path.join(second_stage_dir, f"{name.replace('_intermediate', '')}_final_error{ext}")
                    shutil.copy2(results["first_stage_file"], error_file)
                    logger.info(f"오류 발생 시 대체 파일 생성: {error_file}")
                    results["second_stage_file"] = error_file
                    print(f"오류 발생 시 대체 파일 생성됨: {error_file}")
                except Exception as copy_error:
                    logger.error(f"오류 대체 파일 생성 실패: {str(copy_error)}")
    elif not run_second_stage:
        # 2차 처리를 건너뛰고 성공 상태로 표시
        results["status"] = "success"
        logger.info("2차 처리 건너뜀 (run_second_stage=False)")
    
    # 처리 시간 기록
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    results["end_time"] = end_time.isoformat()
    results["duration_seconds"] = duration
    
    # 최종 상태가 설정되지 않은 경우 성공으로 설정
    if results["status"] == "pending":
        results["status"] = "success"
    
    logger.info(f"전체 워크플로우 완료: {duration:.2f}초 소요, 상태: {results['status']}")
    
    # 상태에 따른 최종 메시지 출력
    if results["status"] == "success":
        success_files = []
        if results["first_stage_file"]:
            success_files.append(f"1차: {results['first_stage_file']}")
        if results["second_stage_file"]:
            success_files.append(f"2차: {results['second_stage_file']}")
        
        print(f"\n[처리 완료] {', '.join(success_files)}")
    else:
        print(f"\n[처리 실패] {results['error']}")
    
    return results

# 직접 실행 시 테스트 코드
if __name__ == "__main__":
    import sys
    
    # 로깅 설정
    logging.basicConfig(level=logging.INFO, 
                        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        
        if len(sys.argv) > 2 and sys.argv[2] == "second_only":
            # 2차 처리만 실행
            process_second_stage(input_file)
        else:
            # 전체 워크플로우 실행
            results = run_complete_workflow(input_file)
            print(f"처리 결과: {results['status']}")
            if results['first_stage_file']:
                print(f"1차 파일: {results['first_stage_file']}")
            if results['second_stage_file']:
                print(f"2차 파일: {results['second_stage_file']}")
    else:
        print("사용법: python excel_processor.py <입력파일경로> [second_only]")

# 비동기 함수 실행을 위한 래퍼 함수
def run_async_func(func, *args, **kwargs):
    """
    비동기 함수를 동기적으로 실행하는 래퍼 함수
    
    Args:
        func: 실행할 비동기 함수
        *args, **kwargs: 함수에 전달할 인수들
        
    Returns:
        함수의 결과값
    """
    try:
        # nest_asyncio로 중첩된 이벤트 루프 지원
        try:
            nest_asyncio.apply()
        except:
            pass
            
        # 이벤트 루프 가져오기
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            # 이벤트 루프가 닫혀있거나 없는 경우 새로 생성
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
        # 비동기 함수 실행
        if loop.is_running():
            # 이미 실행 중인 경우 run_until_complete 대신 다른 방법 사용
            return asyncio.run_coroutine_threadsafe(func(*args, **kwargs), loop).result()
        else:
            return loop.run_until_complete(func(*args, **kwargs))
    except Exception as e:
        logging.error(f"비동기 함수 실행 중 오류: {e}")
        # 빈 리스트 반환 (검색 결과 없음을 의미)
        return [] 