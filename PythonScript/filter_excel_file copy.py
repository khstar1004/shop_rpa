import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from urllib.parse import urlparse
from datetime import datetime

def filter_excel_file():
    input_directory = 'C:\\RPA\\Input'
    target_directory = 'C:\\RPA\\Output'
    
    xlsx_files = [f for f in os.listdir(input_directory) if f.lower().endswith('.xlsx')]
    
    if not xlsx_files:
        print("xlsx 파일이 없습니다.")
        return
    
    file_path = os.path.join(input_directory, xlsx_files[0])
    df = pd.read_excel(file_path)
    
    filtered_df = df[(df['가격차이(2)'] < 0) | (df['가격차이(3)'] < 0)]

    # 원하는 컬럼만 추출
    filtered_df = filtered_df[['날짜', '담당자', '업체명', '업체코드', 'Code', '중분류카테고리', '상품명', '기본수량(1)', '판매단가(V포함)', '본사상품링크', '기본수량(2)', '판매단가(V포함)(2)', '가격차이(2)', '고려기프트 상품링크', '기본수량(3)', '판매단가(V포함)(3)', '가격차이(3)', '가격차이 비율(3)', '공급사명', '공급사 상품링크', '본사 이미지', '고려기프트 이미지', '네이버 이미지']]

    # 컬럼 이름 변경
    filtered_df.rename(columns={
        '날짜': '구분(승인관리:A/가격관리:P)',
        '담당자': '담당자',
        '업체명': '공급사명',
        '업체코드': '공급처코드',
        'Code': '상품코드',
        '중분류카테고리': '카테고리(중분류)',
        '상품명': '상품명',
        '기본수량(1)': '본사 기본수량',
        '판매단가(V포함)': '판매단가1(VAT포함)',
        '본사상품링크': '본사링크',
        '기본수량(2)': '고려 기본수량',
        '판매단가(V포함)(2)': '판매단가2(VAT포함)',
        '가격차이(2)': '고려 가격차이',
        '고려기프트 상품링크': '고려 링크',
        '기본수량(3)': '네이버 기본수량',
        '판매단가(V포함)(3)': '판매단가3 (VAT포함)',
        '가격차이(3)': '네이버 가격차이',
        '가격차이 비율(3)': '네이버가격차이(%)',
        '공급사명': '네이버 공급사명',
        '공급사 상품링크': '네이버 링크',
        '본사 이미지': '해오름(이미지링크)',
        '고려기프트 이미지': '고려기프트(이미지링크)',
        '네이버 이미지': '네이버쇼핑(이미지링크)'
    }, inplace=True)

    current_datetime = datetime.now().strftime('%Y%m%d%H%M%S')
    output_filename = f"UploadExcelFile_{current_datetime}.xlsx"

    output_path = os.path.join(target_directory, output_filename)
    filtered_df.to_excel(output_path, index=False)
    
    # 여기서부터 openpyxl을 사용해서 추가 작업
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 스타일 설정
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_col=23, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

    for col in ws.iter_cols(min_col=1, max_col=23):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 너비와 높이 설정
    for col in ws.iter_cols(min_col=1, max_col=23, max_row=1):
        for cell in col:
            ws.column_dimensions[cell.column_letter].width = 15
            
    for row in ws.iter_rows(min_row=2, max_col=23, max_row=ws.max_row):
        for cell in row:
            ws.row_dimensions[cell.row].height = 100

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # "고려 링크"와 "네이버 링크" 체크
        goleo_link_cell = row[13]  # "고려 링크"는 14번째 컬럼에 위치
        naver_link_cell = row[19]  # "네이버 링크"는 19번째 컬럼에 위치

        if goleo_link_cell.value and not bool(urlparse(goleo_link_cell.value).scheme):
            goleo_link_cell.value = None

        if naver_link_cell.value and not bool(urlparse(naver_link_cell.value).scheme):
            naver_link_cell.value = None

        # 가격차이 체크
        goleo_difference_cell = row[12]  # "고려 가격차이"는 13번째 컬럼에 위치
        naver_difference_cell = row[16]  # "네이버 가격차이"는 17번째 컬럼에 위치

        if goleo_difference_cell.value is not None and goleo_difference_cell.value >= 0:
            for i in [10, 11, 12, 13, 21]:
                row[i].value = None

        if naver_difference_cell.value is not None and naver_difference_cell.value >= 0:
            for i in [14, 15, 16, 17, 18, 19, 22]:
                row[i].value = None

        # '네이버 기본수량' 체크
        naver_quantity_cell = row[14]  # '네이버 기본수량'는 15번째 컬럼에 위치
        naver_difference_percentage_cell = row[17]  # '네이버가격차이(%)'는 18번째 컬럼에 위치

        if (not naver_quantity_cell.value or str(naver_quantity_cell.value).strip() == "") and (naver_difference_percentage_cell.value is not None and naver_difference_percentage_cell.value > -10.0):
            for i in [14, 15, 16, 17, 18, 19, 22]:
                row[i].value = None

        # '고려 기본수량' 체크
        goleo_quantity_cell = row[10]  # '고려 기본수량'은 11번째 컬럼에 위치
        if not (isinstance(goleo_quantity_cell.value, (int, float))):
            goleo_quantity_cell.value = None

    # 특정 컬럼들의 값이 모두 비어 있는지 확인하고, 그럴 경우 해당 행을 삭제
    for row in reversed(list(ws.iter_rows(min_row=2, max_row=ws.max_row))):
        columns_to_check = [10, 11, 12, 13, 14, 15, 16, 17, 18, 19]  # 위에서 제시된 컬럼 인덱스들

        # 해당 행의 컬럼 값들이 모두 None이거나 빈 문자열인지 확인
        if all([(cell.value is None or str(cell.value).strip() == "") for cell in [row[i] for i in columns_to_check]]):
            ws.delete_rows(row[0].row)


    
    wb.save(output_path)
