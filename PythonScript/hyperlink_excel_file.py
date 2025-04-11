import pandas as pd
import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Border, Side

def generate_output_filename(input_filename):
    now = datetime.datetime.now()
    timestamp = now.strftime('%Y%m%d_%H%M%S')
    base_name = os.path.splitext(input_filename)[0]
    output_filename = f"{base_name}_result_{timestamp}.xlsx"
    return output_filename

def hyperlink_excel_file():
    input_directory = 'C:\\RPA\\Input'  # 입력 디렉토리 경로
    target_directory = 'C:\\RPA\\Output'  # 저장할 디렉토리 경로
    
    xlsx_files = [f for f in os.listdir(input_directory) if f.lower().endswith('.xlsx')]
    
    if not xlsx_files:
        print("xlsx 파일이 없습니다.")
        return
    
    file_path = os.path.join(input_directory, xlsx_files[0])
    
    # xlsx 파일에서 데이터프레임 읽기
    df = pd.read_excel(file_path)

    # 워크북 불러오기
    wb = load_workbook(file_path)
    ws = wb.active

    # 테두리 스타일 정의
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # A~AH열의 너비 설정 및 자동 줄 바꿈, 테두리 적용
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
        ws.column_dimensions[col].width = 16
        
        for row in range(1, df.shape[0] + 2):  # 첫 번째 행은 헤더이므로 1을 더해준다.
            cell = ws[f"{col}{row}"]
            cell.alignment = Alignment(wrap_text=True)
            cell.border = thin_border  # 테두리 적용
    
    link_columns = ["본사상품링크", "고려기프트 상품링크", "네이버 쇼핑 링크", "공급사 상품링크"]
    for link_column in link_columns:
        for row_idx, link in enumerate(df[link_column], start=2):  # 첫 번째 행은 헤더이므로 인덱스 2부터 시작
            if pd.notna(link) and (link.startswith("http") or link.startswith("https")):  # 링크가 NaN이 아니며 http 또는 https로 시작할 때만 처리
                cell = ws.cell(row=row_idx, column=df.columns.get_loc(link_column) + 1)  # 열 인덱스 찾기
                cell.value = link  # 셀 값 설정
                cell.hyperlink = link  # 하이퍼링크 설정

    # 색 채우기 위한 PatternFill 객체 생성
    gray_fill = PatternFill(start_color="CCCCCC",
                        end_color="CCCCCC",
                        fill_type="solid")

    yellow_fill = PatternFill(start_color="FFFF00",
                           end_color="FFFF00",
                           fill_type="solid")

    for cell in ws["1:1"]:
        cell.fill = gray_fill

    # for row_idx in range(2, df.shape[0] + 2):  # 첫 번째 행은 헤더이므로 인덱스 2부터 시작
    #     for col_name in ["가격차이(2)", "가격차이(3)"]:
    #         value = df.loc[row_idx - 2, col_name]  # Pandas DataFrame은 0-based indexing
    #         if pd.notna(value) and value < 0:
    #             for col in ws.iter_cols(min_row=row_idx, max_row=row_idx, min_col=1, max_col=ws.max_column):
    #                 for cell in col:
    #                     cell.fill = yellow_fill
    #             break

    for row_idx in range(2, df.shape[0] + 2):  # 첫 번째 행은 헤더이므로 인덱스 2부터 시작
        for col_name in ["가격차이(2)", "가격차이(3)"]:
            value = df.loc[row_idx - 2, col_name]  # Pandas DataFrame은 0-based indexing
            try:
                value = float(value)  # 문자열을 숫자로 변환 시도
                if value < 0:
                    for col in ws.iter_cols(min_row=row_idx, max_row=row_idx, min_col=1, max_col=ws.max_column):
                        for cell in col:
                            cell.fill = yellow_fill
                    break
            except ValueError:
                continue

    # 데이터프레임을 xlsx 파일로 저장
    output_filename = generate_output_filename(xlsx_files[0])
    output_file_path = os.path.join(target_directory, output_filename)
    wb.save(output_file_path)
    print(f"데이터가 {output_file_path}에 저장되었습니다.")


if __name__ == "__main__":
    hyperlink_excel_file()