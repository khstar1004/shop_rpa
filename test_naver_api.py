import os
import sys
import logging
from datetime import datetime
from openpyxl import load_workbook

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger("naver_api_test")

# 프로젝트 루트 경로 추가
root_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(root_dir)

try:
    from core.scraping.naver_crawler import NaverShoppingCrawler
    from utils.caching import FileCache
    from utils.config import load_config
except ImportError as e:
    logger.error(f"필요한 모듈을 가져올 수 없습니다: {e}")
    sys.exit(1)

def check_image_urls_in_excel(file_path):
    """엑셀 파일에서 이미지 URL을 확인"""
    logger.info(f"엑셀 파일 이미지 URL 확인: {file_path}")
    
    if not os.path.exists(file_path):
        logger.error(f"파일이 존재하지 않습니다: {file_path}")
        return
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    image_cells = []
    
    # 모든 셀 검사
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('=IMAGE('):
                image_cells.append((cell.row, cell.column, cell.value))
    
    # 결과 출력
    if image_cells:
        logger.info(f"총 {len(image_cells)}개의 이미지 셀 발견:")
        for row, col, value in image_cells[:5]:  # 처음 5개만 표시
            logger.info(f"  행 {row}, 열 {col}: {value[:50]}...")
        if len(image_cells) > 5:
            logger.info(f"  ... 외 {len(image_cells) - 5}개")
    else:
        logger.warning("이미지 셀을 찾을 수 없습니다!")

def main():
    """네이버 쇼핑 API 테스트"""
    logger.info("=== 네이버 쇼핑 API 테스트 시작 ===")
    
    # 설정 로드
    config = load_config()
    
    # 캐시 초기화
    cache = FileCache(
        cache_dir=config["PATHS"]["CACHE_DIR"],
        duration_seconds=config["PROCESSING"]["CACHE_DURATION"],
        max_size_mb=config["PROCESSING"].get("CACHE_MAX_SIZE_MB", 1024)
    )
    
    # 네이버 크롤러 초기화
    naver_crawler = NaverShoppingCrawler(
        max_retries=config["PROCESSING"]["MAX_RETRIES"],
        cache=cache,
        timeout=config["PROCESSING"].get("REQUEST_TIMEOUT", 30)
    )
    
    # 테스트할 검색어 목록
    test_queries = [
        "도자기 머그컵",
        "사무용 볼펜 세트",
        "usb_메모리",
        "티셔츠"
    ]
    
    # 각 검색어에 대해 테스트
    for query in test_queries:
        logger.info(f"\n검색어 테스트: '{query}'")
        
        # 네이버에서 제품 검색
        products = naver_crawler.search_product(query, max_items=5)
        
        if not products:
            logger.warning(f"'{query}' 검색 결과 없음")
            continue
            
        logger.info(f"검색 결과: {len(products)}개 제품")
        
        # 첫 번째 제품의 상세 정보 출력
        if products and hasattr(products[0], 'name'):
            product = products[0]
            logger.info(f"첫 번째 제품: {product.name}")
            logger.info(f"가격: {product.price}")
            logger.info(f"브랜드: {product.brand}")
            logger.info(f"이미지 URL: {product.image_url[:70]}..." if product.image_url else "이미지 없음")
            logger.info(f"제품 URL: {product.url[:70]}..." if product.url else "URL 없음")
    
    # 이미지를 엑셀에 저장하는 테스트
    output_dir = os.path.join(root_dir, "output")
    os.makedirs(output_dir, exist_ok=True)
    
    # 테스트에 사용할 검색어
    test_query = "도자기 머그컵"
    logger.info(f"\n엑셀 저장 테스트: '{test_query}'")
    
    # 네이버에서 제품 검색
    products = naver_crawler.search_product(test_query, max_items=5)
    
    # 결과가 있으면 엑셀로 저장
    if products:
        from core.processing.excel_manager import ExcelManager
        
        # Excel 매니저 초기화
        excel_manager = ExcelManager(config, logger)
        
        # 출력 파일 설정
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"naver_api_test_{timestamp}.xlsx")
        
        # 스크래핑 결과를 엑셀로 저장
        logger.info(f"제품 {len(products)}개를 엑셀로 저장합니다.")
        try:
            sheet_name = f"naver_test_{timestamp[8:14]}"
            result_file = excel_manager.save_products(
                products, 
                output_file, 
                sheet_name
            )
            logger.info(f"엑셀 파일 저장 완료: {result_file}")
            
            # 저장된 엑셀 파일에서 이미지 URL 확인
            check_image_urls_in_excel(result_file)
            
        except Exception as e:
            logger.error(f"엑셀 저장 중 오류 발생: {e}", exc_info=True)
    else:
        logger.warning(f"'{test_query}' 검색 결과 없음 - 엑셀 저장 테스트 건너뜀")
    
    logger.info("=== 테스트 완료 ===")

if __name__ == "__main__":
    main() 